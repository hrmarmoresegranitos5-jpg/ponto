"""
╔══════════════════════════════════════════════════════════════════╗
║  RELATÓRIOS PDF / EXCEL — Sistema de Ponto v5.0                ║
║  Etapa 5 · Exportação de relatórios em PDF e Excel             ║
╚══════════════════════════════════════════════════════════════════╝

Uso independente:
    from relatorios_pdf import gerar_pdf_calculo, gerar_pdf_banco_horas
    from relatorios_pdf import gerar_excel_calculo, gerar_excel_banco_horas

Requisitos:
    pip install reportlab openpyxl
"""

import os
from datetime import datetime

# ── Cores da paleta do sistema (versão clara para PDF) ──────────────────────
PDF_ACCENT   = (0,   194, 255)   # #00C2FF  → azul destaque
PDF_SUCCESS  = (0,   229, 160)   # #00E5A0  → verde
PDF_DANGER   = (255, 77,  106)   # #FF4D6A  → vermelho
PDF_WARNING  = (255, 184, 48)    # #FFB830  → amarelo
PDF_DARK     = (13,  15,  20)    # #0D0F14  → fundo escuro
PDF_MID      = (138, 144, 168)   # #8A90A8  → texto médio

# ── Helpers de formatação ───────────────────────────────────────────────────
def _fmt_mins(mins: int) -> str:
    """Minutos → 'HH:MM'"""
    m = abs(mins)
    return f"{m // 60:02d}:{m % 60:02d}"

def _saldo_fmt(mins: int) -> str:
    return ("-" if mins < 0 else "+") + _fmt_mins(mins)

def _fd(d: str) -> str:
    """'YYYY-MM-DD' → 'DD/MM/YYYY'"""
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return d

def _now_str() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M")

# ═══════════════════════════════════════════════════════════════════════════════
# PDF — usa ReportLab
# ═══════════════════════════════════════════════════════════════════════════════

def _check_reportlab():
    try:
        import reportlab
    except ImportError:
        raise ImportError(
            "ReportLab não instalado.\n"
            "Execute:  pip install reportlab"
        )

def _rgb(t):
    """Tupla (R,G,B) 0-255 → Color do ReportLab"""
    from reportlab.lib.colors import Color
    return Color(t[0]/255, t[1]/255, t[2]/255)


def _cabecalho_pdf(canvas, doc, titulo: str, subtitulo: str = ""):
    """Desenha cabeçalho padrão em cada página."""
    from reportlab.lib.units import mm
    w, h = doc.pagesize
    # Barra superior escura
    canvas.setFillColor(_rgb(PDF_DARK))
    canvas.rect(0, h - 22*mm, w, 22*mm, fill=1, stroke=0)
    # Linha accent
    canvas.setFillColor(_rgb(PDF_ACCENT))
    canvas.rect(0, h - 23*mm, w, 1.2*mm, fill=1, stroke=0)
    # Título
    canvas.setFillColor(_rgb((234, 237, 245)))
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawString(18*mm, h - 14*mm, titulo)
    # Subtítulo
    if subtitulo:
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(_rgb(PDF_MID))
        canvas.drawString(18*mm, h - 19*mm, subtitulo)
    # Data de geração (direita)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(_rgb(PDF_MID))
    canvas.drawRightString(w - 18*mm, h - 14*mm, f"Gerado em: {_now_str()}")
    # Número de página
    canvas.drawRightString(w - 18*mm, h - 19*mm, f"Pág. {doc.page}")


def gerar_pdf_calculo(dias_calc: list, caminho: str, filtros: dict = None) -> str:
    """
    Gera PDF do relatório de Cálculo de Horas.

    Parâmetros:
        dias_calc  : lista de dicts retornados por CalcEngine.calc_dia()
                     cada dict deve ter: _func_nome, data, dia_nome, horarios,
                     trabalhados_fmt, extras_fmt, faltas_fmt, saldo_fmt,
                     saldo_min, extras_min, faltas_min, alertas
        caminho    : caminho de saída do arquivo .pdf
        filtros    : dict opcional com {'funcionario': str, 'importacao': str}

    Retorna:
        caminho do arquivo gerado
    """
    _check_reportlab()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors

    W, H = A4

    # ── Estilos de parágrafo ─────────────────────────────────────────────────
    st_titulo = ParagraphStyle("titulo",
        fontSize=11, fontName="Helvetica-Bold",
        textColor=_rgb((234, 237, 245)),
        spaceAfter=4)
    st_sub = ParagraphStyle("sub",
        fontSize=8, fontName="Helvetica",
        textColor=_rgb(PDF_MID),
        spaceAfter=12)
    st_normal = ParagraphStyle("normal",
        fontSize=9, fontName="Helvetica",
        textColor=_rgb((138, 144, 168)))

    # ── Sumário calculado ────────────────────────────────────────────────────
    total_trabalhados = sum(d.get("trabalhados_min", 0) for d in dias_calc)
    total_extras      = sum(d.get("extras_min", 0)      for d in dias_calc)
    total_faltas      = sum(d.get("faltas_min", 0)      for d in dias_calc)
    saldo_total       = total_extras - total_faltas
    dias_com_extra    = sum(1 for d in dias_calc if d.get("extras_min", 0) > 0)
    dias_com_falta    = sum(1 for d in dias_calc if d.get("faltas_min", 0) > 0)
    dias_com_alerta   = sum(1 for d in dias_calc if d.get("alertas"))

    # ── Montagem do documento ────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        caminho, pagesize=A4,
        topMargin=28*mm, bottomMargin=18*mm,
        leftMargin=18*mm, rightMargin=18*mm,
        title="Relatório de Cálculo de Horas",
        author="Sistema de Ponto v5.0",
    )

    story = []

    # Cabeçalho de conteúdo
    sub_txt = ""
    if filtros:
        partes = []
        if filtros.get("funcionario"):  partes.append(f"Funcionário: {filtros['funcionario']}")
        if filtros.get("importacao"):   partes.append(f"Importação: {filtros['importacao']}")
        sub_txt = "  |  ".join(partes)
    story.append(Paragraph("Relatório de Cálculo de Horas", st_titulo))
    if sub_txt:
        story.append(Paragraph(sub_txt, st_sub))
    story.append(Spacer(1, 4*mm))

    # Cards de sumário — tabela 2 linhas x 4 colunas
    W_avail = W - 36*mm
    W_card  = W_avail / 4
    col_ws_cards = [W_card] * 4
    resumo_labels = ["TRABALHADO TOTAL", "HORAS EXTRAS", "HORAS FALTAS", "SALDO PERIODO"]
    resumo_vals   = [_fmt_mins(total_trabalhados), _fmt_mins(total_extras),
                     _fmt_mins(total_faltas), _saldo_fmt(saldo_total)]
    resumo_cores  = [PDF_ACCENT, PDF_SUCCESS, PDF_DANGER,
                     PDF_SUCCESS if saldo_total >= 0 else PDF_DANGER]
    cards_tbl = Table([
        [Paragraph(f'<font color="#{_hex(PDF_MID)}" size="7">{lbl}</font>', st_normal)
         for lbl in resumo_labels],
        [Paragraph(f'<font color="#{_hex(cor)}" size="14"><b>{val}</b></font>', st_normal)
         for val, cor in zip(resumo_vals, resumo_cores)],
    ], colWidths=col_ws_cards)
    cards_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _rgb((26, 30, 42))),
        ("GRID",       (0, 0), (-1, -1), 0.5, _rgb((30, 34, 53))),
        ("TOPPADDING",    (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING",    (0, 1), (-1, 1), 2),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("LINEABOVE", (0, 0), (-1, 0), 2, _rgb(PDF_ACCENT)),
    ]))
    story.append(cards_tbl)
    story.append(Spacer(1, 6*mm))

    # Linha info adicional
    info_txt = (f"Total de registros: <b>{len(dias_calc)}</b>  |  "
                f"Dias com extra: <b>{dias_com_extra}</b>  |  "
                f"Dias com falta: <b>{dias_com_falta}</b>  |  "
                f"Dias com alerta: <b>{dias_com_alerta}</b>")
    story.append(Paragraph(f'<font color="#{_hex(PDF_MID)}" size="8">{info_txt}</font>', st_normal))
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_rgb((30, 34, 53))))
    story.append(Spacer(1, 4*mm))

    # Tabela principal de dias
    # Larguras ajustadas para A4 (usável = 493mm, 36mm margens)
    colunas = ["FUNCIONARIO", "DATA", "DIA", "HORARIOS", "TRABALHADO", "EXTRAS", "FALTAS", "SALDO", "ALERTAS"]
    col_w   = [32*mm, 18*mm, 9*mm, 38*mm, 20*mm, 17*mm, 15*mm, 17*mm, 0]
    col_w[-1] = max(10*mm, (W - 36*mm) - sum(col_w[:-1]))

    def _cel(txt, cor=None, bold=False):
        c = f"#{_hex(cor)}" if cor else f"#{_hex(PDF_MID)}"
        b = "<b>" if bold else ""
        be = "</b>" if bold else ""
        return Paragraph(f'<font color="{c}" size="8">{b}{txt}{be}</font>', st_normal)

    linhas = [colunas]  # cabeçalho como texto simples

    for d in dias_calc:
        fn     = str(d.get("_func_nome", ""))
        data   = _fd(str(d.get("data", "")))
        dia    = str(d.get("dia_nome", ""))[:3]
        horas  = "  ".join(str(h) for h in d.get("horarios", []))
        trab   = str(d.get("trabalhados_fmt", "--:--"))
        extra  = str(d.get("extras_fmt", "--:--"))
        falta  = str(d.get("faltas_fmt", "--:--"))
        saldo  = str(d.get("saldo_fmt", "--:--"))
        saldo_m= d.get("saldo_min", 0)
        alerta = " · ".join(d.get("alertas", []))

        cor_saldo = PDF_SUCCESS if saldo_m >= 0 else PDF_DANGER
        cor_extra = PDF_SUCCESS if d.get("extras_min", 0) > 0 else PDF_MID
        cor_falta = PDF_DANGER  if d.get("faltas_min", 0) > 0 else PDF_MID

        linhas.append([
            _cel(fn[:28]),
            _cel(data),
            _cel(dia),
            _cel(horas[:38]),
            _cel(trab, PDF_ACCENT),
            _cel(extra, cor_extra),
            _cel(falta, cor_falta),
            _cel(saldo, cor_saldo, bold=True),
            _cel(alerta[:35], PDF_WARNING if alerta else PDF_MID),
        ])

    tbl = Table(linhas, colWidths=col_w, repeatRows=1)
    estilo = TableStyle([
        # Cabeçalho
        ("BACKGROUND", (0, 0), (-1, 0), _rgb((10, 12, 18))),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _rgb(PDF_ACCENT)),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 7.5),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        # Dados
        ("BACKGROUND", (0, 1), (-1, -1), _rgb((26, 30, 42))),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("GRID",   (0, 0), (-1, -1), 0.3, _rgb((30, 34, 53))),
        ("LINEABOVE", (0, 1), (-1, 1), 1, _rgb(PDF_ACCENT)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [_rgb((26, 30, 42)), _rgb((22, 25, 36))]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])
    tbl.setStyle(estilo)
    story.append(tbl)

    # ── Build com cabeçalho em cada página ───────────────────────────────────
    def _on_page(canv, d):
        sub = f"Sistema de Ponto v5.0  |  {sub_txt}" if sub_txt else "Sistema de Ponto v5.0"
        _cabecalho_pdf(canv, d, "Cálculo de Horas", sub)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return caminho


def gerar_pdf_banco_horas(funcionarios_saldo: list, caminho: str) -> str:
    """
    Gera PDF do Banco de Horas (saldo de todos os funcionários).

    Parâmetros:
        funcionarios_saldo : lista de dicts com:
            { nome, saldo_min, atualizado_em,
              historico: [{tipo, valor_fmt, descricao, data}] }
        caminho : caminho de saída .pdf

    Retorna:
        caminho do arquivo gerado
    """
    _check_reportlab()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable, KeepTogether)
    from reportlab.lib.styles import ParagraphStyle

    doc = SimpleDocTemplate(
        caminho, pagesize=A4,
        topMargin=28*mm, bottomMargin=18*mm,
        leftMargin=18*mm, rightMargin=18*mm,
        title="Banco de Horas",
        author="Sistema de Ponto v5.0",
    )
    W, H = A4

    st_normal = ParagraphStyle("n", fontSize=9, fontName="Helvetica",
                                textColor=_rgb(PDF_MID))
    st_titulo = ParagraphStyle("t", fontSize=11, fontName="Helvetica-Bold",
                                textColor=_rgb((234, 237, 245)), spaceAfter=4)

    def _p(txt, cor=None, size=9, bold=False):
        c = f"#{_hex(cor)}" if cor else f"#{_hex(PDF_MID)}"
        b = "<b>" if bold else ""
        be = "</b>" if bold else ""
        return Paragraph(f'<font color="{c}" size="{size}">{b}{txt}{be}</font>', st_normal)

    story = []
    story.append(Paragraph("Banco de Horas — Saldo Geral", st_titulo))
    story.append(_p(f"Gerado em: {_now_str()}", PDF_MID, size=8))
    story.append(Spacer(1, 6*mm))

    # Tabela de resumo (todos os funcs)
    resumo_data = [["FUNCIONÁRIO", "SALDO", "ATUALIZADO EM"]]
    for f in funcionarios_saldo:
        saldo_m = f.get("saldo_min", 0)
        cor     = PDF_SUCCESS if saldo_m >= 0 else PDF_DANGER
        resumo_data.append([
            _p(str(f.get("nome", "")), (234, 237, 245), bold=True),
            _p(_saldo_fmt(saldo_m), cor, bold=True),
            _p(str(f.get("atualizado_em", ""))[:16] or "—", PDF_MID),
        ])

    tbl_res = Table(resumo_data, colWidths=[80*mm, 40*mm, 50*mm])
    tbl_res.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _rgb((10, 12, 18))),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _rgb(PDF_ACCENT)),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), _rgb((26, 30, 42))),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [_rgb((26, 30, 42)), _rgb((22, 25, 36))]),
        ("GRID",  (0, 0), (-1, -1), 0.3, _rgb((30, 34, 53))),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("LINEABOVE", (0, 1), (-1, 1), 1.5, _rgb(PDF_ACCENT)),
    ]))
    story.append(tbl_res)
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_rgb((30, 34, 53))))
    story.append(Spacer(1, 6*mm))

    # Detalhe por funcionário (histórico)
    story.append(_p("HISTÓRICO DE MOVIMENTAÇÕES", PDF_ACCENT, size=8, bold=True))
    story.append(Spacer(1, 4*mm))

    for f in funcionarios_saldo:
        hist = f.get("historico", [])
        if not hist:
            continue

        saldo_m = f.get("saldo_min", 0)
        cor_saldo = PDF_SUCCESS if saldo_m >= 0 else PDF_DANGER

        bloco = []
        # Cabeçalho do func
        bloco.append(Table([[
            _p(str(f.get("nome", "")), (234, 237, 245), size=10, bold=True),
            _p(f"Saldo: {_saldo_fmt(saldo_m)}", cor_saldo, size=10, bold=True),
        ]], colWidths=[100*mm, 70*mm]))

        bloco.append(Spacer(1, 2*mm))

        # Tabela do histórico
        hist_data = [["DATA", "TIPO", "VALOR", "DESCRIÇÃO"]]
        for m in hist:
            cor_tipo = PDF_SUCCESS if m.get("tipo") == "credito" else PDF_DANGER
            hist_data.append([
                _p(str(m.get("data", ""))[:16], PDF_MID, size=8),
                _p("Crédito" if m.get("tipo") == "credito" else "Débito", cor_tipo, size=8),
                _p(str(m.get("valor_fmt", "")), cor_tipo, size=8, bold=True),
                _p(str(m.get("descricao", ""))[:55], PDF_MID, size=8),
            ])

        tbl_h = Table(hist_data, colWidths=[28*mm, 22*mm, 22*mm, 100*mm])
        tbl_h.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _rgb((15, 18, 25))),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _rgb(PDF_MID)),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 7.5),
            ("BACKGROUND", (0, 1), (-1, -1), _rgb((22, 25, 36))),
            ("GRID", (0, 0), (-1, -1), 0.3, _rgb((30, 34, 53))),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ]))
        bloco.append(tbl_h)
        bloco.append(Spacer(1, 6*mm))

        story.append(KeepTogether(bloco))

    def _on_page(canv, d):
        _cabecalho_pdf(canv, d, "Banco de Horas", "Sistema de Ponto v5.0")

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — usa openpyxl
# ═══════════════════════════════════════════════════════════════════════════════

def _check_openpyxl():
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl não instalado.\n"
            "Execute:  pip install openpyxl"
        )


def _hex(t) -> str:
    """Tupla RGB → hex sem #"""
    return f"{t[0]:02X}{t[1]:02X}{t[2]:02X}"


def gerar_excel_calculo(dias_calc: list, caminho: str, filtros: dict = None) -> str:
    """
    Gera planilha Excel do Cálculo de Horas.

    Retorna: caminho do arquivo gerado
    """
    _check_openpyxl()

    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cálculo de Horas"

    # ── Paleta de fills ───────────────────────────────────────────────────────
    fill_header  = PatternFill("solid", fgColor="0D0F14")
    fill_row1    = PatternFill("solid", fgColor="1A1E2A")
    fill_row2    = PatternFill("solid", fgColor="161929")
    fill_resumo  = PatternFill("solid", fgColor="0A0C12")

    font_title   = Font(name="Segoe UI", size=14, bold=True, color="EAEDF5")
    font_sub     = Font(name="Segoe UI", size=9,  color="8A90A8")
    font_header  = Font(name="Segoe UI", size=9,  bold=True, color="00C2FF")
    font_normal  = Font(name="Segoe UI", size=9,  color="8A90A8")
    font_white   = Font(name="Segoe UI", size=9,  bold=True, color="EAEDF5")
    font_accent  = Font(name="Segoe UI", size=9,  color="00C2FF")
    font_success = Font(name="Segoe UI", size=9,  bold=True, color="00E5A0")
    font_danger  = Font(name="Segoe UI", size=9,  bold=True, color="FF4D6A")
    font_warning = Font(name="Segoe UI", size=9,  color="FFB830")

    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center")
    al_r = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin", color="1E2235")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Linha 1: título ───────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = "Sistema de Ponto — Relatório de Cálculo de Horas"
    ws["A1"].font   = font_title
    ws["A1"].fill   = fill_header
    ws["A1"].alignment = al_l
    ws.row_dimensions[1].height = 28

    # ── Linha 2: filtros / data ───────────────────────────────────────────────
    sub = ""
    if filtros:
        partes = []
        if filtros.get("funcionario"): partes.append(f"Funcionário: {filtros['funcionario']}")
        if filtros.get("importacao"):  partes.append(f"Importação: {filtros['importacao']}")
        sub = "  |  ".join(partes) + "  |  " if partes else ""
    ws.merge_cells("A2:G2")
    ws["A2"] = sub + f"Gerado em: {_now_str()}"
    ws["A2"].font = font_sub
    ws["A2"].fill = fill_header
    ws["A2"].alignment = al_l
    ws.merge_cells("H2:I2")
    ws["H2"] = f"v5.0"
    ws["H2"].font = font_sub
    ws["H2"].fill = fill_header
    ws["H2"].alignment = al_r
    ws.row_dimensions[2].height = 16

    # ── Linha 3: sumário ──────────────────────────────────────────────────────
    total_trabalhados = sum(d.get("trabalhados_min", 0) for d in dias_calc)
    total_extras      = sum(d.get("extras_min", 0)      for d in dias_calc)
    total_faltas      = sum(d.get("faltas_min", 0)      for d in dias_calc)
    saldo_total       = total_extras - total_faltas

    ws.row_dimensions[3].height = 4  # espaço
    ws.row_dimensions[4].height = 20

    resumo_labels = ["TRABALHADO TOTAL", "HORAS EXTRAS", "HORAS FALTAS", "SALDO PERÍODO"]
    resumo_vals   = [_fmt_mins(total_trabalhados), _fmt_mins(total_extras),
                     _fmt_mins(total_faltas), _saldo_fmt(saldo_total)]
    resumo_cores  = ["00C2FF", "00E5A0", "FF4D6A",
                     "00E5A0" if saldo_total >= 0 else "FF4D6A"]

    for i, (lbl, val, cor) in enumerate(zip(resumo_labels, resumo_vals, resumo_cores)):
        col_a = i*2 + 1   # A, C, E, G
        col_b = col_a + 1
        cl_a = get_column_letter(col_a)
        cl_b = get_column_letter(col_b)
        ws.merge_cells(f"{cl_a}4:{cl_b}4")
        ws[f"{cl_a}4"] = lbl
        ws[f"{cl_a}4"].font = Font(name="Segoe UI", size=7, color="8A90A8")
        ws[f"{cl_a}4"].fill = fill_resumo
        ws[f"{cl_a}4"].alignment = al_l
        ws.row_dimensions[5].height = 22
        ws.merge_cells(f"{cl_a}5:{cl_b}5")
        ws[f"{cl_a}5"] = val
        ws[f"{cl_a}5"].font = Font(name="Segoe UI", size=13, bold=True, color=cor)
        ws[f"{cl_a}5"].fill = fill_resumo
        ws[f"{cl_a}5"].alignment = al_l

    ws.row_dimensions[6].height = 6  # espaço

    # ── Linha 7: cabeçalho da tabela ─────────────────────────────────────────
    headers = ["FUNCIONÁRIO", "DATA", "DIA", "HORÁRIOS", "TRABALHADO", "EXTRAS", "FALTAS", "SALDO", "ALERTAS"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=i, value=h)
        c.font   = font_header
        c.fill   = fill_header
        c.alignment = al_c
        c.border = border
    ws.row_dimensions[7].height = 20

    # ── Dados ────────────────────────────────────────────────────────────────
    for row_idx, d in enumerate(dias_calc, 8):
        fill = fill_row1 if row_idx % 2 == 0 else fill_row2
        ws.row_dimensions[row_idx].height = 16

        saldo_m = d.get("saldo_min", 0)
        extras_m = d.get("extras_min", 0)
        faltas_m = d.get("faltas_min", 0)
        fn_extra  = font_success if extras_m > 0 else font_normal
        fn_falta  = font_danger  if faltas_m > 0 else font_normal
        fn_saldo  = font_success if saldo_m  >= 0 else font_danger

        vals = [
            (str(d.get("_func_nome", "")),           font_white,   al_l),
            (_fd(str(d.get("data", ""))),             font_normal,  al_c),
            (str(d.get("dia_nome", ""))[:3],          font_normal,  al_c),
            ("  ".join(str(h) for h in d.get("horarios", [])), font_normal, al_l),
            (str(d.get("trabalhados_fmt", "--:--")),  font_accent,  al_c),
            (str(d.get("extras_fmt", "--:--")),       fn_extra,     al_c),
            (str(d.get("faltas_fmt", "--:--")),       fn_falta,     al_c),
            (str(d.get("saldo_fmt",  "--:--")),       fn_saldo,     al_c),
            (" · ".join(d.get("alertas", [])),        font_warning, al_l),
        ]
        for col_idx, (val, fnt, aln) in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = fnt; c.fill = fill; c.alignment = aln; c.border = border

    # ── Larguras das colunas ─────────────────────────────────────────────────
    col_widths = [28, 12, 8, 34, 13, 12, 11, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congela cabeçalho
    ws.freeze_panes = "A8"

    wb.save(caminho)
    return caminho


def gerar_excel_banco_horas(funcionarios_saldo: list, caminho: str) -> str:
    """
    Gera planilha Excel do Banco de Horas.

    Retorna: caminho do arquivo gerado
    """
    _check_openpyxl()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"

    fill_h  = PatternFill("solid", fgColor="0D0F14")
    fill_r1 = PatternFill("solid", fgColor="1A1E2A")
    fill_r2 = PatternFill("solid", fgColor="161929")

    thin = Side(style="thin", color="1E2235")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:D1")
    ws["A1"] = "Banco de Horas — Saldo Geral"
    ws["A1"].font      = Font(name="Segoe UI", size=14, bold=True, color="EAEDF5")
    ws["A1"].fill      = fill_h
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Gerado em: {_now_str()}  |  Sistema de Ponto v5.0"
    ws["A2"].font      = Font(name="Segoe UI", size=9, color="8A90A8")
    ws["A2"].fill      = fill_h
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for i, h in enumerate(["FUNCIONÁRIO", "SALDO", "ATUALIZADO EM", "OBSERVAÇÃO"], 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Segoe UI", size=9, bold=True, color="00C2FF")
        c.fill = fill_h
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    ws.row_dimensions[4].height = 20

    for ri, f in enumerate(funcionarios_saldo, 5):
        fill = fill_r1 if ri % 2 == 0 else fill_r2
        saldo_m = f.get("saldo_min", 0)
        cor = "00E5A0" if saldo_m >= 0 else "FF4D6A"
        dados = [
            (f.get("nome", ""),                    Font(name="Segoe UI", size=9, bold=True, color="EAEDF5"), "left"),
            (_saldo_fmt(saldo_m),                   Font(name="Segoe UI", size=10, bold=True, color=cor),     "center"),
            (str(f.get("atualizado_em", ""))[:16], Font(name="Segoe UI", size=9, color="8A90A8"),              "center"),
            ("",                                    Font(name="Segoe UI", size=9, color="8A90A8"),              "left"),
        ]
        for ci, (v, fn, al) in enumerate(dados, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = fn; c.fill = fill
            c.alignment = Alignment(horizontal=al, vertical="center")
            c.border = brd
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.freeze_panes = "A5"

    # ── Aba por funcionário (histórico) ───────────────────────────────────────
    for f in funcionarios_saldo:
        hist = f.get("historico", [])
        if not hist:
            continue
        nome_aba = str(f.get("nome", "Func"))[:28]
        ws2 = wb.create_sheet(title=nome_aba)

        ws2.merge_cells("A1:D1")
        ws2["A1"] = f"Histórico — {nome_aba}"
        ws2["A1"].font = Font(name="Segoe UI", size=12, bold=True, color="EAEDF5")
        ws2["A1"].fill = fill_h
        ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 24

        saldo_m = f.get("saldo_min", 0)
        cor_s = "00E5A0" if saldo_m >= 0 else "FF4D6A"
        ws2.merge_cells("A2:D2")
        ws2["A2"] = f"Saldo atual: {_saldo_fmt(saldo_m)}"
        ws2["A2"].font = Font(name="Segoe UI", size=10, bold=True, color=cor_s)
        ws2["A2"].fill = fill_h
        ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[2].height = 18
        ws2.row_dimensions[3].height = 6

        for ci, h in enumerate(["DATA", "TIPO", "VALOR", "DESCRIÇÃO"], 1):
            c = ws2.cell(row=4, column=ci, value=h)
            c.font = Font(name="Segoe UI", size=9, bold=True, color="00C2FF")
            c.fill = fill_h
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
        ws2.row_dimensions[4].height = 18

        for ri, m in enumerate(hist, 5):
            fill = fill_r1 if ri % 2 == 0 else fill_r2
            tipo = m.get("tipo", "")
            cor_t = "00E5A0" if tipo == "credito" else "FF4D6A"
            tipo_label = "Crédito" if tipo == "credito" else "Débito"
            dados = [
                (str(m.get("data", ""))[:16], Font(name="Segoe UI", size=9, color="8A90A8"),  "center"),
                (tipo_label,                   Font(name="Segoe UI", size=9, color=cor_t),     "center"),
                (str(m.get("valor_fmt", "")),  Font(name="Segoe UI", size=9, bold=True, color=cor_t), "center"),
                (str(m.get("descricao", "")),  Font(name="Segoe UI", size=9, color="8A90A8"),  "left"),
            ]
            for ci, (v, fn, al) in enumerate(dados, 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.font = fn; c.fill = fill
                c.alignment = Alignment(horizontal=al, vertical="center")
                c.border = brd
            ws2.row_dimensions[ri].height = 16

        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 50
        ws2.freeze_panes = "A5"

    wb.save(caminho)
    return caminho
