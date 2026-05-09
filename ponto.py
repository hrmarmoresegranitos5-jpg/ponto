"""
╔══════════════════════════════════════════════════════════════════╗
║  SISTEMA DE PONTO AUTOMÁTICO — v5.0                             ║
║  Etapa 1 · Estrutura, DB, Cadastro de Funcionários              ║
║  Etapa 2 · Importação CSV / TXT / XLS / XLSX                    ║
║  Etapa 3 · Cálculo automático de jornada, extras e saldo        ║
║  Etapa 4 · Banco de Horas acumulado + retiradas                 ║
║  Etapa 5 · Relatórios PDF e exportação Excel
║  Etapa 4.6 · Notificações automáticas do funcionário                    ║
╚══════════════════════════════════════════════════════════════════╝
Requisitos : Python 3.8+  (tkinter incluso no Windows)
Opcionais  : pip install openpyxl xlrd   (.xlsx/.xls)
             pip install reportlab        (relatórios PDF)
Execute    : python ponto.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3, os, re
from datetime import datetime, timedelta
from collections import defaultdict

# ── Sincronização com a nuvem — bibliotecas opcionais ────────────────────────
try:
    import requests as _requests
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

try:
    import configparser as _configparser
    _CONFIGPARSER_OK = True
except ImportError:
    _CONFIGPARSER_OK = False

# ── Módulo de relatórios (Etapa 5) — importação opcional ─────────────────────
try:
    from relatorios_pdf import (gerar_pdf_calculo, gerar_pdf_banco_horas,
                                  gerar_excel_calculo, gerar_excel_banco_horas)
    _RELATORIOS_OK = True
except ImportError:
    _RELATORIOS_OK = False

# ── Módulo de notificações (Etapa 4.6) — importação opcional ─────────────────
try:
    from notificacoes import (GerenciadorNotif, BotaoSino, ToastInterno,
                               TipoNotif, CentralNotificacoes)
    _NOTIF_OK = True
except ImportError:
    _NOTIF_OK = False

# ──────────────────────────────────────────────────────────────────────────────
# GLOBAIS
# ──────────────────────────────────────────────────────────────────────────────
APP_TITLE   = "Sistema de Ponto"
APP_VERSION = "v5.1"
DB_FILE     = "ponto.db"
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ponto_config.ini")


# ── Leitura/gravação da configuração de nuvem ────────────────────────────────
def _cfg_load() -> dict:
    """Lê SERVIDOR_URL e ADMIN_KEY do ponto_config.ini (ou variáveis de ambiente)."""
    cfg = {"servidor_url": "", "admin_key": ""}
    if _CONFIGPARSER_OK and os.path.exists(CONFIG_FILE):
        p = _configparser.ConfigParser()
        p.read(CONFIG_FILE, encoding="utf-8")
        sec = p["nuvem"] if "nuvem" in p else {}
        cfg["servidor_url"] = sec.get("servidor_url", "").strip()
        cfg["admin_key"]    = sec.get("admin_key",    "").strip()
    # Variáveis de ambiente têm prioridade
    cfg["servidor_url"] = os.environ.get("SERVIDOR_URL", cfg["servidor_url"]).strip()
    cfg["admin_key"]    = os.environ.get("ADMIN_KEY",    cfg["admin_key"]   ).strip()
    return cfg


def _cfg_save(servidor_url: str, admin_key: str):
    """Salva as configurações de nuvem no ponto_config.ini."""
    if not _CONFIGPARSER_OK:
        return
    p = _configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        p.read(CONFIG_FILE, encoding="utf-8")
    if "nuvem" not in p:
        p["nuvem"] = {}
    p["nuvem"]["servidor_url"] = servidor_url.strip()
    p["nuvem"]["admin_key"]    = admin_key.strip()
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        p.write(f)


# ── Sincronização ─────────────────────────────────────────────────────────────
def sincronizar_nuvem(db) -> tuple:
    """
    Envia saldo + últimos movimentos de todos os funcionários para o servidor.
    Retorna (True, mensagem) em sucesso ou (False, erro) em falha.
    """
    if not _REQUESTS_OK:
        return False, "Biblioteca 'requests' não instalada.\nExecute:  pip install requests"

    cfg = _cfg_load()
    url = cfg["servidor_url"]
    key = cfg["admin_key"]

    if not url:
        return False, (
            "URL do servidor não configurada.\n"
            "Vá em Configurações → Nuvem e informe a URL do Railway."
        )

    payload = {"funcionarios": []}
    for func in db.get_funcionarios(apenas_ativos=False):
        fid       = func["id"]
        saldo, movs = db.get_banco_horas_func(fid)
        payload["funcionarios"].append({
            "nome":       func["nome"],
            "pin":        func["pin"] if "pin" in func.keys() else None,
            "saldo_min":  saldo,
            "salario":    func["salario"],
            "tipo_pag":   func["tipo_pag"],
            "jornada_h":  func["jornada_h"],
            "movimentos": [
                {
                    "tipo":       m["tipo"],
                    "minutos":    m["minutos"],
                    "descricao":  m["descricao"] or "",
                    "referencia": m["referencia"] or "",
                    "criado_em":  m["criado_em"],
                }
                for m in movs
            ],
        })

    endpoint = url.rstrip("/") + "/api/admin/sincronizar"
    headers  = {"Content-Type": "application/json"}
    if key:
        headers["X-Admin-Key"] = key

    try:
        resp = _requests.post(endpoint, json=payload, headers=headers, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            return True, f"✔ Sincronizado! {data.get('atualizados', '?')} funcionário(s) atualizados."
        return False, f"Erro {resp.status_code}: {resp.text[:300]}"
    except Exception as e:
        return False, f"Falha na conexão:\n{e}"


def enviar_notificacao_servidor(func_id=None, tipo="info", titulo="", mensagem="", icone="🔔") -> bool:
    """
    Envia uma notificação para o servidor (aparecerá no app do funcionário).
    func_id=None → envia para todos.
    Retorna True se enviou com sucesso.
    """
    if not _REQUESTS_OK:
        return False
    cfg = _cfg_load()
    url = cfg["servidor_url"]
    key = cfg["admin_key"]
    if not url:
        return False
    endpoint = url.rstrip("/") + "/api/admin/notificacao"
    headers  = {"Content-Type": "application/json"}
    if key:
        headers["X-Admin-Key"] = key
    payload = {"tipo": tipo, "titulo": titulo, "mensagem": mensagem, "icone": icone}
    if func_id:
        payload["func_id"] = func_id
    try:
        resp = _requests.post(endpoint, json=payload, headers=headers, timeout=8)
        return resp.status_code == 200
    except Exception:
        return False


C = {
    "bg_dark":    "#0D0F14",  "bg_panel":   "#13161E",
    "bg_card":    "#1A1E2A",  "bg_input":   "#0F1219",
    "sidebar":    "#0A0C12",  "accent":     "#00C2FF",
    "accent_dim": "#005599",  "success":    "#00E5A0",
    "warning":    "#FFB830",  "danger":     "#FF4D6A",
    "purple":     "#AA88FF",  "text_hi":    "#EAEDF5",
    "text_mid":   "#8A90A8",  "text_lo":    "#4A5068",
    "border":     "#1E2235",  "hover":      "#1F2438",
    "white":      "#FFFFFF",
    # extras para cálculo
    "extra_bg":   "#0A1A0A",  "falta_bg":   "#1A0A0A",
    "atencao_bg": "#1A1400",
}

FONT_TITLE  = ("Segoe UI", 22, "bold")
FONT_HEADER = ("Segoe UI", 13, "bold")
FONT_BODY   = ("Segoe UI", 11)
FONT_SMALL  = ("Segoe UI",  9)
FONT_MONO   = ("Consolas", 10)
FONT_BIG    = ("Segoe UI", 16, "bold")

DIAS_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]

# ──────────────────────────────────────────────────────────────────────────────
# ENGINE DE CÁLCULO DE HORAS  (Etapa 3)
# ──────────────────────────────────────────────────────────────────────────────
def _hm(s: str) -> int:
    """'HH:MM' → minutos desde meia-noite."""
    h, m = map(int, s.split(":"))
    return h * 60 + m

def _fmt(mins: int) -> str:
    """Minutos (absolutos) → 'HH:MM'."""
    m = abs(mins)
    return f"{m // 60:02d}:{m % 60:02d}"

def _saldo_fmt(mins: int) -> str:
    """Minutos (com sinal) → '+HH:MM' ou '-HH:MM'."""
    return f"{'-' if mins < 0 else '+'}{_fmt(mins)}"

def _semana_iso(data_str: str) -> str:
    """Retorna 'YYYY-Www' para agrupar por semana."""
    try:
        dt = datetime.strptime(data_str, "%Y-%m-%d")
        return dt.strftime("%Y-W%W")
    except Exception:
        return "?"


class CalcEngine:
    """
    Regras da empresa
    ─────────────────
    Seg–Sex  07:00–12:00  e  14:00–17:00  →  8 h
    Sábado   07:00–11:00                  →  4 h

    Regras de batida
    ─────────────────
    2 batidas → trabalho direto (sem almoço)
    4 batidas → entrada / saída-almoço / retorno / saída
    Ímpar     → calcula o que der e avisa
    """

    # Referências para detectar atraso / saída antecipada (em minutos)
    _REF_SEG_SEX = dict(entrada=_hm("07:00"), saida_almoco=_hm("12:00"),
                        retorno=_hm("14:00"), saida=_hm("17:00"))
    _REF_SAB     = dict(entrada=_hm("07:00"), saida=_hm("11:00"))
    _TOLERANCIA  = 5   # minutos de tolerância

    def calc_dia(self, data_str: str, horarios: list, jornada_h: float) -> dict:
        """Retorna dicionário completo com todos os campos calculados."""
        jornada_min = int(jornada_h * 60)
        alertas: list[str] = []

        try:
            dt = datetime.strptime(data_str, "%Y-%m-%d")
        except Exception:
            dt = None

        dia_idx    = dt.weekday() if dt else 0      # 0=seg … 6=dom
        dia_nome   = DIAS_PT[dia_idx]
        is_sabado  = dia_idx == 5
        is_domingo = dia_idx == 6
        ref        = self._REF_SAB if is_sabado else self._REF_SEG_SEX

        qtd         = len(horarios)
        trabalhados = 0

        if is_domingo and qtd == 0:
            alertas.append("Domingo sem registro")

        elif qtd == 0:
            alertas.append("Sem batidas registradas")

        elif qtd == 2:
            e = _hm(horarios[0]); s = _hm(horarios[1])
            trabalhados = max(0, s - e)
            T = self._TOLERANCIA
            if e > ref["entrada"] + T:
                alertas.append(f"Atraso ({horarios[0]})")
            if is_sabado:
                if s < ref["saida"] - T:
                    alertas.append(f"Saída antecipada ({horarios[1]})")
            else:
                ref_saida_2bat = _hm("17:00") if jornada_min >= 8*60 else _hm(
                    f"{7 + int(jornada_h):02d}:00")
                if s < ref_saida_2bat - T:
                    alertas.append(f"Saída antecipada ({horarios[1]})")

        elif qtd == 4:
            e1  = _hm(horarios[0]); sa  = _hm(horarios[1])
            ret = _hm(horarios[2]); e2  = _hm(horarios[3])
            manha = max(0, sa - e1);  tarde = max(0, e2 - ret)
            trabalhados = manha + tarde
            T = self._TOLERANCIA
            if e1  > ref["entrada"]      + T: alertas.append(f"Atraso ({horarios[0]})")
            if sa  < ref["saida_almoco"] - T: alertas.append(f"Almoço cedo ({horarios[1]})")
            if ret > ref["retorno"]      + 15: alertas.append(f"Retorno tardio ({horarios[2]})")
            if e2  < ref["saida"]        - T: alertas.append(f"Saída antecipada ({horarios[3]})")

        elif qtd % 2 == 0:
            for i in range(0, qtd, 2):
                trabalhados += max(0, _hm(horarios[i+1]) - _hm(horarios[i]))
        else:
            for i in range(0, qtd - 1, 2):
                trabalhados += max(0, _hm(horarios[i+1]) - _hm(horarios[i]))
            alertas.append("Nº ímpar de batidas — verificar")

        saldo  = trabalhados - jornada_min
        extras = max(0,  saldo)
        faltas = max(0, -saldo)

        if extras > 120:
            alertas.append("Excesso de horas (+2h)")
        if faltas > 30:
            alertas.append("Falta significativa")

        # status visual
        if faltas > 0 or "Nº ímpar" in " ".join(alertas):
            status = "falta"
        elif extras > 0:
            status = "extra"
        else:
            status = "ok"

        return {
            "data":            data_str,
            "dia_nome":        dia_nome,
            "horarios":        horarios,
            "qtd":             qtd,
            "trabalhados_min": trabalhados,
            "jornada_min":     jornada_min,
            "extras_min":      extras,
            "faltas_min":      faltas,
            "saldo_min":       saldo,
            "alertas":         alertas,
            "status":          status,
            # formatados
            "trabalhados_fmt": _fmt(trabalhados),
            "jornada_fmt":     _fmt(jornada_min),
            "extras_fmt":      _fmt(extras),
            "faltas_fmt":      _fmt(faltas),
            "saldo_fmt":       _saldo_fmt(saldo),
            "semana":          _semana_iso(data_str),
        }

    def calc_periodo(self, dias: list) -> dict:
        """Soma todos os dias e retorna totais + saldo de semanas."""
        total_trab = sum(d["trabalhados_min"] for d in dias)
        total_jorn = sum(d["jornada_min"]     for d in dias)
        total_ext  = sum(d["extras_min"]      for d in dias)
        total_flt  = sum(d["faltas_min"]      for d in dias)
        saldo      = total_trab - total_jorn

        # Agrupa por semana
        semanas: dict[str, int] = defaultdict(int)
        for d in dias:
            semanas[d["semana"]] += d["saldo_min"]

        return {
            "dias":           len(dias),
            "trabalhados_min": total_trab,
            "jornada_min":     total_jorn,
            "extras_min":      total_ext,
            "faltas_min":      total_flt,
            "saldo_min":       saldo,
            "trabalhados_fmt": _fmt(total_trab),
            "jornada_fmt":     _fmt(total_jorn),
            "extras_fmt":      _fmt(total_ext),
            "faltas_fmt":      _fmt(total_flt),
            "saldo_fmt":       _saldo_fmt(saldo),
            "semanas":         dict(semanas),
        }


# ──────────────────────────────────────────────────────────────────────────────
# ENGINE DE PARSING  (Etapa 2 — inalterado)
# ──────────────────────────────────────────────────────────────────────────────
TIME_RE      = re.compile(r'\b(\d{1,2}:\d{2}(?::\d{2})?)\b')
DATE_RE      = re.compile(
    r'\b(\d{4}[-/]\d{2}[-/]\d{2}|\d{2}[-/]\d{2}[-/]\d{4}|\d{2}[-/]\d{2}[-/]\d{2})\b')
KNUP_LINE_RE = re.compile(
    r'(?P<id>\d{5,12})\s+(?P<date>\d{2}[/-]\d{2}[/-]\d{2,4})\s+'
    r'(?P<time>\d{2}:\d{2}(?::\d{2})?)')

def _norm_date(raw):
    for fmt in ("%Y-%m-%d","%Y/%m/%d","%d-%m-%Y","%d/%m/%Y","%d-%m-%y","%d/%m/%y"):
        try: return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except: continue
    return None

def _norm_time(raw):
    p = raw.strip().split(":")
    return f"{int(p[0]):02d}:{int(p[1]):02d}" if len(p) >= 2 else raw.strip()


# ──────────────────────────────────────────────────────────────────────────────
# PARSER ESPECÍFICO PARA O RELÓGIO DE PONTO
# Suporta os três formatos exportados pelo relógio:
#   1. RegistroPresença.xls  — tabela mensal com dias como colunas
#   2. RelatórioAnormal.xls  — uma linha por funcionário/dia (entrada/saída)
#   3. RelatórioPresença.xls — cartão de ponto detalhado, múltiplos funcs por aba
# ──────────────────────────────────────────────────────────────────────────────
def _rt(v):
    """Converte datetime.time ou string 'HH:MM' para 'HH:MM'. Retorna None se inválido."""
    if v is None or str(v).strip().lower() in ('nenhum', 'none', '', 'nan'): return None
    try:
        import datetime as _dt
        if isinstance(v, _dt.time):
            return f"{v.hour:02d}:{v.minute:02d}"
    except Exception: pass
    s = str(v).strip()
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m: return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    return None

def _rd(v, ano_ref=None):
    """Converte datetime ou '01 4ª' para 'YYYY-MM-DD'."""
    try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.strftime("%Y-%m-%d")
    except Exception: pass
    if ano_ref:
        m = re.match(r'(\d{1,2})\s+\S', str(v).strip())
        if m:
            try:
                from datetime import date as _date
                return _date(ano_ref[0], ano_ref[1], int(m.group(1))).strftime("%Y-%m-%d")
            except Exception: pass
    return None

def _periodo(texto):
    """Extrai (ano, mes) de texto como '01/04/2026~25/04/2026'."""
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(texto))
    if m: return int(m.group(3)), int(m.group(2))
    return None


class RelogioPontoParser:
    """
    Parser dedicado aos arquivos exportados pelo relógio de ponto.
    Detecta automaticamente o formato pelo nome/conteúdo da aba.
    """

    def parse(self, filepath):
        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".xls":
            filepath = self._xls_para_xlsx(filepath)
            if not filepath:
                return [], ["Não foi possível converter o arquivo .xls.\n"
                            "Certifique-se de que o LibreOffice está instalado."]
        try:
            from openpyxl import load_workbook as _lwb
            wb = _lwb(filepath, read_only=True, data_only=True)
        except Exception as e:
            return [], [f"Erro ao abrir arquivo: {e}"]

        for sheet in wb.sheetnames:
            nome_aba = sheet.lower()
            if any(x in nome_aba for x in ('registro', 'registro de pres')):
                return self._parse_registro(wb[sheet])
            if 'anormal' in nome_aba:
                return self._parse_anormal(wb[sheet])
            if any(x in nome_aba for x in ('cartão', 'cart', '1,2', '4,5', 'presença de func')):
                return self._parse_cartao(wb[sheet])

        # Detecção por conteúdo
        for sheet in wb.sheetnames:
            rows_top = list(wb[sheet].iter_rows(values_only=True, max_row=8))
            flat = ' '.join(str(c) for r in rows_top for c in r if c).lower()
            if 'anormal' in flat:
                return self._parse_anormal(wb[sheet])
            if 'cartão de ponto' in flat or 'cartao de ponto' in flat:
                return self._parse_cartao(wb[sheet])
            if 'idusuário' in flat or 'idusuario' in flat:
                return self._parse_registro(wb[sheet])

        return [], ["Formato não reconhecido. Use RegistroPresença, RelatórioAnormal ou RelatórioPresença."]

    def _xls_para_xlsx(self, fp):
        """Le .xls com xlrd3 (puro Python, sem LibreOffice) e salva como .xlsx."""
        import tempfile, subprocess, sys
        try:
            import xlrd3 as xlrd
        except ImportError:
            try:
                subprocess.run([sys.executable, "-m", "pip", "install", "xlrd3"],
                               capture_output=True, timeout=60)
                import xlrd3 as xlrd
            except Exception:
                return None
        try:
            from openpyxl import Workbook
            import datetime as _dt
            wb_in  = xlrd.open_workbook(fp)
            wb_out = Workbook()
            wb_out.remove(wb_out.active)
            for sheet_idx in range(wb_in.nsheets):
                ws_in  = wb_in.sheet_by_index(sheet_idx)
                ws_out = wb_out.create_sheet(title=ws_in.name)
                for row_idx in range(ws_in.nrows):
                    for col_idx in range(ws_in.ncols):
                        cell = ws_in.cell(row_idx, col_idx)
                        val  = cell.value
                        if cell.ctype == 3:
                            try:
                                tup = xlrd.xldate_as_tuple(val, wb_in.datemode)
                                if tup[0] == 0:
                                    val = _dt.time(tup[3], tup[4], tup[5])
                                else:
                                    val = _dt.datetime(*tup)
                            except Exception:
                                pass
                        elif cell.ctype == 2 and val == int(val):
                            val = int(val)
                        ws_out.cell(row=row_idx + 1, column=col_idx + 1, value=val)
            out_path = os.path.join(tempfile.mkdtemp(),
                                    os.path.splitext(os.path.basename(fp))[0] + ".xlsx")
            wb_out.save(out_path)
            return out_path
        except Exception:
            return None

    # ── Formato 1: RegistroPresença ──────────────────────────────────────────
    # Cabeçalho a cada bloco de 3 linhas:
    #   L1: IDUsuário: X | Nome: fulano | Dep.: marmoraria
    #   L2: 1  2  3 ... 25  (números dos dias)
    #   L3: batidas de cada dia separadas por \n dentro da célula
    def _parse_registro(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        resultado, avisos = [], []
        periodo = None
        for row in rows[:6]:
            for c in row:
                if c and 'Data de presen' in str(c):
                    periodo = _periodo(str(c)); break

        i = 0
        while i < len(rows):
            row = rows[i]
            flat = ' '.join(str(c).lower() for c in row if c)
            if 'idusuário:' in flat or 'idusuario:' in flat:
                # Extrai nome
                nome = None
                for j, c in enumerate(row):
                    if c and 'nome' in str(c).lower() and j + 1 < len(row) and rows[i][j+1]:
                        nome = str(rows[i][j+1]).strip().lower(); break
                if nome and i + 2 < len(rows):
                    dias_row = rows[i + 1]
                    bat_row  = rows[i + 2]
                    for ci, dia_num in enumerate(dias_row):
                        if dia_num is None: continue
                        s = str(dia_num).strip()
                        if not s.isdigit(): continue
                        dia = int(s)
                        celula = bat_row[ci] if ci < len(bat_row) else None
                        if celula is None: continue
                        horarios = []
                        for h in str(celula).split('\n'):
                            t = _rt(h.strip())
                            if t: horarios.append(t)
                        if horarios and periodo:
                            try:
                                from datetime import date as _date
                                ds = _date(periodo[0], periodo[1], dia).strftime("%Y-%m-%d")
                                resultado.append({"funcionario_raw": nome, "data": ds,
                                    "horarios": sorted(horarios), "qtd_batidas": len(horarios), "erros": []})
                            except Exception: pass
                i += 3
            else:
                i += 1
        if not resultado:
            avisos.append("RegistroPresença: nenhuma batida encontrada.")
        return resultado, avisos

    # ── Formato 2: RelatórioAnormal ──────────────────────────────────────────
    # Colunas: IDUsuário | Nome | Dep. | Data | E.manhã | S.manhã | E.tarde | S.tarde | …
    def _parse_anormal(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        resultado, avisos = [], []
        # Acha linha de cabeçalho
        hdr = None
        for i, row in enumerate(rows):
            flat = [str(c).lower() if c else '' for c in row]
            if any('idusuário' in x or 'idusuario' in x for x in flat):
                hdr = i; break
        if hdr is None:
            return [], ["RelatórioAnormal: cabeçalho não encontrado."]
        for row in rows[hdr + 2:]:
            if not row or row[0] is None: continue
            try:
                id_usr = row[0]
                nome   = str(row[1]).strip().lower() if row[1] else None
                data_v = row[3]
                em = _rt(row[4]) if len(row) > 4 else None
                sm = _rt(row[5]) if len(row) > 5 else None
                et = _rt(row[6]) if len(row) > 6 else None
                st = _rt(row[7]) if len(row) > 7 else None
                if not nome or not isinstance(id_usr, (int, float)): continue
                from datetime import datetime as _dt
                if not isinstance(data_v, _dt): continue
                ds = data_v.strftime("%Y-%m-%d")
                horarios = [h for h in [em, sm, et, st] if h]
                if horarios:
                    resultado.append({"funcionario_raw": nome, "data": ds,
                        "horarios": sorted(horarios), "qtd_batidas": len(horarios), "erros": []})
            except Exception as e:
                avisos.append(f"Linha ignorada: {e}")
        if not resultado:
            avisos.append("RelatórioAnormal: nenhuma batida com horário encontrada.")
        return resultado, avisos

    # ── Formato 3: RelatórioPresença (cartão de ponto) ───────────────────────
    # Vários funcionários lado a lado em blocos de ~15 colunas.
    # Cabeçalho: 'Dep.' | dep | … | 'Nome' | nome_func
    # Dados: '01 4ª' | E.manhã | – | S.manhã | – | – | E.tarde | – | S.tarde | E.ext | – | S.ext
    def _parse_cartao(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        resultado, avisos = [], []
        periodo = None
        for row in rows[:6]:
            for c in row:
                if c and '~' in str(c):
                    periodo = _periodo(str(c)); break

        # Mapa col_bloco → nome do funcionário
        func_map = {}
        nome_row_idx = None
        for i, row in enumerate(rows[:15]):
            for ci, c in enumerate(row):
                if c and str(c).strip() == 'Nome':
                    if ci + 1 < len(row) and row[ci + 1]:
                        nome = str(row[ci + 1]).strip().lower()
                        bloco = (ci // 15) * 15
                        func_map[bloco] = nome
                        nome_row_idx = i
            if nome_row_idx and i > nome_row_idx + 1: break

        if not func_map:
            return [], ["RelatórioPresença: funcionários não identificados."]

        # Acha linha de início dos dados (após 'Cartão de ponto' + 2 sub-cabeçalhos)
        data_start = None
        for i, row in enumerate(rows):
            flat = [str(c) for c in row if c]
            if any('Cartão de ponto' in x or 'Cartao de ponto' in x for x in flat):
                data_start = i + 3; break
        if data_start is None:
            # Fallback: primeira linha com padrão '01 4ª'
            for i, row in enumerate(rows):
                if any(c and re.match(r'^\d{2}\s+\S', str(c)) for c in row):
                    data_start = i; break
        if data_start is None:
            return [], ["RelatórioPresença: dados não encontrados."]

        # Offsets das colunas de horário dentro de cada bloco de 15 cols
        HORA_OFFSETS = [1, 3, 6, 8, 10, 12]
        for row in rows[data_start:]:
            if not row or all(c is None for c in row): continue
            for bloco, nome in func_map.items():
                if bloco >= len(row): continue
                data_v = row[bloco]
                if data_v is None: continue
                ds = _rd(data_v, periodo)
                if not ds: continue
                horarios = []
                for off in HORA_OFFSETS:
                    ci = bloco + off
                    if ci < len(row):
                        t = _rt(row[ci])
                        if t: horarios.append(t)
                if horarios:
                    resultado.append({"funcionario_raw": nome, "data": ds,
                        "horarios": sorted(horarios), "qtd_batidas": len(horarios), "erros": []})

        if not resultado:
            avisos.append("RelatórioPresença: nenhuma batida encontrada.")
        return resultado, avisos


class PontoParser:
    """
    Parser principal: tenta primeiro o RelogioPontoParser (formato do relógio).
    Se não reconhecer, usa o parser genérico para CSV/TXT.
    """
    def __init__(self):
        self._relogio = RelogioPontoParser()

    def parse(self, filepath):
        ext = os.path.splitext(filepath)[1].lower()
        # Arquivos Excel → sempre tenta o parser do relógio primeiro
        if ext in (".xls", ".xlsx"):
            regs, avisos = self._relogio.parse(filepath)
            if regs:
                return regs, avisos
            # Se não encontrou nada, avisa
            return regs, avisos or ["Arquivo Excel não reconhecido. "
                                     "Use RegistroPresença, RelatórioAnormal ou RelatórioPresença."]
        # CSV/TXT → parser genérico
        try:
            rows = self._read_text(filepath)
        except Exception as e:
            return [], [f"Erro ao abrir: {e}"]
        return self._parse_rows(rows)

    def _read_text(self, fp):
        for enc in ("utf-8","latin-1","cp1252","utf-8-sig"):
            try:
                with open(fp, encoding=enc, errors="replace") as f: return f.readlines()
            except: continue
        return []

    def _parse_rows(self, lines):
        avisos = []; bucket = defaultdict(lambda: {"horarios": set(), "erros": []})
        knup = sum(1 for l in lines if KNUP_LINE_RE.search(l))
        if knup:
            for line in lines:
                m = KNUP_LINE_RE.search(line)
                if m:
                    fid = m.group("id").lstrip("0") or m.group("id")
                    d   = _norm_date(m.group("date"))
                    if d: bucket[(fid, d)]["horarios"].add(_norm_time(m.group("time")))
        else:
            cur_func = "Desconhecido"; cur_date = None
            for line in lines:
                line = line.strip()
                if not line or line.startswith("#"): continue
                m = re.match(r'(?:funcion[aá]rio|nome|colaborador)\s*[:\-]?\s*(.+)', line, re.I)
                if m: cur_func = m.group(1).strip(); continue
                if any(sep in line for sep in (",",";","\t")):
                    sep = "," if "," in line else (";" if ";" in line else "\t")
                    self._cols([c.strip().strip('"') for c in line.split(sep)], cur_func, bucket)
                    continue
                times = TIME_RE.findall(line); dates = DATE_RE.findall(line)
                ex = None
                for d in dates:
                    nd = _norm_date(d)
                    if nd: ex = nd; cur_date = nd; break
                if not ex: ex = cur_date
                if times and ex:
                    for t in times: bucket[(cur_func, ex)]["horarios"].add(_norm_time(t))
        if not bucket:
            avisos.append("Nenhum registro reconhecido. Verifique o formato.")
        result = []
        for (fr, ds), v in sorted(bucket.items()):
            h = sorted(v["horarios"])
            result.append({"funcionario_raw":fr,"data":ds,"horarios":h,
                           "qtd_batidas":len(h),"erros":v["erros"]})
        return result, avisos

    def _cols(self, cols, default_func, bucket):
        fr = default_func; ds = None; hrs = []
        for col in cols:
            nd = _norm_date(col)
            if nd and not ds: ds = nd; continue
            if TIME_RE.fullmatch(col.strip()): hrs.append(_norm_time(col)); continue
            if col and not ds and not TIME_RE.search(col):
                if re.match(r'^[A-Za-zÀ-ÿ\s]+$', col) and len(col) > 2: fr = col.strip()
                elif re.match(r'^\d{3,12}$', col): fr = col.strip()
        if ds and hrs:
            for h in hrs: bucket[(fr, ds)]["horarios"].add(h)


# ──────────────────────────────────────────────────────────────────────────────
# BANCO DE DADOS
# ──────────────────────────────────────────────────────────────────────────────
class Database:
    def __init__(self, path=DB_FILE):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self._migrate()
        self._seed()
        # ── Gerenciador de notificações (Etapa 4.6) ───────────────────────────
        if _NOTIF_OK:
            self.notif = GerenciadorNotif(self.conn)
        else:
            self.notif = None

    def _migrate(self):
        self.conn.executescript("""
        CREATE TABLE IF NOT EXISTS funcionarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE, salario REAL NOT NULL,
            tipo_pag TEXT NOT NULL,   jornada_h REAL NOT NULL,
            observacoes TEXT,         ativo INTEGER NOT NULL DEFAULT 1,
            criado_em TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS importacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            arquivo TEXT NOT NULL,    status TEXT NOT NULL DEFAULT 'ok',
            importado_em TEXT NOT NULL, total_reg INTEGER DEFAULT 0,
            total_erros INTEGER DEFAULT 0, avisos TEXT
        );
        CREATE TABLE IF NOT EXISTS batidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            importacao_id INTEGER NOT NULL, funcionario_id INTEGER,
            funcionario_raw TEXT NOT NULL,  data TEXT NOT NULL,
            horarios TEXT NOT NULL,         qtd_batidas INTEGER NOT NULL DEFAULT 0,
            tem_erro INTEGER NOT NULL DEFAULT 0, erros TEXT,
            FOREIGN KEY (importacao_id) REFERENCES importacoes(id),
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        CREATE TABLE IF NOT EXISTS banco_horas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL UNIQUE,
            saldo_min INTEGER NOT NULL DEFAULT 0,
            atualizado_em TEXT NOT NULL,
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        CREATE TABLE IF NOT EXISTS banco_horas_mov (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            minutos INTEGER NOT NULL,
            descricao TEXT,
            referencia TEXT,
            criado_em TEXT NOT NULL,
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        CREATE TABLE IF NOT EXISTS pagamentos_horas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            minutos_pagos INTEGER NOT NULL,
            valor_pago REAL NOT NULL,
            tipo TEXT NOT NULL,
            descricao TEXT,
            criado_em TEXT NOT NULL,
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        """)
        self.conn.commit()

    def _seed(self):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Adiciona coluna pin se ainda não existe (migração segura)
        try:
            self.conn.execute("ALTER TABLE funcionarios ADD COLUMN pin TEXT DEFAULT NULL")
            self.conn.commit()
        except Exception:
            pass
        for nome, sal, tipo, jorn, obs, pin in [
            ("hangel",   1500.0, "mes",    8.0, "Pagamento mensal",        "1111"),
            ("fabricio", 400.0,  "semana", 8.0, "Pagamento semanal",       "1234"),
            ("gibs",     1000.0, "mes",    8.0, "Pagamento mensal",        "2345"),
            ("hugo",     1900.0, "mes",    8.0, "Pagamento mensal",        "3456"),
            ("tiago",    400.0,  "mes",    4.0, "Meio periodo - 4h/dia",   "4567"),
        ]:
            self.conn.execute(
                "INSERT OR IGNORE INTO funcionarios(nome,salario,tipo_pag,jornada_h,observacoes,criado_em,pin) VALUES(?,?,?,?,?,?,?)",
                (nome, sal, tipo, jorn, obs, now, pin))
        self.conn.commit()
        # Garante que todo funcionário tem linha no banco_horas
        for f in self.get_funcionarios(apenas_ativos=False):
            self.conn.execute(
                "INSERT OR IGNORE INTO banco_horas(funcionario_id,saldo_min,atualizado_em) VALUES(?,0,?)",
                (f["id"], now))
        self.conn.commit()

    # ── Funcionários ──────────────────────────────────────────────────────────
    def get_funcionarios(self, apenas_ativos=True):
        q = "SELECT * FROM funcionarios" + (" WHERE ativo=1" if apenas_ativos else "")
        return self.conn.execute(q + " ORDER BY nome").fetchall()

    def get_funcionario(self, fid):
        return self.conn.execute("SELECT * FROM funcionarios WHERE id=?", (fid,)).fetchone()

    def save_funcionario(self, data, fid=None):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if fid:
            self.conn.execute(
                "UPDATE funcionarios SET nome=?,salario=?,tipo_pag=?,jornada_h=?,observacoes=? WHERE id=?",
                (data["nome"],data["salario"],data["tipo_pag"],data["jornada_h"],data["observacoes"],fid))
        else:
            self.conn.execute(
                "INSERT INTO funcionarios(nome,salario,tipo_pag,jornada_h,observacoes,criado_em) VALUES(?,?,?,?,?,?)",
                (data["nome"],data["salario"],data["tipo_pag"],data["jornada_h"],data["observacoes"],now))
        self.conn.commit()

    def toggle_ativo(self, fid):
        self.conn.execute("UPDATE funcionarios SET ativo=1-ativo WHERE id=?", (fid,))
        self.conn.commit()

    def delete_funcionario(self, fid):
        self.conn.execute("DELETE FROM funcionarios WHERE id=?", (fid,))
        self.conn.commit()

    # ── Importações ───────────────────────────────────────────────────────────
    def save_importacao(self, arquivo, registros, avisos):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = self.conn.execute(
            "INSERT INTO importacoes(arquivo,status,importado_em,total_reg,total_erros,avisos) VALUES(?,?,?,?,?,?)",
            (arquivo, "aviso" if avisos else "ok", now, len(registros),
             sum(1 for r in registros if r.get("erros")), "\n".join(avisos)))
        imp_id = cur.lastrowid
        funcs  = {f["nome"].lower(): f["id"] for f in self.get_funcionarios(apenas_ativos=False)}
        func_jornada = {f["id"]: f["jornada_h"] for f in self.get_funcionarios(apenas_ativos=False)}

        # Acumula extras por funcionário nessa importação para creditar no banco
        extras_por_func: dict = {}

        engine = CalcEngine()
        for reg in registros:
            raw = reg["funcionario_raw"]; fid = self._match(raw, funcs)
            self.conn.execute(
                "INSERT INTO batidas(importacao_id,funcionario_id,funcionario_raw,data,horarios,qtd_batidas,tem_erro,erros) VALUES(?,?,?,?,?,?,?,?)",
                (imp_id, fid, raw, reg["data"], ",".join(reg["horarios"]),
                 reg["qtd_batidas"], 1 if reg.get("erros") else 0,
                 "; ".join(reg["erros"]) if reg.get("erros") else ""))
            # Calcula extras do dia e acumula para o banco de horas
            if fid and reg["horarios"] and not reg.get("erros"):
                jorn = func_jornada.get(fid, 8.0)
                calc = engine.calc_dia(reg["data"], reg["horarios"], jorn)
                extras_min = calc.get("extras_min", 0)
                if extras_min > 0:
                    extras_por_func[fid] = extras_por_func.get(fid, 0) + extras_min

        self.conn.commit()

        # Credita banco de horas automaticamente
        for fid, total_extras in extras_por_func.items():
            self.creditar_banco_horas(
                fid, total_extras,
                f"Importação: {os.path.basename(arquivo)}",
                referencia=str(imp_id)
            )

        # ── Sincronização automática com a nuvem ──────────────────────────
        try:
            ok, msg = sincronizar_nuvem(self)
            self._ultimo_sync = (ok, msg)
        except Exception:
            self._ultimo_sync = (False, "Erro ao sincronizar.")

        return imp_id

    def _match(self, raw, funcs):
        rl = raw.lower().strip()
        if rl in funcs: return funcs[rl]
        for n, fid in funcs.items():
            if rl in n or n in rl: return fid
        return None

    def get_importacoes(self):
        return self.conn.execute("SELECT * FROM importacoes ORDER BY importado_em DESC").fetchall()

    def get_importacao(self, imp_id):
        return self.conn.execute("SELECT * FROM importacoes WHERE id=?", (imp_id,)).fetchone()

    def get_batidas(self, imp_id=None, func_id=None):
        q = "SELECT b.*, f.nome as func_nome FROM batidas b LEFT JOIN funcionarios f ON f.id=b.funcionario_id WHERE 1=1"
        a = []
        if imp_id  is not None: q += " AND b.importacao_id=?"; a.append(imp_id)
        if func_id is not None: q += " AND b.funcionario_id=?"; a.append(func_id)
        return self.conn.execute(q + " ORDER BY b.funcionario_raw, b.data", a).fetchall()

    def delete_importacao(self, imp_id):
        self.conn.execute("DELETE FROM batidas     WHERE importacao_id=?", (imp_id,))
        self.conn.execute("DELETE FROM importacoes WHERE id=?",            (imp_id,))
        self.conn.commit()

    # ── Banco de Horas ────────────────────────────────────────────────────────
    def get_banco_horas(self):
        """Retorna saldo atual de todos os funcionários."""
        return self.conn.execute("""
            SELECT f.id, f.nome, f.pin, f.salario, f.tipo_pag, f.jornada_h,
                   COALESCE(b.saldo_min, 0) as saldo_min, b.atualizado_em
            FROM funcionarios f
            LEFT JOIN banco_horas b ON b.funcionario_id = f.id
            WHERE f.ativo = 1
            ORDER BY f.nome
        """).fetchall()

    def get_folha_pagamento(self):
        """Retorna funcionários com saldo > 0 e valor monetário calculado."""
        rows = self.conn.execute("""
            SELECT f.id, f.nome, f.salario, f.tipo_pag, f.jornada_h,
                   COALESCE(b.saldo_min, 0) as saldo_min, b.atualizado_em
            FROM funcionarios f
            LEFT JOIN banco_horas b ON b.funcionario_id = f.id
            WHERE f.ativo = 1
            ORDER BY f.nome
        """).fetchall()
        result = []
        for r in rows:
            sal   = r["salario"]
            tipo  = r["tipo_pag"]
            jorn  = r["jornada_h"]
            # valor por hora: salário / horas do período
            if tipo == "semana":
                horas_periodo = jorn * 5
            else:
                horas_periodo = jorn * 5 * 4.333
            valor_hora = sal / horas_periodo if horas_periodo else 0
            saldo_h    = r["saldo_min"] / 60.0
            valor_dev  = saldo_h * valor_hora
            # total já pago
            pago_row = self.conn.execute(
                "SELECT COALESCE(SUM(valor_pago),0) as total FROM pagamentos_horas WHERE funcionario_id=?",
                (r["id"],)).fetchone()
            total_pago = pago_row["total"] if pago_row else 0
            result.append({
                "id":          r["id"],
                "nome":        r["nome"],
                "salario":     sal,
                "tipo_pag":    tipo,
                "jornada_h":   jorn,
                "saldo_min":   r["saldo_min"],
                "valor_hora":  valor_hora,
                "valor_devido": valor_dev,
                "total_pago":  total_pago,
                "a_pagar":     max(0, valor_dev - total_pago),
                "atualizado_em": r["atualizado_em"] or "",
            })
        return result

    def registrar_pagamento_horas(self, func_id, minutos_pagos, valor_pago, tipo, descricao):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute(
            "INSERT INTO pagamentos_horas(funcionario_id,minutos_pagos,valor_pago,tipo,descricao,criado_em)"
            " VALUES(?,?,?,?,?,?)",
            (func_id, minutos_pagos, valor_pago, tipo, descricao, now))
        self.conn.commit()
        # ── Notificação automática (local) ────────────────────────────────────
        if self.notif:
            func = self.get_funcionario(func_id)
            nome = func["nome"].capitalize() if func else "Funcionário"
            brl = f"R$ {valor_pago:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            self.notif.pagamento_lancado(nome, func_id, brl)
        # ── Notificação remota (app do funcionário) ────────────────────────────
        try:
            func = self.get_funcionario(func_id)
            nome = func["nome"].capitalize() if func else "Funcionário"
            brl  = f"R$ {valor_pago:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            enviar_notificacao_servidor(
                func_id=func_id,
                tipo="pagamento",
                titulo=f"💵 Pagamento registrado: {brl}",
                mensagem=f"Um pagamento de {brl} foi lançado para {nome}. {descricao or ''}".strip(),
                icone="💵"
            )
        except Exception:
            pass

    def get_historico_pagamentos(self, func_id):
        return self.conn.execute(
            "SELECT * FROM pagamentos_horas WHERE funcionario_id=? ORDER BY criado_em DESC LIMIT 30",
            (func_id,)).fetchall()

    def get_banco_horas_func(self, func_id):
        """Retorna saldo + histórico de movimentos de um funcionário."""
        saldo = self.conn.execute(
            "SELECT saldo_min FROM banco_horas WHERE funcionario_id=?", (func_id,)
        ).fetchone()
        movs = self.conn.execute("""
            SELECT * FROM banco_horas_mov
            WHERE funcionario_id=?
            ORDER BY criado_em DESC LIMIT 50
        """, (func_id,)).fetchall()
        return (saldo["saldo_min"] if saldo else 0), movs

    def creditar_banco_horas(self, func_id, minutos, descricao, referencia=""):
        """Adiciona minutos ao banco (chamado automaticamente na importação)."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute(
            "INSERT OR IGNORE INTO banco_horas(funcionario_id,saldo_min,atualizado_em) VALUES(?,0,?)",
            (func_id, now))
        self.conn.execute(
            "UPDATE banco_horas SET saldo_min=saldo_min+?, atualizado_em=? WHERE funcionario_id=?",
            (minutos, now, func_id))
        self.conn.execute(
            "INSERT INTO banco_horas_mov(funcionario_id,tipo,minutos,descricao,referencia,criado_em) VALUES(?,?,?,?,?,?)",
            (func_id, "credito", minutos, descricao, referencia, now))
        self.conn.commit()
        # ── Notificação automática ────────────────────────────────────────────
        if self.notif and minutos != 0:
            func = self.get_funcionario(func_id)
            nome = func["nome"].capitalize() if func else "Funcionário"
            # Saldo novo
            saldo_row = self.conn.execute(
                "SELECT saldo_min FROM banco_horas WHERE funcionario_id=?",
                (func_id,)).fetchone()
            saldo_min = saldo_row["saldo_min"] if saldo_row else 0
            if minutos > 0:
                h, m = divmod(abs(minutos), 60)
                horas_str = f"+{h}h{m:02d}" if h else f"+{m}min"
                self.notif.horas_extras(nome, func_id, horas_str,
                                        referencia or "")
            else:
                h, m = divmod(abs(minutos), 60)
                tempo_str = f"{h}h{m:02d}" if h else f"{m} minutos"
                self.notif.horas_negativas(nome, func_id, tempo_str)
            # Notificação de saldo sempre após crédito
            if saldo_min != 0:
                # Calcula valor
                func_data = self.get_funcionario(func_id)
                if func_data:
                    sal = func_data["salario"]
                    tipo = func_data["tipo_pag"]
                    if tipo == "hora":
                        valor_hora = sal
                    elif tipo == "semana":
                        valor_hora = sal / (func_data["jornada_h"] * 5)
                    else:
                        valor_hora = sal / (func_data["jornada_h"] * 22)
                    valor = (saldo_min / 60) * valor_hora * 1.5
                    brl = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    self.notif.saldo_atualizado(nome, func_id, brl)

    def debitar_banco_horas(self, func_id, minutos, descricao):
        """Remove minutos do banco (retirada aprovada pelo gestor)."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute(
            "UPDATE banco_horas SET saldo_min=saldo_min-?, atualizado_em=? WHERE funcionario_id=?",
            (minutos, now, func_id))
        self.conn.execute(
            "INSERT INTO banco_horas_mov(funcionario_id,tipo,minutos,descricao,referencia,criado_em) VALUES(?,?,?,?,?,?)",
            (func_id, "debito", minutos, descricao, "", now))
        self.conn.commit()
        # ── Sincroniza com a nuvem ─────────────────────────────────────────
        try:
            ok, msg = sincronizar_nuvem(self)
            self._ultimo_sync = (ok, msg)
        except Exception:
            self._ultimo_sync = (False, "Erro ao sincronizar.")

    def get_pin_func(self, func_id):
        row = self.conn.execute("SELECT pin FROM funcionarios WHERE id=?", (func_id,)).fetchone()
        return row["pin"] if row else None

    def set_pin_func(self, func_id, pin):
        self.conn.execute("UPDATE funcionarios SET pin=? WHERE id=?", (pin, func_id))
        self.conn.commit()

    def get_stats(self):
        f = self.conn.execute("SELECT COUNT(*) FROM funcionarios WHERE ativo=1").fetchone()[0]
        i = self.conn.execute("SELECT COUNT(*) FROM importacoes").fetchone()[0]
        r = self.conn.execute("SELECT COUNT(*) FROM batidas").fetchone()[0]
        return f, i, r


# ──────────────────────────────────────────────────────────────────────────────
# WIDGETS
# ──────────────────────────────────────────────────────────────────────────────
def _tv_style():
    s = ttk.Style(); s.theme_use("default")
    s.configure("Dark.Treeview", background=C["bg_card"], foreground=C["text_hi"],
        rowheight=40, fieldbackground=C["bg_card"], borderwidth=0, font=("Segoe UI",11))
    s.configure("Dark.Treeview.Heading", background=C["bg_dark"], foreground=C["text_mid"],
        font=("Segoe UI",9,"bold"), borderwidth=0, relief="flat")
    s.map("Dark.Treeview", background=[("selected",C["accent_dim"])],
          foreground=[("selected",C["white"])])
    s.layout("Dark.Treeview", [("Dark.Treeview.treearea", {"sticky":"nswe"})])

class Sep(tk.Frame):
    def __init__(self, p, **kw): super().__init__(p, bg=C["border"], height=1, **kw)

class Btn(tk.Frame):
    _clr = {"primary":(C["accent"],C["bg_dark"]), "success":(C["success"],C["bg_dark"]),
            "danger": (C["danger"],C["white"]),   "outline":(C["border"],C["text_hi"]),
            "warning":(C["warning"],C["bg_dark"])}
    def __init__(self, parent, text, cmd=None, style="primary", icon="", **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self._bg, self._fg = self._clr.get(style, self._clr["primary"])
        self._cmd = cmd
        self.lbl = tk.Label(self, text=(f"{icon}  {text}" if icon else text),
            bg=self._bg, fg=self._fg, font=("Segoe UI",10,"bold"),
            cursor="hand2", padx=18, pady=8)
        self.lbl.pack(fill="both", expand=True)
        self.lbl.bind("<Button-1>", lambda e: self._cmd() if self._cmd else None)
        self.lbl.bind("<Enter>",    lambda e: self.lbl.config(bg=self._lite()))
        self.lbl.bind("<Leave>",    lambda e: self.lbl.config(bg=self._bg))
    def _lite(self):
        try:
            r,g,b = int(self._bg[1:3],16),int(self._bg[3:5],16),int(self._bg[5:7],16)
            return f"#{min(255,r+25):02x}{min(255,g+25):02x}{min(255,b+25):02x}"
        except: return self._bg

class Card(tk.Frame):
    def __init__(self, parent, title, value, sub="", color=None, **kw):
        super().__init__(parent, bg=C["bg_card"],
                         highlightbackground=C["border"], highlightthickness=1, **kw)
        color = color or C["accent"]
        tk.Frame(self, bg=color, height=3).pack(fill="x")
        inn = tk.Frame(self, bg=C["bg_card"], padx=20, pady=16); inn.pack(fill="both", expand=True)
        tk.Label(inn, text=title,      bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
        tk.Label(inn, text=str(value), bg=C["bg_card"], fg=color, font=("Segoe UI",28,"bold")).pack(anchor="w", pady=0)
        if sub: tk.Label(inn, text=sub, bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w", pady=0)

def _fd(d):
    try: return datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
    except: return d


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — DASHBOARD
# ──────────────────────────────────────────────────────────────────────────────
class PageDashboard(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db; self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], pady=30, padx=36); hdr.pack(fill="x")
        tk.Label(hdr, text="Dashboard", bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(hdr, text=f"Visão geral  ·  {datetime.now().strftime('%d/%m/%Y')}",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Sep(self).pack(fill="x", padx=36)

        cf = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=24); cf.pack(fill="x")
        f, i, r = self.db.get_stats()
        for idx, (t,v,s,c) in enumerate([
            ("Funcionários Ativos", f, "cadastrados",     C["accent"]),
            ("Importações",         i, "arquivos lidos",  C["success"]),
            ("Registros de Ponto",  r, "batidas no banco",C["warning"]),
        ]):
            Card(cf, t, v, s, c).grid(row=0, column=idx, padx=16, sticky="ew")
            cf.columnconfigure(idx, weight=1)

        Sep(self).pack(fill="x", padx=36)
        info = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=24); info.pack(fill="both", expand=True)
        info.columnconfigure(0, weight=1); info.columnconfigure(1, weight=1)

        # Equipe
        lf = tk.Frame(info, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        lf.grid(row=0, column=0, padx=12, sticky="nsew")
        tk.Frame(lf, bg=C["accent"], height=3).pack(fill="x")
        li = tk.Frame(lf, bg=C["bg_card"], padx=20, pady=16); li.pack(fill="both", expand=True)
        tk.Label(li, text="Equipe Cadastrada", bg=C["bg_card"], fg=C["text_hi"], font=FONT_HEADER).pack(anchor="w", pady=12)
        for nome, sal, jorn, color in [
            ("Hangel",   "R$ 1.500/mês",  "8h/dia", C["accent"]),
            ("Fabricio", "R$ 400/semana", "8h/dia", C["success"]),
            ("Gibs",     "R$ 1.000/mês",  "8h/dia", C["warning"]),
            ("Hugo",     "R$ 1.900/mês",  "8h/dia", C["purple"]),
            ("Tiago",    "R$ 400/mês",    "4h/dia", C["text_mid"]),
        ]:
            rw = tk.Frame(li, bg=C["bg_card"], pady=6); rw.pack(fill="x")
            tk.Label(rw, text="●",   bg=C["bg_card"], fg=color, font=("Segoe UI",8)).pack(side="left", padx=8)
            tk.Label(rw, text=nome,  bg=C["bg_card"], fg=C["text_hi"], font=("Segoe UI",11,"bold"), width=10, anchor="w").pack(side="left")
            tk.Label(rw, text=sal,   bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL, width=14, anchor="w").pack(side="left")
            tk.Label(rw, text=jorn,  bg=C["bg_card"], fg=color, font=FONT_SMALL).pack(side="right")

        # Status
        rf = tk.Frame(info, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        rf.grid(row=0, column=1, padx=0, sticky="nsew")
        tk.Frame(rf, bg=C["success"], height=3).pack(fill="x")
        ri = tk.Frame(rf, bg=C["bg_card"], padx=20, pady=16); ri.pack(fill="both", expand=True)
        tk.Label(ri, text="Status do Sistema", bg=C["bg_card"], fg=C["text_hi"], font=FONT_HEADER).pack(anchor="w", pady=12)
        for sym, lbl, val, color in [
            ("✓","Banco de dados",      "Conectado",   C["success"]),
            ("✓","Cadastro de pessoal", "Operacional", C["success"]),
            ("✓","Importação de ponto", "Operacional", C["success"]),
            ("✓","Relatórios",          "Disponível",  C["success"]),
            ("✓","Cálculo de horas",    "Operacional", C["success"]),
            ("○","Folha de pagamento",  "Etapa 4",     C["warning"]),
        ]:
            rw = tk.Frame(ri, bg=C["bg_card"], pady=4); rw.pack(fill="x")
            tk.Label(rw, text=sym, bg=C["bg_card"], fg=color, font=("Segoe UI",10,"bold"), width=2).pack(side="left")
            tk.Label(rw, text=lbl, bg=C["bg_card"], fg=C["text_mid"], font=FONT_BODY, width=22, anchor="w").pack(side="left", padx=4)
            tk.Label(rw, text=val, bg=C["bg_card"], fg=color, font=FONT_SMALL).pack(side="right")


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — FUNCIONÁRIOS
# ──────────────────────────────────────────────────────────────────────────────
class PageFuncionarios(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db; self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Funcionários",         bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Cadastro e gestão de pessoal", bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Btn(hdr, "Novo Funcionário", self._form, icon="+").pack(side="right", pady=4)
        Sep(self).pack(fill="x", padx=36)

        tf = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=20); tf.pack(fill="both", expand=True)
        _tv_style()
        self.tree = ttk.Treeview(tf, columns=("nome","salario","tipo","jornada","obs","status"),
                                  show="headings", style="Dark.Treeview", selectmode="browse")
        for col,(lbl,w) in {"nome":("FUNCIONÁRIO",200),"salario":("SALÁRIO",140),
            "tipo":("TIPO PGTO",120),"jornada":("JORNADA",90),"obs":("OBSERVAÇÕES",280),"status":("STATUS",90)}.items():
            self.tree.heading(col, text=lbl)
            self.tree.column(col, width=w, anchor="center" if col in ("jornada","status") else "w")
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self._load()

        br = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); br.pack(fill="x")
        Btn(br, "Editar",          self._edit,   style="outline", icon="✎").pack(side="left", padx=10)
        Btn(br, "Ativar/Desativar",self._toggle, style="outline", icon="⏼").pack(side="left", padx=10)
        Btn(br, "Excluir",         self._delete, style="danger",  icon="✕").pack(side="left")

    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        for f in self.db.get_funcionarios(apenas_ativos=False):
            sal  = f"R$ {f['salario']:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            tipo = "Semanal" if f["tipo_pag"]=="semana" else "Mensal"
            jorn = f"{f['jornada_h']:.0f}h/dia"
            stat = "Ativo" if f["ativo"] else "Inativo"
            self.tree.insert("","end", iid=str(f["id"]),
                values=(f["nome"],sal,tipo,jorn,f["observacoes"] or "",stat),
                tags=("ativo" if f["ativo"] else "inativo",))
        self.tree.tag_configure("inativo", foreground=C["text_lo"])

    def _sel(self):
        s = self.tree.selection()
        if not s: messagebox.showwarning("Seleção","Selecione um funcionário."); return None
        return int(s[0])

    def _form(self, fid=None):   FormFuncionario(self, self.db, fid=fid, on_save=self._load)
    def _edit(self):
        fid = self._sel()
        if fid: self._form(fid)
    def _toggle(self):
        fid = self._sel()
        if fid: self.db.toggle_ativo(fid); self._load()
    def _delete(self):
        fid = self._sel()
        if not fid: return
        f = self.db.get_funcionario(fid)
        if messagebox.askyesno("Confirmar", f"Excluir '{f['nome']}' permanentemente?", icon="warning"):
            self.db.delete_funcionario(fid); self._load()

class FormFuncionario(tk.Toplevel):
    def __init__(self, parent, db, fid=None, on_save=None):
        super().__init__(parent)
        self.db, self.fid, self.on_save = db, fid, on_save
        self.title("Editar" if fid else "Novo Funcionário")
        self.configure(bg=C["bg_card"]); self.resizable(False,False); self.grab_set()
        sw,sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"520x530+{(sw-520)//2}+{(sh-530)//2}")
        self._build()
        if fid: self._pop()

    def _build(self):
        tk.Frame(self, bg=C["accent"], height=4).pack(fill="x")
        c = tk.Frame(self, bg=C["bg_card"], padx=32, pady=24); c.pack(fill="both", expand=True)
        tk.Label(c, text="Editar Funcionário" if self.fid else "Novo Funcionário",
                 bg=C["bg_card"], fg=C["text_hi"], font=FONT_HEADER).pack(anchor="w", pady=16)
        def fld(lbl):
            tk.Label(c, text=lbl, bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w", pady=2)
        fld("NOME COMPLETO"); self.vn = tk.StringVar()
        tk.Entry(c, textvariable=self.vn, bg=C["bg_input"], fg=C["text_hi"], insertbackground=C["accent"], font=FONT_BODY, relief="flat").pack(fill="x", ipady=10)
        tk.Frame(c, bg=C["border"], height=1).pack(fill="x")
        fld("SALÁRIO (R$)"); self.vs = tk.StringVar()
        tk.Entry(c, textvariable=self.vs, bg=C["bg_input"], fg=C["text_hi"], insertbackground=C["accent"], font=FONT_BODY, relief="flat").pack(fill="x", ipady=10)
        tk.Frame(c, bg=C["border"], height=1).pack(fill="x")
        fld("TIPO DE PAGAMENTO"); self.vt = tk.StringVar(value="mes")
        tf = tk.Frame(c, bg=C["bg_card"]); tf.pack(fill="x", pady=0)
        for val,lbl in [("mes","Mensal"),("semana","Semanal")]:
            tk.Radiobutton(tf, text=lbl, variable=self.vt, value=val, bg=C["bg_card"],
                fg=C["text_hi"], selectcolor=C["bg_input"], activebackground=C["bg_card"],
                activeforeground=C["accent"], font=FONT_BODY).pack(side="left", padx=24)
        fld("JORNADA DIÁRIA (HORAS)"); self.vj = tk.StringVar(value="8")
        tk.Entry(c, textvariable=self.vj, bg=C["bg_input"], fg=C["text_hi"], insertbackground=C["accent"], font=FONT_BODY, relief="flat").pack(fill="x", ipady=10)
        tk.Frame(c, bg=C["border"], height=1).pack(fill="x")
        fld("OBSERVAÇÕES"); self.vo = tk.StringVar()
        tk.Entry(c, textvariable=self.vo, bg=C["bg_input"], fg=C["text_hi"], insertbackground=C["accent"], font=FONT_BODY, relief="flat").pack(fill="x", ipady=10)
        tk.Frame(c, bg=C["border"], height=1).pack(fill="x")
        br = tk.Frame(c, bg=C["bg_card"]); br.pack(fill="x", pady=0)
        Btn(br, "Cancelar", self.destroy, style="outline").pack(side="left")
        Btn(br, "Salvar",   self._save,   style="primary").pack(side="right")

    def _pop(self):
        f = self.db.get_funcionario(self.fid)
        if f:
            self.vn.set(f["nome"]); self.vs.set(str(f["salario"]))
            self.vt.set(f["tipo_pag"]); self.vj.set(str(f["jornada_h"]))
            self.vo.set(f["observacoes"] or "")

    def _save(self):
        nome = self.vn.get().strip()
        if not nome: messagebox.showerror("Erro","Nome obrigatório.",parent=self); return
        try:    sal  = float(self.vs.get().replace(",","."))
        except: messagebox.showerror("Erro","Salário inválido.",parent=self); return
        try:    jorn = float(self.vj.get().replace(",","."))
        except: messagebox.showerror("Erro","Jornada inválida.",parent=self); return
        self.db.save_funcionario(
            {"nome":nome,"salario":sal,"tipo_pag":self.vt.get(),
             "jornada_h":jorn,"observacoes":self.vo.get().strip()}, fid=self.fid)
        if self.on_save: self.on_save()
        self.destroy()


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — IMPORTAR PONTO
# ──────────────────────────────────────────────────────────────────────────────
class PageImportacao(tk.Frame):
    def __init__(self, parent, db, app_ref=None, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db; self.app_ref = app_ref
        self.parser = PontoParser(); self._regs = []; self._avisos = []; self._path = ""
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Importar Ponto",    bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Selecione o arquivo exportado pelo relógio",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Btn(hdr, "Selecionar Arquivo", self._pick, style="primary", icon="📂").pack(side="right", pady=4)
        Sep(self).pack(fill="x", padx=36)

        fb = tk.Frame(self, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        fb.pack(fill="x", padx=36, pady=0)
        fi = tk.Frame(fb, bg=C["bg_card"], padx=20, pady=14); fi.pack(fill="x")
        tk.Label(fi, text="Arquivo:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(side="left")
        self.lbl_file = tk.Label(fi, text="nenhum selecionado", bg=C["bg_card"], fg=C["text_lo"], font=FONT_BODY)
        self.lbl_file.pack(side="left", padx=0)

        self.aviso_frame = tk.Frame(self, bg=C["bg_panel"])

        pf = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); pf.pack(fill="both", expand=True)
        tk.Label(pf, text="PRÉ-VISUALIZAÇÃO DOS DADOS LIDOS",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=8)
        tw = tk.Frame(pf, bg=C["bg_panel"]); tw.pack(fill="both", expand=True)
        _tv_style()
        self.tree = ttk.Treeview(tw, columns=("func","data","qtd","horarios","status"),
                                  show="headings", style="Dark.Treeview", selectmode="browse")
        for col,(lbl,w) in {"func":("FUNCIONÁRIO",180),"data":("DATA",110),
            "qtd":("BATIDAS",80),"horarios":("HORÁRIOS",380),"status":("STATUS",90)}.items():
            self.tree.heading(col, text=lbl)
            self.tree.column(col, width=w, anchor="center" if col in ("qtd","status") else "w")
        self.tree.tag_configure("ok",    foreground=C["text_hi"])
        self.tree.tag_configure("impar", foreground=C["warning"])
        self.tree.tag_configure("erro",  foreground=C["danger"])
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)

        bot = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); bot.pack(fill="x")
        self.lbl_count = tk.Label(bot, text="", bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY)
        self.lbl_count.pack(side="left")
        self.btn_imp = Btn(bot, "Confirmar Importação", self._confirm, style="success", icon="✓")
        self.btn_imp.pack(side="right")
        self.btn_imp.lbl.config(state="disabled", fg=C["text_lo"])

    def _pick(self):
        path = filedialog.askopenfilename(title="Selecionar arquivo",
            filetypes=[("Suportados","*.csv *.txt *.xls *.xlsx"),("Todos","*.*")])
        if not path: return
        self._path = path
        self.lbl_file.config(text=os.path.basename(path), fg=C["accent"])
        for r in self.tree.get_children(): self.tree.delete(r)
        self.aviso_frame.pack_forget()
        regs, avisos = self.parser.parse(path)
        self._regs, self._avisos = regs, avisos
        for reg in regs:
            qtd = reg["qtd_batidas"]; err = bool(reg.get("erros"))
            tag = "erro" if err else ("impar" if qtd%2!=0 else "ok")
            stat = "ERRO" if err else ("ÍMPAR" if qtd%2!=0 else "OK")
            self.tree.insert("","end", values=(reg["funcionario_raw"],_fd(reg["data"]),
                qtd, "  |  ".join(reg["horarios"]), stat), tags=(tag,))
        if avisos:
            self.aviso_frame.pack(fill="x", padx=36, pady=0)
            for w in self.aviso_frame.winfo_children(): w.destroy()
            box = tk.Frame(self.aviso_frame, bg="#2A1A0A",
                           highlightbackground=C["warning"], highlightthickness=1); box.pack(fill="x")
            tk.Frame(box, bg=C["warning"], height=2).pack(fill="x")
            inn = tk.Frame(box, bg="#2A1A0A", padx=16, pady=10); inn.pack(fill="x")
            tk.Label(inn, text="Avisos:", bg="#2A1A0A", fg=C["warning"], font=("Segoe UI",10,"bold")).pack(anchor="w")
            for av in avisos:
                tk.Label(inn, text=f"  • {av}", bg="#2A1A0A", fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
        n = len(regs)
        self.lbl_count.config(text=f"{n} registro(s) encontrado(s)")
        self.btn_imp.lbl.config(state="normal" if n>0 else "disabled",
                                 fg=C["bg_dark"] if n>0 else C["text_lo"])

    def _confirm(self):
        if not self._regs: return
        self.db.save_importacao(os.path.basename(self._path), self._regs, self._avisos)
        # ── Notificação de fechamento semanal ─────────────────────────────────
        if self.db.notif:
            self.db.notif.fechamento_semanal()
        messagebox.showinfo("Concluído", f"{len(self._regs)} registros importados!\nRedirecionando para Cálculo de Horas.")
        self._regs = []; self._path = ""
        self.lbl_file.config(text="nenhum selecionado", fg=C["text_lo"])
        for r in self.tree.get_children(): self.tree.delete(r)
        self.lbl_count.config(text="")
        self.btn_imp.lbl.config(state="disabled", fg=C["text_lo"])
        if self.app_ref: self.app_ref._show("calculo")


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — RELATÓRIOS IMPORTADOS
# ──────────────────────────────────────────────────────────────────────────────
class PageRelatorios(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db; self._imp_id = None; self._fnc_id = None
        self._all_imps = []; self._func_ids = {}
        self._build(); self._load_imps()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Relatórios Importados", bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Registros do relógio organizados por funcionário e data",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Btn(hdr, "Excluir Importação", self._del_imp, style="danger", icon="✕").pack(side="right", pady=4)
        Sep(self).pack(fill="x", padx=36)

        flt = tk.Frame(self, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        flt.pack(fill="x", padx=36, pady=0)
        tk.Frame(flt, bg=C["accent"], height=2).pack(fill="x")
        fi = tk.Frame(flt, bg=C["bg_card"], padx=20, pady=12); fi.pack(fill="x")
        tk.Label(fi, text="Importação:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(side="left")
        self.cb_imp = ttk.Combobox(fi, state="readonly", width=38, font=FONT_BODY)
        self.cb_imp.pack(side="left", padx=24)
        self.cb_imp.bind("<<ComboboxSelected>>", self._on_imp)
        tk.Label(fi, text="Funcionário:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(side="left")
        self.cb_func = ttk.Combobox(fi, state="readonly", width=20, font=FONT_BODY)
        self.cb_func.pack(side="left", padx=24)
        self.cb_func.bind("<<ComboboxSelected>>", self._on_func)
        Btn(fi, "Todos", self._todos, style="outline").pack(side="left")

        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=3); body.columnconfigure(1, weight=2)

        lw = tk.Frame(body, bg=C["bg_panel"]); lw.grid(row=0, column=0, sticky="nsew", padx=12)
        tk.Label(lw, text="REGISTROS POR DIA", bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        tw = tk.Frame(lw, bg=C["bg_panel"]); tw.pack(fill="both", expand=True)
        _tv_style()
        self.tree = ttk.Treeview(tw, columns=("func","data","qtd","horarios","status"),
                                  show="headings", style="Dark.Treeview", selectmode="browse")
        for col,(lbl,w) in {"func":("FUNCIONÁRIO",160),"data":("DATA",100),
            "qtd":("BATIDAS",70),"horarios":("HORÁRIOS",260),"status":("STATUS",80)}.items():
            self.tree.heading(col, text=lbl)
            self.tree.column(col, width=w, anchor="center" if col in ("qtd","status") else "w")
        self.tree.tag_configure("ok",    foreground=C["text_hi"])
        self.tree.tag_configure("impar", foreground=C["warning"])
        self.tree.tag_configure("erro",  foreground=C["danger"])
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sel)

        rw = tk.Frame(body, bg=C["bg_panel"]); rw.grid(row=0, column=1, sticky="nsew")
        tk.Label(rw, text="DETALHE DO DIA", bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        self.det_card = tk.Frame(rw, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        self.det_card.pack(fill="both", expand=True)
        tk.Frame(self.det_card, bg=C["accent"], height=3).pack(fill="x")
        self.det_inner = tk.Frame(self.det_card, bg=C["bg_card"], padx=24, pady=20)
        self.det_inner.pack(fill="both", expand=True)
        self._det_vazio()

        ft = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); ft.pack(fill="x")
        self.lbl_tot = tk.Label(ft, text="", bg=C["bg_panel"], fg=C["text_mid"], font=FONT_SMALL)
        self.lbl_tot.pack(side="left")

    def _load_imps(self):
        self._all_imps = self.db.get_importacoes()
        opts = ["(todas as importações)"] + [f"[{i['id']}]  {i['arquivo']}  |  {i['importado_em'][:16]}" for i in self._all_imps]
        self.cb_imp["values"] = opts
        if opts: self.cb_imp.current(0)
        funcs = self.db.get_funcionarios(apenas_ativos=False)
        self._func_ids = {f["nome"]: f["id"] for f in funcs}
        self.cb_func["values"] = ["(todos)"] + [f["nome"] for f in funcs]
        if self.cb_func["values"]: self.cb_func.current(0)
        self._load_bat()

    def _load_bat(self, imp_id=None, func_id=None):
        for r in self.tree.get_children(): self.tree.delete(r)
        rows = self.db.get_batidas(imp_id=imp_id, func_id=func_id)
        for row in rows:
            fn  = row["func_nome"] or row["funcionario_raw"]
            qtd = row["qtd_batidas"]; err = bool(row["tem_erro"])
            tag = "erro" if err else ("impar" if qtd%2!=0 else "ok")
            stat = "ERRO" if err else ("ÍMPAR" if qtd%2!=0 else "OK")
            hrs = "  |  ".join(row["horarios"].split(",")) if row["horarios"] else "—"
            self.tree.insert("","end", iid=str(row["id"]),
                values=(fn, _fd(row["data"]), qtd, hrs, stat), tags=(tag,))
        t = len(rows)
        e = sum(1 for r in rows if r["tem_erro"])
        ip = sum(1 for r in rows if r["qtd_batidas"]%2!=0 and not r["tem_erro"])
        self.lbl_tot.config(text=f"{t} registro(s)   |   {ip} ímpares   |   {e} erro(s)")
        self._det_vazio()

    def _on_imp(self, e=None):
        idx = self.cb_imp.current()
        self._imp_id = self._all_imps[idx-1]["id"] if idx > 0 else None
        self._load_bat(self._imp_id, self._fnc_id)
    def _on_func(self, e=None):
        idx = self.cb_func.current()
        self._fnc_id = None if idx==0 else self._func_ids.get(self.cb_func.get())
        self._load_bat(self._imp_id, self._fnc_id)
    def _todos(self):
        self._imp_id = self._fnc_id = None
        self.cb_imp.current(0); self.cb_func.current(0); self._load_bat()
    def _on_sel(self, e=None):
        sel = self.tree.selection()
        if not sel: return
        rows = self.db.get_batidas()
        row  = next((r for r in rows if r["id"]==int(sel[0])), None)
        if row: self._det_render(row)
    def _del_imp(self):
        if not self._imp_id: messagebox.showinfo("Info","Selecione uma importação específica."); return
        imp = self.db.get_importacao(self._imp_id)
        if messagebox.askyesno("Confirmar", f"Excluir '{imp['arquivo']}'?", icon="warning"):
            self.db.delete_importacao(self._imp_id); self._imp_id = None; self._load_imps()

    def _det_vazio(self):
        for w in self.det_inner.winfo_children(): w.destroy()
        tk.Label(self.det_inner, text="Selecione um registro\npara ver o detalhe",
                 bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",12), justify="center").pack(expand=True)

    def _det_render(self, row):
        for w in self.det_inner.winfo_children(): w.destroy()
        fn  = row["func_nome"] or row["funcionario_raw"]
        hrs = [h.strip() for h in row["horarios"].split(",") if h.strip()]
        qtd = row["qtd_batidas"]
        tk.Label(self.det_inner, text=fn,       bg=C["bg_card"], fg=C["text_hi"], font=("Segoe UI",15,"bold")).pack(anchor="w")
        tk.Label(self.det_inner, text=_fd(row["data"]), bg=C["bg_card"], fg=C["accent"], font=("Segoe UI",11)).pack(anchor="w", pady=16)
        Sep(self.det_inner).pack(fill="x", pady=14)
        LABELS = ["Entrada","Saída almoço","Retorno","Saída","Extra ent.","Extra saída"]
        for i, hora in enumerate(hrs):
            lbl = LABELS[i] if i < len(LABELS) else f"Batida {i+1}"
            color = C["accent"] if i%2==0 else C["success"]
            rf = tk.Frame(self.det_inner, bg=C["bg_card"], pady=5); rf.pack(fill="x")
            tk.Frame(rf, bg=color, width=3).pack(side="left", fill="y", padx=12)
            tk.Label(rf, text=lbl,  bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL, width=14, anchor="w").pack(side="left")
            tk.Label(rf, text=hora, bg=C["bg_card"], fg=C["text_hi"], font=("Segoe UI",14,"bold")).pack(side="right")
        Sep(self.det_inner).pack(fill="x", pady=8)
        tk.Label(self.det_inner, text=f"{qtd} batida(s)", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
        if qtd%2!=0:
            wb = tk.Frame(self.det_inner, bg="#2A1A0A", pady=8, padx=12); wb.pack(fill="x", pady=0)
            tk.Label(wb, text="Número ímpar de batidas", bg="#2A1A0A", fg=C["warning"], font=("Segoe UI",10,"bold")).pack(anchor="w")
            tk.Label(wb, text="Possível batida faltante ou esquecida.", bg="#2A1A0A", fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
        if row["erros"]:
            eb = tk.Frame(self.det_inner, bg="#200A0A", pady=8, padx=12); eb.pack(fill="x", pady=0)
            tk.Label(eb, text="Erro de leitura", bg="#200A0A", fg=C["danger"], font=("Segoe UI",10,"bold")).pack(anchor="w")
            tk.Label(eb, text=row["erros"], bg="#200A0A", fg=C["text_mid"], font=FONT_SMALL, wraplength=260, justify="left").pack(anchor="w")
        if not row["funcionario_id"]:
            ub = tk.Frame(self.det_inner, bg="#0A1520", pady=8, padx=12); ub.pack(fill="x", pady=0)
            tk.Label(ub, text="Funcionário não identificado", bg="#0A1520", fg=C["accent"], font=("Segoe UI",10,"bold")).pack(anchor="w")
            tk.Label(ub, text=f"ID: {row['funcionario_raw']}", bg="#0A1520", fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — CÁLCULO DE HORAS  (Etapa 3 — nova)
# ──────────────────────────────────────────────────────────────────────────────
class PageCalculo(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db     = db
        self.engine = CalcEngine()
        self._func_id   = None
        self._imp_id    = None
        self._all_funcs = []
        self._all_imps  = []
        self._dias_calc  = []   # resultado de calc_dia por linha selecionada
        self._build()
        self._load_filtros()

    # ── Construção da interface ───────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Cálculo de Horas", bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Jornada · Horas extras · Faltas · Saldo diário e semanal",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Btn(hdr, "Recalcular", self._recalc, style="primary", icon="↻").pack(side="right", pady=4)
        Btn(hdr, "Excel", self._exportar_excel, style="outline", icon="📊").pack(side="right", pady=4, padx=6)
        Btn(hdr, "PDF",   self._exportar_pdf,   style="outline", icon="📄").pack(side="right", pady=4, padx=4)
        Sep(self).pack(fill="x", padx=36)

        # Filtros
        flt = tk.Frame(self, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
        flt.pack(fill="x", padx=36, pady=0)
        tk.Frame(flt, bg=C["accent"], height=2).pack(fill="x")
        fi = tk.Frame(flt, bg=C["bg_card"], padx=20, pady=12); fi.pack(fill="x")

        tk.Label(fi, text="Funcionário:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(side="left")
        self.cb_func = ttk.Combobox(fi, state="readonly", width=22, font=FONT_BODY)
        self.cb_func.pack(side="left", padx=24)
        self.cb_func.bind("<<ComboboxSelected>>", lambda e: self._recalc())

        tk.Label(fi, text="Importação:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(side="left")
        self.cb_imp = ttk.Combobox(fi, state="readonly", width=34, font=FONT_BODY)
        self.cb_imp.pack(side="left", padx=24)
        self.cb_imp.bind("<<ComboboxSelected>>", lambda e: self._recalc())
        Btn(fi, "Tudo", self._reset_filtros, style="outline").pack(side="left")

        # Cards de totais
        self.cards_frame = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); self.cards_frame.pack(fill="x")
        self._build_cards()

        Sep(self).pack(fill="x", padx=36)

        # Corpo: tabela + detalhe
        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=3); body.columnconfigure(1, weight=2)

        # Tabela de dias
        lw = tk.Frame(body, bg=C["bg_panel"]); lw.grid(row=0, column=0, sticky="nsew", padx=14)
        tk.Label(lw, text="REGISTRO DIÁRIO",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        tw = tk.Frame(lw, bg=C["bg_panel"]); tw.pack(fill="both", expand=True)
        _tv_style()
        cols = ("func","data","dia","horarios","trab","extra","falta","saldo","alertas")
        self.tree = ttk.Treeview(tw, columns=cols, show="headings",
                                  style="Dark.Treeview", selectmode="browse")
        widths = {"func":130,"data":95,"dia":50,"horarios":200,"trab":75,"extra":75,"falta":75,"saldo":75,"alertas":180}
        labels = {"func":"FUNCIONÁRIO","data":"DATA","dia":"DIA","horarios":"HORÁRIOS",
                  "trab":"TRABALHADO","extra":"EXTRAS","falta":"FALTAS","saldo":"SALDO","alertas":"ALERTAS"}
        for col in cols:
            self.tree.heading(col, text=labels[col])
            self.tree.column(col, width=widths[col],
                anchor="center" if col in ("dia","trab","extra","falta","saldo") else "w")
        # Tags de cor
        self.tree.tag_configure("extra",   foreground=C["success"])
        self.tree.tag_configure("falta",   foreground=C["danger"])
        self.tree.tag_configure("ok",      foreground=C["text_hi"])
        self.tree.tag_configure("atencao", foreground=C["warning"])
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sel)

        # Painel de detalhe / saldo semanal
        rw = tk.Frame(body, bg=C["bg_panel"]); rw.grid(row=0, column=1, sticky="nsew")

        # Detalhe dia
        tk.Label(rw, text="DETALHE DO DIA",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        self.det_card = tk.Frame(rw, bg=C["bg_card"],
                                  highlightbackground=C["border"], highlightthickness=1)
        self.det_card.pack(fill="x")
        tk.Frame(self.det_card, bg=C["accent"], height=3).pack(fill="x")
        self.det_inner = tk.Frame(self.det_card, bg=C["bg_card"], padx=20, pady=16)
        self.det_inner.pack(fill="both", expand=True)
        self._det_vazio()

        # Saldo semanal
        tk.Label(rw, text="SALDO SEMANAL",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        self.sem_card = tk.Frame(rw, bg=C["bg_card"],
                                  highlightbackground=C["border"], highlightthickness=1)
        self.sem_card.pack(fill="both", expand=True)
        tk.Frame(self.sem_card, bg=C["warning"], height=3).pack(fill="x")
        self.sem_inner = tk.Frame(self.sem_card, bg=C["bg_card"], padx=20, pady=14)
        self.sem_inner.pack(fill="both", expand=True)
        tk.Label(self.sem_inner, text="Selecione um funcionário\npara ver o saldo semanal",
                 bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",11), justify="center").pack(expand=True)

        # Rodapé legenda
        leg = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=12); leg.pack(fill="x")
        for cor, txt in [(C["success"],"■ Saldo positivo / horas extras"),
                         (C["danger"], "■ Saldo negativo / faltas"),
                         (C["warning"],"■ Atenção / alertas")]:
            tk.Label(leg, text=txt, bg=C["bg_panel"], fg=cor, font=FONT_SMALL).pack(side="left", padx=24)

    def _build_cards(self):
        for w in self.cards_frame.winfo_children(): w.destroy()
        self.cards_frame.columnconfigure(0, weight=1)
        self.cards_frame.columnconfigure(1, weight=1)
        self.cards_frame.columnconfigure(2, weight=1)
        self.cards_frame.columnconfigure(3, weight=1)
        placeholders = [
            ("Total Trabalhado", "--:--", "horas no período",  C["accent"]),
            ("Horas Extras",     "--:--", "acima da jornada",  C["success"]),
            ("Horas Faltantes",  "--:--", "abaixo da jornada", C["danger"]),
            ("Saldo Geral",      "--:--", "positivo ou negativo", C["warning"]),
        ]
        self._card_labels = {}
        for i, (t, v, s, c) in enumerate(placeholders):
            card = tk.Frame(self.cards_frame, bg=C["bg_card"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.grid(row=0, column=i, padx=14, sticky="ew", pady=4)
            tk.Frame(card, bg=c, height=3).pack(fill="x")
            inn = tk.Frame(card, bg=C["bg_card"], padx=16, pady=12); inn.pack(fill="both", expand=True)
            tk.Label(inn, text=t, bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
            lv = tk.Label(inn, text=v, bg=C["bg_card"], fg=c, font=("Segoe UI",22,"bold"))
            lv.pack(anchor="w", pady=0)
            tk.Label(inn, text=s, bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w")
            self._card_labels[i] = (lv, c)

    # ── Dados ─────────────────────────────────────────────────────────────────
    def _load_filtros(self):
        self._all_funcs = self.db.get_funcionarios(apenas_ativos=False)
        self._all_imps  = self.db.get_importacoes()
        fn_opts  = ["(todos os funcionários)"] + [f["nome"] for f in self._all_funcs]
        imp_opts = ["(todas as importações)"]  + [f"[{i['id']}]  {i['arquivo']}  |  {i['importado_em'][:16]}" for i in self._all_imps]
        self.cb_func["values"] = fn_opts;  self.cb_func.current(0)
        self.cb_imp["values"]  = imp_opts; self.cb_imp.current(0)
        self._recalc()

    def _reset_filtros(self):
        self.cb_func.current(0); self.cb_imp.current(0); self._recalc()

    def _recalc(self):
        # Descobre filtros
        fi = self.cb_func.current()
        ii = self.cb_imp.current()
        func_id = self._all_funcs[fi-1]["id"] if fi > 0 else None
        imp_id  = self._all_imps[ii-1]["id"]  if ii > 0 else None

        # Busca batidas
        rows = self.db.get_batidas(imp_id=imp_id, func_id=func_id)

        # Para cada batida, pega a jornada do funcionário cadastrado
        func_jornada: dict[int, float] = {}
        for f in self._all_funcs:
            func_jornada[f["id"]] = f["jornada_h"]

        # Calcula dia a dia
        dias_calc = []
        for row in rows:
            if not row["horarios"]: continue
            hrs  = [h.strip() for h in row["horarios"].split(",") if h.strip()]
            fid  = row["funcionario_id"]
            jorn = func_jornada.get(fid, 8.0)   # default 8h se não vinculado
            calc = self.engine.calc_dia(row["data"], hrs, jorn)
            calc["_row"]      = row
            calc["_func_nome"]= row["func_nome"] or row["funcionario_raw"]
            dias_calc.append(calc)

        self._dias_calc = dias_calc

        # Atualiza tabela
        for r in self.tree.get_children(): self.tree.delete(r)
        for d in dias_calc:
            fn   = d["_func_nome"]
            alrt = " · ".join(d["alertas"]) if d["alertas"] else ""
            tag  = "falta"   if d["faltas_min"] > 0   else \
                   "atencao" if d["alertas"]           else \
                   "extra"   if d["extras_min"] > 0    else "ok"
            self.tree.insert("","end",
                values=(fn, _fd(d["data"]), d["dia_nome"],
                        "  ".join(d["horarios"]),
                        d["trabalhados_fmt"], d["extras_fmt"],
                        d["faltas_fmt"], d["saldo_fmt"], alrt),
                tags=(tag,))

        # Atualiza cards de totais
        totais = self.engine.calc_periodo(dias_calc) if dias_calc else None
        vals = [
            (totais["trabalhados_fmt"] if totais else "--:--", C["accent"]),
            (totais["extras_fmt"]      if totais else "--:--", C["success"]),
            (totais["faltas_fmt"]      if totais else "--:--", C["danger"]),
            (totais["saldo_fmt"]       if totais else "--:--",
             C["success"] if totais and totais["saldo_min"]>=0 else C["danger"]),
        ]
        for i, (v, c) in enumerate(vals):
            lbl, _ = self._card_labels[i]
            lbl.config(text=v, fg=c)

        # Atualiza saldo semanal
        self._render_saldo_semanal(totais)
        self._det_vazio()

    # ── Detalhe do dia (painel direito) ───────────────────────────────────────
    def _on_sel(self, e=None):
        sel = self.tree.selection()
        if not sel: return
        idx = self.tree.index(sel[0])
        if 0 <= idx < len(self._dias_calc):
            self._det_render(self._dias_calc[idx])

    def _det_vazio(self):
        for w in self.det_inner.winfo_children(): w.destroy()
        tk.Label(self.det_inner, text="Selecione um dia\npara ver o detalhe",
                 bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",11), justify="center").pack(expand=True)

    def _det_render(self, d):
        for w in self.det_inner.winfo_children(): w.destroy()

        # Nome + data
        tk.Label(self.det_inner, text=d["_func_nome"],
                 bg=C["bg_card"], fg=C["text_hi"], font=("Segoe UI",13,"bold")).pack(anchor="w")
        tk.Label(self.det_inner, text=f"{_fd(d['data'])}  —  {d['dia_nome']}",
                 bg=C["bg_card"], fg=C["accent"], font=FONT_BODY).pack(anchor="w", pady=12)
        Sep(self.det_inner).pack(fill="x", pady=12)

        # Grade de métricas 2×2
        mg = tk.Frame(self.det_inner, bg=C["bg_card"]); mg.pack(fill="x", pady=12)
        mg.columnconfigure(0, weight=1); mg.columnconfigure(1, weight=1)
        metrics = [
            ("Trabalhado",   d["trabalhados_fmt"], C["accent"],   0, 0),
            ("Jornada",      d["jornada_fmt"],     C["text_mid"], 0, 1),
            ("Extras",       d["extras_fmt"],      C["success"],  1, 0),
            ("Faltas",       d["faltas_fmt"],      C["danger"],   1, 1),
        ]
        for lbl, val, color, row, col in metrics:
            mf = tk.Frame(mg, bg=C["bg_input"], padx=12, pady=10)
            mf.grid(row=row, column=col, padx=(0 if col>0 else 0, 6 if col==0 else 0),
                    pady=6, sticky="ew")
            tk.Label(mf, text=lbl, bg=C["bg_input"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w")
            tk.Label(mf, text=val, bg=C["bg_input"], fg=color, font=("Segoe UI",16,"bold")).pack(anchor="w")

        # Saldo grande
        saldo_cor = C["success"] if d["saldo_min"] >= 0 else C["danger"]
        saldo_bg  = C["extra_bg"] if d["saldo_min"] >= 0 else C["falta_bg"]
        sf = tk.Frame(self.det_inner, bg=saldo_bg, pady=10, padx=12); sf.pack(fill="x", pady=12)
        tk.Label(sf, text="SALDO DO DIA", bg=saldo_bg, fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")
        tk.Label(sf, text=d["saldo_fmt"], bg=saldo_bg, fg=saldo_cor, font=FONT_BIG).pack(anchor="w")

        # Batidas com rótulos
        if d["horarios"]:
            Sep(self.det_inner).pack(fill="x", pady=10)
            LABELS = ["Entrada","Saída almoço","Retorno","Saída","Extra ent.","Extra saída","Bat.7","Bat.8"]
            for i, hora in enumerate(d["horarios"]):
                lbl   = LABELS[i] if i < len(LABELS) else f"Batida {i+1}"
                color = C["accent"] if i%2==0 else C["success"]
                rf = tk.Frame(self.det_inner, bg=C["bg_card"], pady=3); rf.pack(fill="x")
                tk.Frame(rf, bg=color, width=3).pack(side="left", fill="y", padx=10)
                tk.Label(rf, text=lbl,  bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL, width=14, anchor="w").pack(side="left")
                tk.Label(rf, text=hora, bg=C["bg_card"], fg=C["text_hi"],  font=("Segoe UI",13,"bold")).pack(side="right")

        # Alertas
        if d["alertas"]:
            Sep(self.det_inner).pack(fill="x", pady=8)
            af = tk.Frame(self.det_inner, bg=C["atencao_bg"], pady=8, padx=12); af.pack(fill="x")
            tk.Label(af, text="Alertas", bg=C["atencao_bg"], fg=C["warning"], font=("Segoe UI",10,"bold")).pack(anchor="w")
            for al in d["alertas"]:
                tk.Label(af, text=f"  • {al}", bg=C["atencao_bg"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")

    # ── Saldo semanal ─────────────────────────────────────────────────────────
    def _render_saldo_semanal(self, totais):
        for w in self.sem_inner.winfo_children(): w.destroy()
        if not totais or not totais.get("semanas"):
            tk.Label(self.sem_inner, text="Sem dados de semanas",
                     bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",11), justify="center").pack(expand=True)
            return
        semanas = sorted(totais["semanas"].items())
        for sem, saldo in semanas:
            # Formata "2025-W20" → "Semana 20 / 2025"
            try:
                partes = sem.split("-W")
                sem_label = f"Sem. {partes[1]} / {partes[0]}"
            except Exception:
                sem_label = sem
            cor = C["success"] if saldo >= 0 else C["danger"]
            rw  = tk.Frame(self.sem_inner, bg=C["bg_card"], pady=7); rw.pack(fill="x")
            tk.Label(rw, text=sem_label, bg=C["bg_card"], fg=C["text_mid"],
                     font=FONT_BODY, width=14, anchor="w").pack(side="left")
            # Barra de saldo proporcional
            bar_max = 480   # 8h em pixels relativos
            bar_px  = min(abs(saldo) * 80 // 60, 100)
            tk.Frame(rw, bg=cor, height=16, width=max(4, bar_px)).pack(side="left", padx=8)
            tk.Label(rw, text=_saldo_fmt(saldo), bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",11,"bold")).pack(side="right")

    # ── Exportação (Etapa 5) ──────────────────────────────────────────────────
    def _filtros_atuais(self) -> dict:
        fi = self.cb_func.current()
        ii = self.cb_imp.current()
        nome_func = self._all_funcs[fi-1]["nome"] if fi > 0 else "Todos"
        nome_imp  = (self._all_imps[ii-1]["arquivo"] if ii > 0
                     else "Todas as importações")
        return {"funcionario": nome_func, "importacao": nome_imp}

    def _exportar_pdf(self):
        if not _RELATORIOS_OK:
            messagebox.showwarning(
                "Módulo ausente",
                "O módulo relatorios_pdf.py não foi encontrado.\n"
                "Coloque-o na mesma pasta que ponto.py e tente novamente."
            ); return
        if not self._dias_calc:
            messagebox.showinfo("Sem dados", "Recalcule primeiro para ter dados a exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            title="Salvar relatório PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
            initialfile=f"calculo_horas_{datetime.now():%Y%m%d_%H%M}.pdf",
        )
        if not caminho: return
        try:
            gerar_pdf_calculo(self._dias_calc, caminho, self._filtros_atuais())
            messagebox.showinfo("PDF gerado", f"Arquivo salvo em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro ao gerar PDF", str(e))

    def _exportar_excel(self):
        if not _RELATORIOS_OK:
            messagebox.showwarning(
                "Módulo ausente",
                "O módulo relatorios_pdf.py não foi encontrado.\n"
                "Coloque-o na mesma pasta que ponto.py e tente novamente."
            ); return
        if not self._dias_calc:
            messagebox.showinfo("Sem dados", "Recalcule primeiro para ter dados a exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            title="Salvar planilha Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"calculo_horas_{datetime.now():%Y%m%d_%H%M}.xlsx",
        )
        if not caminho: return
        try:
            gerar_excel_calculo(self._dias_calc, caminho, self._filtros_atuais())
            messagebox.showinfo("Excel gerado", f"Arquivo salvo em:\n{caminho}")
        except ImportError:
            messagebox.showwarning("openpyxl ausente",
                "Execute:  pip install openpyxl\ne tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro ao gerar Excel", str(e))


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — BANCO DE HORAS  (Etapa 4 — nova)
# ──────────────────────────────────────────────────────────────────────────────
class PageBancoHoras(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db
        self._all_funcs = []
        self._sel_func_id = None
        self._build()
        self._load()

    # ── Interface ─────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Banco de Horas", bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Saldo acumulado de horas extras · Registrar retiradas",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Btn(hdr, "↻  Atualizar", self._load, style="outline").pack(side="right", pady=4)
        Btn(hdr, "Excel", self._exportar_excel, style="outline", icon="📊").pack(side="right", pady=4, padx=4)
        Btn(hdr, "PDF",   self._exportar_pdf,   style="outline", icon="📄").pack(side="right", pady=4, padx=4)
        Sep(self).pack(fill="x", padx=36)

        # Corpo dividido: lista funcionários (esq) + detalhe (dir)
        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=20); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=2); body.columnconfigure(1, weight=3)

        # ── Coluna esquerda: tabela de saldos ─────────────────────────────────
        lw = tk.Frame(body, bg=C["bg_panel"]); lw.grid(row=0, column=0, sticky="nsew", padx=16)
        tk.Label(lw, text="SALDO POR FUNCIONÁRIO",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=8)

        tw = tk.Frame(lw, bg=C["bg_panel"]); tw.pack(fill="both", expand=True)
        _tv_style()
        self.tree = ttk.Treeview(tw, columns=("nome","saldo","atualizado"),
                                  show="headings", style="Dark.Treeview", selectmode="browse")
        self.tree.heading("nome",       text="FUNCIONÁRIO")
        self.tree.heading("saldo",      text="SALDO")
        self.tree.heading("atualizado", text="ATUALIZADO EM")
        self.tree.column("nome",       width=160, anchor="w")
        self.tree.column("saldo",      width=100, anchor="center")
        self.tree.column("atualizado", width=140, anchor="center")
        self.tree.tag_configure("positivo", foreground=C["success"])
        self.tree.tag_configure("negativo", foreground=C["danger"])
        self.tree.tag_configure("zero",     foreground=C["text_mid"])
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sel)

        # ── Coluna direita: detalhe + retirada ────────────────────────────────
        rw = tk.Frame(body, bg=C["bg_panel"]); rw.grid(row=0, column=1, sticky="nsew")

        # Card de saldo destaque
        tk.Label(rw, text="SALDO ATUAL",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        self.saldo_card = tk.Frame(rw, bg=C["bg_card"],
                                    highlightbackground=C["border"], highlightthickness=1)
        self.saldo_card.pack(fill="x")
        self._bar_saldo = tk.Frame(self.saldo_card, bg=C["accent"], height=3); self._bar_saldo.pack(fill="x")
        si = tk.Frame(self.saldo_card, bg=C["bg_card"], padx=24, pady=20); si.pack(fill="x")
        self._lbl_func  = tk.Label(si, text="Selecione um funcionário", bg=C["bg_card"],
                                    fg=C["text_mid"], font=("Segoe UI",12)); self._lbl_func.pack(anchor="w")
        self._lbl_saldo = tk.Label(si, text="--:--", bg=C["bg_card"],
                                    fg=C["accent"], font=("Segoe UI",36,"bold")); self._lbl_saldo.pack(anchor="w", pady=0)
        self._lbl_sub   = tk.Label(si, text="", bg=C["bg_card"],
                                    fg=C["text_lo"], font=FONT_SMALL); self._lbl_sub.pack(anchor="w")

        # Formulário de retirada
        Sep(rw).pack(fill="x", pady=12)
        tk.Label(rw, text="REGISTRAR RETIRADA",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        ret_card = tk.Frame(rw, bg=C["bg_card"],
                             highlightbackground=C["border"], highlightthickness=1)
        ret_card.pack(fill="x")
        tk.Frame(ret_card, bg=C["warning"], height=3).pack(fill="x")
        ri = tk.Frame(ret_card, bg=C["bg_card"], padx=20, pady=16); ri.pack(fill="x")

        # Linha 1: horas e minutos
        r1 = tk.Frame(ri, bg=C["bg_card"]); r1.pack(fill="x", pady=10)
        tk.Label(r1, text="Horas:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_BODY, width=8, anchor="w").pack(side="left")
        self.ent_horas = tk.Entry(r1, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                   insertbackground=C["accent"], width=6,
                                   highlightbackground=C["border"], highlightthickness=1)
        self.ent_horas.pack(side="left", padx=16)
        self.ent_horas.insert(0, "0")
        tk.Label(r1, text="Min:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_BODY, width=5, anchor="w").pack(side="left")
        self.ent_min = tk.Entry(r1, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                 insertbackground=C["accent"], width=6,
                                 highlightbackground=C["border"], highlightthickness=1)
        self.ent_min.pack(side="left")
        self.ent_min.insert(0, "0")

        # Linha 2: descrição
        r2 = tk.Frame(ri, bg=C["bg_card"]); r2.pack(fill="x", pady=14)
        tk.Label(r2, text="Motivo:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_BODY, width=8, anchor="w").pack(side="left")
        self.ent_desc = tk.Entry(r2, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                  insertbackground=C["accent"],
                                  highlightbackground=C["border"], highlightthickness=1)
        self.ent_desc.pack(side="left", fill="x", expand=True)

        Btn(ri, "Confirmar Retirada", self._confirmar_retirada, style="warning", icon="−").pack(anchor="w")

        # PIN do funcionário
        Sep(rw).pack(fill="x", pady=12)
        tk.Label(rw, text="PIN DO FUNCIONÁRIO (para consulta no celular)",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        pin_card = tk.Frame(rw, bg=C["bg_card"],
                             highlightbackground=C["border"], highlightthickness=1)
        pin_card.pack(fill="x")
        tk.Frame(pin_card, bg=C["purple"], height=3).pack(fill="x")
        pi = tk.Frame(pin_card, bg=C["bg_card"], padx=20, pady=14); pi.pack(fill="x")
        pr = tk.Frame(pi, bg=C["bg_card"]); pr.pack(fill="x")
        tk.Label(pr, text="PIN:", bg=C["bg_card"], fg=C["text_mid"], font=FONT_BODY, width=8, anchor="w").pack(side="left")
        self.ent_pin = tk.Entry(pr, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                 insertbackground=C["accent"], width=10, show="*",
                                 highlightbackground=C["border"], highlightthickness=1)
        self.ent_pin.pack(side="left", padx=10)
        Btn(pr, "Salvar PIN", self._salvar_pin, style="outline").pack(side="left")
        self._lbl_pin_info = tk.Label(pi, text="", bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL)
        self._lbl_pin_info.pack(anchor="w", pady=0)

        # Histórico de movimentos
        Sep(rw).pack(fill="x", pady=12)
        tk.Label(rw, text="HISTÓRICO",
                 bg=C["bg_panel"], fg=C["text_mid"], font=("Segoe UI",9,"bold")).pack(anchor="w", pady=6)
        hist_fr = tk.Frame(rw, bg=C["bg_card"],
                            highlightbackground=C["border"], highlightthickness=1)
        hist_fr.pack(fill="both", expand=True)
        tk.Frame(hist_fr, bg=C["accent_dim"], height=3).pack(fill="x")
        self.hist_inner = tk.Frame(hist_fr, bg=C["bg_card"], padx=16, pady=12)
        self.hist_inner.pack(fill="both", expand=True)
        tk.Label(self.hist_inner, text="Selecione um funcionário",
                 bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",11)).pack(expand=True)

    # ── Dados ─────────────────────────────────────────────────────────────────
    def _load(self):
        self._all_funcs = list(self.db.get_banco_horas())
        for r in self.tree.get_children(): self.tree.delete(r)
        for row in self._all_funcs:
            saldo = row["saldo_min"]
            saldo_txt = _saldo_fmt(saldo)
            atu = row["atualizado_em"][:16] if row["atualizado_em"] else "-"
            tag = "positivo" if saldo > 0 else ("negativo" if saldo < 0 else "zero")
            self.tree.insert("", "end", iid=str(row["id"]),
                             values=(row["nome"], saldo_txt, atu), tags=(tag,))

    def _on_sel(self, e=None):
        sel = self.tree.selection()
        if not sel: return
        fid = int(sel[0])
        self._sel_func_id = fid
        # Acha nome e saldo
        row = next((r for r in self._all_funcs if r["id"] == fid), None)
        if not row: return
        saldo = row["saldo_min"]
        saldo_cor = C["success"] if saldo >= 0 else C["danger"]
        self._bar_saldo.config(bg=saldo_cor)
        self._lbl_func.config(text=row["nome"], fg=C["text_hi"])
        self._lbl_saldo.config(text=_saldo_fmt(saldo), fg=saldo_cor)
        horas = abs(saldo) // 60; mins = abs(saldo) % 60
        sub = f"{'Positivo' if saldo>=0 else 'Negativo'} — {horas}h {mins}min acumulados"
        self._lbl_sub.config(text=sub)
        # PIN atual
        self.ent_pin.delete(0, "end")
        pin = self.db.get_pin_func(fid)
        if pin:
            self.ent_pin.insert(0, pin)
            self._lbl_pin_info.config(text=f"PIN atual: {'*'*len(pin)} ({len(pin)} dígitos)")
        else:
            self._lbl_pin_info.config(text="Sem PIN cadastrado")
        # Histórico
        _, movs = self.db.get_banco_horas_func(fid)
        self._render_historico(movs)

    def _render_historico(self, movs):
        for w in self.hist_inner.winfo_children(): w.destroy()
        if not movs:
            tk.Label(self.hist_inner, text="Sem movimentações ainda",
                     bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",10)).pack(expand=True)
            return
        canvas = tk.Canvas(self.hist_inner, bg=C["bg_card"], highlightthickness=0)
        sb = ttk.Scrollbar(self.hist_inner, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=C["bg_card"])
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        for mov in movs:
            is_cred = mov["tipo"] == "credito"
            cor  = C["success"] if is_cred else C["warning"]
            icon = "+" if is_cred else "−"
            mins = mov["minutos"]
            val  = f"{icon}{mins//60:02d}:{mins%60:02d}"
            rw = tk.Frame(inner, bg=C["bg_card"], pady=5); rw.pack(fill="x")
            tk.Frame(rw, bg=cor, width=3).pack(side="left", fill="y", padx=10)
            lb = tk.Frame(rw, bg=C["bg_card"]); lb.pack(side="left", fill="x", expand=True)
            tk.Label(lb, text=mov["descricao"] or "—", bg=C["bg_card"],
                     fg=C["text_hi"], font=FONT_SMALL, anchor="w").pack(anchor="w")
            dt = mov["criado_em"][:16] if mov["criado_em"] else ""
            tk.Label(lb, text=dt, bg=C["bg_card"],
                     fg=C["text_lo"], font=("Segoe UI",8), anchor="w").pack(anchor="w")
            tk.Label(rw, text=val, bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",12,"bold")).pack(side="right", padx=6)
            Sep(inner).pack(fill="x")

    # ── Ações ─────────────────────────────────────────────────────────────────
    def _confirmar_retirada(self):
        if not self._sel_func_id:
            messagebox.showwarning("Atenção", "Selecione um funcionário primeiro.")
            return
        try:
            h = int(self.ent_horas.get() or 0)
            m = int(self.ent_min.get()   or 0)
        except ValueError:
            messagebox.showerror("Erro", "Horas e minutos devem ser números inteiros.")
            return
        total_min = h * 60 + m
        if total_min <= 0:
            messagebox.showwarning("Atenção", "Informe um valor maior que zero.")
            return
        desc = self.ent_desc.get().strip() or "Retirada aprovada pelo gestor"
        # Checa saldo disponível
        row = next((r for r in self._all_funcs if r["id"] == self._sel_func_id), None)
        saldo_atual = row["saldo_min"] if row else 0
        nome = row["nome"] if row else "funcionário"
        if total_min > saldo_atual:
            resp = messagebox.askyesno(
                "Saldo insuficiente",
                f"{nome} tem apenas {_fmt(saldo_atual)} no banco.\n"
                f"Você quer retirar {_fmt(total_min)} mesmo assim?\n"
                f"O saldo ficará negativo."
            )
            if not resp: return
        ok = messagebox.askyesno(
            "Confirmar retirada",
            f"Retirar {_fmt(total_min)} do banco de {nome}?\nMotivo: {desc}"
        )
        if not ok: return
        self.db.debitar_banco_horas(self._sel_func_id, total_min, desc)
        messagebox.showinfo("Sucesso", f"Retirada de {_fmt(total_min)} registrada para {nome}.")
        self.ent_horas.delete(0, "end"); self.ent_horas.insert(0, "0")
        self.ent_min.delete(0, "end");   self.ent_min.insert(0, "0")
        self.ent_desc.delete(0, "end")
        self._load()
        # Reseleciona o mesmo funcionário
        try: self.tree.selection_set(str(self._sel_func_id)); self._on_sel()
        except Exception: pass

    def _salvar_pin(self):
        if not self._sel_func_id:
            messagebox.showwarning("Atenção", "Selecione um funcionário primeiro.")
            return
        pin = self.ent_pin.get().strip()
        if not pin.isdigit() or len(pin) < 4:
            messagebox.showerror("Erro", "O PIN deve ter pelo menos 4 dígitos numéricos.")
            return
        self.db.set_pin_func(self._sel_func_id, pin)
        self._lbl_pin_info.config(text=f"PIN salvo: {'*'*len(pin)} ({len(pin)} dígitos)")
        messagebox.showinfo("Sucesso", "PIN atualizado com sucesso!")

    # ── Exportação (Etapa 5) ──────────────────────────────────────────────────
    def _dados_para_exportar(self) -> list:
        """Monta lista de dicts com saldo + histórico de todos os funcionários."""
        dados = []
        for f in self._all_funcs:
            fid = f["id"]
            saldo_min, movs = self.db.get_banco_horas_func(fid)
            historico = []
            for m in movs:
                is_cred = m["tipo"] == "credito"
                historico.append({
                    "tipo":      m["tipo"],
                    "valor_fmt": ("+" if is_cred else "−") + _fmt(m["minutos"]),
                    "descricao": m["descricao"] or "",
                    "data":      (m["criado_em"] or "")[:16],
                })
            dados.append({
                "nome":         f["nome"],
                "saldo_min":    saldo_min,
                "atualizado_em": "",
                "historico":    historico,
            })
        return dados

    def _exportar_pdf(self):
        if not _RELATORIOS_OK:
            messagebox.showwarning(
                "Módulo ausente",
                "O módulo relatorios_pdf.py não foi encontrado.\n"
                "Coloque-o na mesma pasta que ponto.py e tente novamente."
            ); return
        dados = self._dados_para_exportar()
        if not dados:
            messagebox.showinfo("Sem dados", "Nenhum funcionário encontrado."); return
        caminho = filedialog.asksaveasfilename(
            title="Salvar Banco de Horas PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
            initialfile=f"banco_horas_{datetime.now():%Y%m%d_%H%M}.pdf",
        )
        if not caminho: return
        try:
            gerar_pdf_banco_horas(dados, caminho)
            messagebox.showinfo("PDF gerado", f"Arquivo salvo em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro ao gerar PDF", str(e))

    def _exportar_excel(self):
        if not _RELATORIOS_OK:
            messagebox.showwarning(
                "Módulo ausente",
                "O módulo relatorios_pdf.py não foi encontrado.\n"
                "Coloque-o na mesma pasta que ponto.py e tente novamente."
            ); return
        dados = self._dados_para_exportar()
        if not dados:
            messagebox.showinfo("Sem dados", "Nenhum funcionário encontrado."); return
        caminho = filedialog.asksaveasfilename(
            title="Salvar Banco de Horas Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"banco_horas_{datetime.now():%Y%m%d_%H%M}.xlsx",
        )
        if not caminho: return
        try:
            gerar_excel_banco_horas(dados, caminho)
            messagebox.showinfo("Excel gerado", f"Arquivo salvo em:\n{caminho}")
        except ImportError:
            messagebox.showwarning("openpyxl ausente",
                "Execute:  pip install openpyxl\ne tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro ao gerar Excel", str(e))


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────────────────────────
class PageConfiguracoes(tk.Frame):
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db; self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=30); hdr.pack(fill="x")
        tk.Label(hdr, text="Configurações", bg=C["bg_panel"], fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(hdr, text="Parâmetros do sistema", bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=0)
        Sep(self).pack(fill="x", padx=36)

        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=24); body.pack(fill="both", expand=True)

        # ── Seção: Sincronização com a Nuvem ─────────────────────────────
        tk.Label(body, text="☁  Sincronização com a Nuvem", bg=C["bg_panel"],
                 fg=C["accent"], font=("Segoe UI",10,"bold")).pack(anchor="w", pady=8)

        cloud_card = tk.Frame(body, bg=C["bg_card"],
                              highlightbackground=C["border"], highlightthickness=1)
        cloud_card.pack(fill="x")

        cfg = _cfg_load()

        # URL do servidor
        row_url = tk.Frame(cloud_card, bg=C["bg_card"], padx=20, pady=12); row_url.pack(fill="x")
        tk.Label(row_url, text="URL do Servidor (Railway)", bg=C["bg_card"],
                 fg=C["text_mid"], font=FONT_BODY, width=26, anchor="w").pack(side="left")
        self._url_var = tk.StringVar(value=cfg["servidor_url"])
        ent_url = tk.Entry(row_url, textvariable=self._url_var, bg=C["bg_input"],
                           fg=C["text_hi"], font=FONT_MONO, insertbackground=C["accent"],
                           relief="flat", width=46)
        ent_url.pack(side="left", padx=8)
        tk.Label(row_url, text="ex: https://sistema-ponto-production.up.railway.app",
                 bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL).pack(side="left")

        Sep(cloud_card).pack(fill="x", padx=20)

        # Admin Key
        row_key = tk.Frame(cloud_card, bg=C["bg_card"], padx=20, pady=12); row_key.pack(fill="x")
        tk.Label(row_key, text="ADMIN_KEY (opcional)", bg=C["bg_card"],
                 fg=C["text_mid"], font=FONT_BODY, width=26, anchor="w").pack(side="left")
        self._key_var = tk.StringVar(value=cfg["admin_key"])
        ent_key = tk.Entry(row_key, textvariable=self._key_var, bg=C["bg_input"],
                           fg=C["text_hi"], font=FONT_MONO, insertbackground=C["accent"],
                           relief="flat", width=46, show="●")
        ent_key.pack(side="left", padx=8)
        tk.Label(row_key, text="Deve ser a mesma do Railway → Variables → ADMIN_KEY",
                 bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL).pack(side="left")

        Sep(cloud_card).pack(fill="x", padx=20)

        # Botões + status
        row_btn = tk.Frame(cloud_card, bg=C["bg_card"], padx=20, pady=14); row_btn.pack(fill="x")
        self._sync_status = tk.Label(row_btn, text="", bg=C["bg_card"],
                                     fg=C["success"], font=FONT_SMALL)
        self._sync_status.pack(side="right", padx=0)

        Btn(row_btn, "Sincronizar Agora", cmd=self._sync_manual,
            style="primary", icon="↑").pack(side="right")
        Btn(row_btn, "Salvar URL / Chave", cmd=self._salvar_cfg,
            style="secondary", icon="💾").pack(side="right", padx=8)

        # Aviso requests
        if not _REQUESTS_OK:
            warn = tk.Frame(cloud_card, bg="#2a1a00", padx=20, pady=10); warn.pack(fill="x")
            tk.Label(warn,
                     text="⚠  Biblioteca 'requests' não instalada.  Execute:  pip install requests",
                     bg="#2a1a00", fg="#ffaa00", font=FONT_SMALL).pack(anchor="w")

        # ── Seção: Sistema ────────────────────────────────────────────────
        for title, items in [
            ("Sistema", [
                ("Versão",         APP_VERSION),
                ("Banco de dados", os.path.abspath(DB_FILE)),
            ]),
            ("Jornada Padrão", [
                ("Seg–Sex entrada",   "07:00"),
                ("Seg–Sex saída",     "17:00  (com almoço 12:00–14:00)"),
                ("Sábado entrada",    "07:00"),
                ("Sábado saída",      "11:00"),
                ("Tolerância",        "5 minutos"),
            ]),
            ("Etapas Concluídas", [
                ("✓ Etapa 1", "Estrutura, DB, Cadastro de Funcionários"),
                ("✓ Etapa 2", "Importação CSV / TXT / XLS / XLSX"),
                ("✓ Etapa 3", "Cálculo de jornada, extras, faltas e saldo"),
                ("✓ Etapa 4", "Banco de Horas acumulado + retiradas + PIN"),
                ("✓ Etapa 5", "Relatórios PDF e exportação Excel"),
                ("✓ Etapa 6", "Deploy na nuvem (Railway + PostgreSQL)"),
                ("✓ Etapa 7", "Sincronização automática PC → Nuvem → App"),
            ]),
        ]:
            tk.Label(body, text=title, bg=C["bg_panel"],
                     fg=C["accent"], font=("Segoe UI",10,"bold")).pack(anchor="w", pady=8)
            card = tk.Frame(body, bg=C["bg_card"], highlightbackground=C["border"], highlightthickness=1)
            card.pack(fill="x")
            for lbl, val in items:
                row = tk.Frame(card, bg=C["bg_card"], padx=20, pady=10); row.pack(fill="x")
                tk.Label(row, text=lbl, bg=C["bg_card"], fg=C["text_mid"],
                         font=FONT_BODY, width=26, anchor="w").pack(side="left")
                tk.Label(row, text=val, bg=C["bg_card"], fg=C["text_hi"], font=FONT_MONO).pack(side="left")
                Sep(card).pack(fill="x", padx=20)

    def _salvar_cfg(self):
        _cfg_save(self._url_var.get(), self._key_var.get())
        self._sync_status.config(text="✔ Configuração salva!", fg=C["success"])
        self.after(3000, lambda: self._sync_status.config(text=""))

    def _sync_manual(self):
        _cfg_save(self._url_var.get(), self._key_var.get())
        self._sync_status.config(text="Sincronizando…", fg=C["text_mid"])
        self.update_idletasks()

        def _run():
            ok, msg = sincronizar_nuvem(self.db)
            color = C["success"] if ok else C["danger"]
            self._sync_status.config(text=msg, fg=color)
            self.after(6000, lambda: self._sync_status.config(text=""))

        self.after(50, _run)


# ──────────────────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────
# ETAPA 4.3 — ABA FINANCEIRO
# ──────────────────────────────────────────────────────────────────────────────

class PageFinanceiro(tk.Frame):
    """Aba Financeiro — acompanhamento individual de horas extras em dinheiro."""

    # Cores extras usadas só aqui
    _GOLD   = "#FFD166"
    _TEAL   = "#06D6A0"
    _CORAL  = "#EF476F"
    _INDIGO = "#7B5EA7"

    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db
        self._sel_id   = None
        self._dados    = {}       # cache do funcionário selecionado
        self._mult_var = tk.DoubleVar(value=1.5)   # multiplicador hora extra
        self._build()
        self._load_funcionarios()

    # ── DB helpers (migração segura) ──────────────────────────────────────────
    def _ensure_tables(self):
        """Cria tabelas específicas do financeiro se ainda não existem."""
        self.db.conn.executescript("""
        CREATE TABLE IF NOT EXISTS financeiro_config (
            funcionario_id INTEGER PRIMARY KEY,
            multiplicador  REAL    NOT NULL DEFAULT 1.5,
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        CREATE TABLE IF NOT EXISTS financeiro_mov (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            data           TEXT    NOT NULL,
            tipo           TEXT    NOT NULL,
            valor          REAL    NOT NULL,
            observacao     TEXT,
            criado_em      TEXT    NOT NULL,
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        );
        """)
        self.db.conn.commit()

    def _get_mult(self, fid):
        self._ensure_tables()
        row = self.db.conn.execute(
            "SELECT multiplicador FROM financeiro_config WHERE funcionario_id=?", (fid,)
        ).fetchone()
        return row[0] if row else 1.5

    def _set_mult(self, fid, mult):
        self._ensure_tables()
        self.db.conn.execute(
            "INSERT OR REPLACE INTO financeiro_config(funcionario_id,multiplicador) VALUES(?,?)",
            (fid, mult))
        self.db.conn.commit()

    def _get_movimentos(self, fid):
        self._ensure_tables()
        return self.db.conn.execute(
            "SELECT * FROM financeiro_mov WHERE funcionario_id=? ORDER BY data DESC, criado_em DESC",
            (fid,)
        ).fetchall()

    def _add_movimento(self, fid, data, tipo, valor, obs):
        self._ensure_tables()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.db.conn.execute(
            "INSERT INTO financeiro_mov(funcionario_id,data,tipo,valor,observacao,criado_em)"
            " VALUES(?,?,?,?,?,?)",
            (fid, data, tipo, valor, obs, now))
        self.db.conn.commit()

    def _del_movimento(self, mov_id):
        self.db.conn.execute("DELETE FROM financeiro_mov WHERE id=?", (mov_id,))
        self.db.conn.commit()

    # ── Cálculo financeiro ────────────────────────────────────────────────────
    def _calc(self, f, mult):
        """Retorna dicionário com todos os valores financeiros do funcionário."""
        sal   = f["salario"]
        tipo  = f["tipo_pag"]
        # Valor da hora conforme tipo de pagamento
        if tipo == "semana":
            vh = sal / 44.0
        else:
            vh = sal / 220.0
        vhe = vh * mult   # valor hora extra

        # Horas extras acumuladas (banco de horas)
        saldo_min = f["saldo_min"] if "saldo_min" in f.keys() else 0
        valor_bruto = (saldo_min / 60.0) * vhe

        # Pagamentos já registrados (tabela pagamentos_horas existente)
        row = self.db.conn.execute(
            "SELECT COALESCE(SUM(valor_pago),0) as total FROM pagamentos_horas WHERE funcionario_id=?",
            (f["id"],)).fetchone()
        pago_oficial = row[0] if row else 0.0

        # Adiantamentos registrados no financeiro_mov
        movs = self._get_movimentos(f["id"])
        total_adiant = sum(m["valor"] for m in movs if m["tipo"] == "adiantamento")
        total_bonus  = sum(m["valor"] for m in movs if m["tipo"] == "bonus")

        saldo_pendente = valor_bruto - pago_oficial - total_adiant + total_bonus

        return {
            "salario":        sal,
            "tipo_pag":       tipo,
            "valor_hora":     vh,
            "valor_hora_ext": vhe,
            "mult":           mult,
            "saldo_min":      saldo_min,
            "valor_bruto":    valor_bruto,
            "pago_oficial":   pago_oficial,
            "total_adiant":   total_adiant,
            "total_bonus":    total_bonus,
            "saldo_pendente": saldo_pendente,
            "movimentos":     movs,
        }

    # ── Formatadores ──────────────────────────────────────────────────────────
    @staticmethod
    def _brl(v):
        s = f"R$ {abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        return ("- " if v < 0 else "") + s

    @staticmethod
    def _hmin(m):
        h = int(abs(m)) // 60; mn = int(abs(m)) % 60
        return f"{h:02d}h{mn:02d}"

    # ── Construção da interface ───────────────────────────────────────────────
    def _build(self):
        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=22); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="💳  Financeiro", bg=C["bg_panel"], fg=C["text_hi"],
                 font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Acompanhamento individual · Horas extras em dinheiro",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w")
        Btn(hdr, "↻  Atualizar", self._refresh, style="outline").pack(side="right", pady=4)
        Sep(self).pack(fill="x", padx=36)

        # ── Corpo principal: sidebar esquerda + área direita ─────────────────
        body = tk.Frame(self, bg=C["bg_panel"]); body.pack(fill="both", expand=True)

        # Sidebar: lista de funcionários
        self._build_sidebar(body)

        # Área direita (conteúdo dinâmico)
        self._right = tk.Frame(body, bg=C["bg_panel"]); self._right.pack(side="left", fill="both", expand=True)
        self._show_placeholder()

    def _build_sidebar(self, parent):
        sb = tk.Frame(parent, bg=C["sidebar"], width=200); sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        tk.Label(sb, text="FUNCIONÁRIOS", bg=C["sidebar"], fg=C["text_lo"],
                 font=(FONT_SMALL[0], 8, "bold")).pack(anchor="w", padx=16, pady=(16,4))
        Sep(sb).pack(fill="x", padx=12, pady=2)

        scroll_frame = tk.Frame(sb, bg=C["sidebar"]); scroll_frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(scroll_frame, bg=C["sidebar"], highlightthickness=0)
        vsb    = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        self._func_list = tk.Frame(canvas, bg=C["sidebar"])
        self._func_list.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=self._func_list, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
        self._func_canvas = canvas

    def _load_funcionarios(self):
        for w in self._func_list.winfo_children(): w.destroy()
        funcs = self.db.get_banco_horas()  # já tem saldo_min
        self._funcs = funcs
        for f in funcs:
            fid  = f["id"]
            nome = f["nome"].capitalize()
            btn_frame = tk.Frame(self._func_list, bg=C["sidebar"], cursor="hand2")
            btn_frame.pack(fill="x")
            dot_color = C["success"] if f["saldo_min"] > 0 else C["text_lo"]
            tk.Label(btn_frame, text="●", bg=C["sidebar"], fg=dot_color,
                     font=("Segoe UI",8)).pack(side="left", padx=(12,4), pady=12)
            tk.Label(btn_frame, text=nome, bg=C["sidebar"], fg=C["text_hi"],
                     font=FONT_BODY, anchor="w").pack(side="left", fill="x", expand=True)
            bar = tk.Frame(btn_frame, bg=C["sidebar"], width=3); bar.pack(side="right", fill="y")

            def _click(e=None, fid=fid, frame=btn_frame, b=bar):
                self._select_func(fid, frame, b)

            def _enter(e, fr=btn_frame): fr.config(bg=C["hover"])
            def _leave(e, fr=btn_frame, fid=fid):
                fr.config(bg=C["bg_card"] if self._sel_id == fid else C["sidebar"])

            for w in btn_frame.winfo_children():
                w.bind("<Button-1>", _click)
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)
            btn_frame.bind("<Button-1>", _click)
            btn_frame.bind("<Enter>",    _enter)
            btn_frame.bind("<Leave>",    _leave)
            btn_frame._bar = bar

        self._func_btns = {f["id"]: self._func_list.winfo_children()[i]
                           for i, f in enumerate(funcs)}

    def _select_func(self, fid, frame=None, bar=None):
        # Reseta visual de todos
        for w in self._func_list.winfo_children():
            w.config(bg=C["sidebar"])
            for c in w.winfo_children(): c.config(bg=C["sidebar"])
            if hasattr(w, "_bar"): w._bar.config(bg=C["sidebar"])
        # Destaca selecionado
        if frame:
            frame.config(bg=C["bg_card"])
            for c in frame.winfo_children(): c.config(bg=C["bg_card"])
            if bar: bar.config(bg=C["accent"])
        self._sel_id = fid
        f = next((x for x in self._funcs if x["id"] == fid), None)
        if f:
            self._render_financeiro(f)

    def _show_placeholder(self):
        for w in self._right.winfo_children(): w.destroy()
        ph = tk.Frame(self._right, bg=C["bg_panel"]); ph.pack(expand=True)
        tk.Label(ph, text="💳", bg=C["bg_panel"], fg=C["text_lo"],
                 font=("Segoe UI",48)).pack(pady=8)
        tk.Label(ph, text="Selecione um funcionário", bg=C["bg_panel"],
                 fg=C["text_mid"], font=("Segoe UI",14)).pack()
        tk.Label(ph, text="para ver o resumo financeiro",
                 bg=C["bg_panel"], fg=C["text_lo"], font=FONT_BODY).pack()

    # ── Renderização do painel financeiro ─────────────────────────────────────
    def _render_financeiro(self, f):
        for w in self._right.winfo_children(): w.destroy()

        mult  = self._get_mult(f["id"])
        dados = self._calc(f, mult)
        self._dados = dados
        self._cur_f  = f

        # Container com scroll
        canvas = tk.Canvas(self._right, bg=C["bg_panel"], highlightthickness=0)
        vsb    = ttk.Scrollbar(self._right, orient="vertical", command=canvas.yview)
        inner  = tk.Frame(canvas, bg=C["bg_panel"])
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        pad = dict(padx=28, pady=6)

        # ── Nome + configurações de multiplicador ─────────────────────────────
        top = tk.Frame(inner, bg=C["bg_panel"], padx=28, pady=16); top.pack(fill="x")
        name_f = tk.Frame(top, bg=C["bg_panel"]); name_f.pack(side="left", fill="x", expand=True)
        tk.Label(name_f, text=f["nome"].upper(), bg=C["bg_panel"],
                 fg=C["text_hi"], font=("Segoe UI",16,"bold")).pack(anchor="w")
        tipo_str = "Semanal (÷ 44h)" if dados["tipo_pag"] == "semana" else "Mensal (÷ 220h)"
        tk.Label(name_f, text=f"Tipo de pagamento: {tipo_str}",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_SMALL).pack(anchor="w")

        # Multiplicador configurável — botões de seleção rápida
        mult_f = tk.Frame(top, bg=C["bg_card"],
                          highlightbackground=C["border"], highlightthickness=1)
        mult_f.pack(side="right", padx=4)
        tk.Frame(mult_f, bg=self._GOLD, height=2).pack(fill="x")
        mi = tk.Frame(mult_f, bg=C["bg_card"], padx=14, pady=10); mi.pack()
        tk.Label(mi, text="Hora extra — multiplicador", bg=C["bg_card"],
                 fg=C["text_lo"], font=(FONT_SMALL[0],8,"bold")).pack(anchor="w")

        # Botões rápidos 1.5× / 2× / 3×
        def _set_mult(v):
            self._mult_var.set(v)
            self._salvar_mult(f["id"])
            self._select_func(f["id"])  # recarrega tela com novo valor

        btn_row = tk.Frame(mi, bg=C["bg_card"]); btn_row.pack(fill="x", pady=6)
        for label, val in [("1.5× (+50%)", 1.5), ("2× (dobro)", 2.0), ("3× (triplo)", 3.0)]:
            is_sel = abs(mult - val) < 0.01
            cor_bg = self._GOLD if is_sel else C["bg_input"]
            cor_fg = C["bg_card"] if is_sel else C["text_mid"]
            tk.Button(btn_row, text=label, bg=cor_bg, fg=cor_fg,
                      relief="flat", cursor="hand2",
                      font=("Segoe UI", 9, "bold" if is_sel else "normal"),
                      padx=8, pady=4,
                      command=lambda v=val: _set_mult(v)).pack(side="left", padx=2)

        # Campo manual para valores personalizados
        mrow = tk.Frame(mi, bg=C["bg_card"]); mrow.pack(fill="x", pady=2)
        tk.Label(mrow, text="Ou digite:", bg=C["bg_card"], fg=C["text_lo"],
                 font=FONT_SMALL).pack(side="left", padx=(0,4))
        self._mult_var.set(mult)
        mult_ent = tk.Entry(mrow, textvariable=self._mult_var, width=5,
                            bg=C["bg_input"], fg=self._GOLD, font=("Segoe UI",11,"bold"),
                            insertbackground=self._GOLD, relief="flat",
                            highlightbackground=C["border"], highlightthickness=1)
        mult_ent.pack(side="left")
        tk.Label(mrow, text="×", bg=C["bg_card"], fg=C["text_lo"],
                 font=("Segoe UI",11)).pack(side="left", padx=4)
        Btn(mi, "Salvar", cmd=lambda: self._salvar_mult(f["id"]),
            style="warning").pack(anchor="e", pady=2)

        Sep(inner).pack(fill="x", padx=28, pady=4)

        # ── Cartões de resumo (linha 1) ───────────────────────────────────────
        row1 = tk.Frame(inner, bg=C["bg_panel"], padx=28, pady=4); row1.pack(fill="x")
        cards1 = [
            ("SALÁRIO",              self._brl(dados["salario"]),        f'{"Mensal" if dados["tipo_pag"]=="mes" else "Semanal"}', C["accent"]),
            ("VALOR DA HORA",        self._brl(dados["valor_hora"]),     f'Salário ÷ {"220h" if dados["tipo_pag"]=="mes" else "44h"}', C["purple"] if True else C["purple"]),
            ("HORA EXTRA (×{:.2g})".format(dados["mult"]),
                                     self._brl(dados["valor_hora_ext"]), f'Hora × {dados["mult"]}×', self._GOLD),
        ]
        for title, val, sub, cor in cards1:
            cf = tk.Frame(row1, bg=C["bg_card"],
                          highlightbackground=C["border"], highlightthickness=1)
            cf.pack(side="left", fill="x", expand=True, padx=4)
            tk.Frame(cf, bg=cor, height=3).pack(fill="x")
            ci = tk.Frame(cf, bg=C["bg_card"], padx=16, pady=14); ci.pack(fill="x")
            tk.Label(ci, text=title, bg=C["bg_card"], fg=C["text_lo"],
                     font=(FONT_SMALL[0],8,"bold")).pack(anchor="w")
            tk.Label(ci, text=val,   bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",20,"bold")).pack(anchor="w", pady=2)
            tk.Label(ci, text=sub,   bg=C["bg_card"], fg=C["text_lo"],
                     font=FONT_SMALL).pack(anchor="w")

        # ── Cartões de resumo (linha 2) ───────────────────────────────────────
        row2 = tk.Frame(inner, bg=C["bg_panel"], padx=28, pady=4); row2.pack(fill="x")
        sinal_pend = C["success"] if dados["saldo_pendente"] >= 0 else self._CORAL
        cards2 = [
            ("HRS EXTRAS ACUMULADAS", self._hmin(dados["saldo_min"]),     "Banco de horas",            C["text_hi"]),
            ("VALOR ACUMULADO",        self._brl(dados["valor_bruto"]),   "Horas × valor h. extra",    self._TEAL),
            ("ADIANTAMENTOS",          self._brl(dados["total_adiant"]),  "Pagamentos registrados",     self._CORAL),
            ("SALDO RESTANTE",         self._brl(dados["saldo_pendente"]),"A receber",                 sinal_pend),
        ]
        for title, val, sub, cor in cards2:
            cf = tk.Frame(row2, bg=C["bg_card"],
                          highlightbackground=C["border"], highlightthickness=1)
            cf.pack(side="left", fill="x", expand=True, padx=4)
            tk.Frame(cf, bg=cor, height=3).pack(fill="x")
            ci = tk.Frame(cf, bg=C["bg_card"], padx=16, pady=14); ci.pack(fill="x")
            tk.Label(ci, text=title, bg=C["bg_card"], fg=C["text_lo"],
                     font=(FONT_SMALL[0],8,"bold")).pack(anchor="w")
            tk.Label(ci, text=val,   bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",18,"bold")).pack(anchor="w", pady=2)
            tk.Label(ci, text=sub,   bg=C["bg_card"], fg=C["text_lo"],
                     font=FONT_SMALL).pack(anchor="w")

        Sep(inner).pack(fill="x", padx=28, pady=8)

        # ── Área inferior: histórico + ações ─────────────────────────────────
        bot = tk.Frame(inner, bg=C["bg_panel"], padx=28, pady=0); bot.pack(fill="both", expand=True)
        bot.columnconfigure(0, weight=3); bot.columnconfigure(1, weight=2)

        # ── Histórico financeiro (esquerda) ───────────────────────────────────
        hf = tk.Frame(bot, bg=C["bg_panel"]); hf.grid(row=0, column=0, sticky="nsew", padx=(0,12))

        tk.Label(hf, text="HISTÓRICO FINANCEIRO", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,6))

        hist_card = tk.Frame(hf, bg=C["bg_card"],
                              highlightbackground=C["border"], highlightthickness=1)
        hist_card.pack(fill="both", expand=True)
        tk.Frame(hist_card, bg=C["accent"], height=3).pack(fill="x")

        hist_canvas = tk.Canvas(hist_card, bg=C["bg_card"], highlightthickness=0, height=280)
        hscroll = ttk.Scrollbar(hist_card, orient="vertical", command=hist_canvas.yview)
        self._hist_inner = tk.Frame(hist_canvas, bg=C["bg_card"])
        self._hist_inner.bind("<Configure>",
            lambda e: hist_canvas.configure(scrollregion=hist_canvas.bbox("all")))
        hist_canvas.create_window((0,0), window=self._hist_inner, anchor="nw")
        hist_canvas.configure(yscrollcommand=hscroll.set)
        hscroll.pack(side="right", fill="y"); hist_canvas.pack(fill="both", expand=True)
        self._hist_canvas_ref = hist_canvas

        self._render_historico(dados["movimentos"], dados)

        # ── Painel de ações (direita) ─────────────────────────────────────────
        af = tk.Frame(bot, bg=C["bg_panel"]); af.grid(row=0, column=1, sticky="nsew")

        # Registrar adiantamento
        tk.Label(af, text="REGISTRAR MOVIMENTO", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,6))

        ac = tk.Frame(af, bg=C["bg_card"],
                      highlightbackground=C["border"], highlightthickness=1)
        ac.pack(fill="x")
        tk.Frame(ac, bg=self._CORAL, height=3).pack(fill="x")
        ai = tk.Frame(ac, bg=C["bg_card"], padx=18, pady=16); ai.pack(fill="x")

        # Tipo
        tk.Label(ai, text="Tipo:", bg=C["bg_card"], fg=C["text_mid"],
                 font=FONT_BODY).pack(anchor="w")
        self._tipo_mov = tk.StringVar(value="adiantamento")
        tipo_f = tk.Frame(ai, bg=C["bg_card"]); tipo_f.pack(fill="x", pady=4)
        for txt, val, cor in [("Adiantamento", "adiantamento", self._CORAL),
                               ("Bônus",        "bonus",        self._TEAL)]:
            rb = tk.Radiobutton(tipo_f, text=txt, variable=self._tipo_mov, value=val,
                                bg=C["bg_card"], fg=C["text_hi"],
                                selectcolor=C["bg_dark"], activebackground=C["bg_card"],
                                activeforeground=cor, font=FONT_BODY)
            rb.pack(side="left", padx=4)

        # Data
        tk.Label(ai, text="Data (DD/MM/AAAA):", bg=C["bg_card"],
                 fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=(8,0))
        self._ent_data = tk.Entry(ai, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                   insertbackground=C["accent"], relief="flat",
                                   highlightbackground=C["border"], highlightthickness=1)
        self._ent_data.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self._ent_data.pack(fill="x", pady=2)

        # Valor
        tk.Label(ai, text="Valor (R$):", bg=C["bg_card"],
                 fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=(8,0))
        self._ent_valor_mov = tk.Entry(ai, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                        insertbackground=C["accent"], relief="flat",
                                        highlightbackground=C["border"], highlightthickness=1)
        self._ent_valor_mov.pack(fill="x", pady=2)

        # Observação
        tk.Label(ai, text="Observação:", bg=C["bg_card"],
                 fg=C["text_mid"], font=FONT_BODY).pack(anchor="w", pady=(8,0))
        self._ent_obs = tk.Entry(ai, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                  insertbackground=C["accent"], relief="flat",
                                  highlightbackground=C["border"], highlightthickness=1)
        self._ent_obs.pack(fill="x", pady=2)

        Btn(ai, "✓  Registrar", cmd=lambda: self._registrar_mov(f),
            style="danger").pack(anchor="w", pady=(12,0))

        # Gerar comprovante
        Sep(af).pack(fill="x", pady=12)
        tk.Label(af, text="COMPROVANTE", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,6))

        gc = tk.Frame(af, bg=C["bg_card"],
                      highlightbackground=C["border"], highlightthickness=1)
        gc.pack(fill="x")
        tk.Frame(gc, bg=C["success"], height=3).pack(fill="x")
        gi = tk.Frame(gc, bg=C["bg_card"], padx=18, pady=14); gi.pack(fill="x")

        Btn(gi, "📄  Gerar PDF",       cmd=lambda: self._gerar_pdf(f, dados),       style="success").pack(fill="x", pady=2)
        Btn(gi, "💬  Texto WhatsApp",   cmd=lambda: self._gerar_whatsapp(f, dados),  style="outline").pack(fill="x", pady=2)
        Btn(gi, "📋  Resumo Completo",  cmd=lambda: self._gerar_resumo(f, dados),    style="outline").pack(fill="x", pady=2)

    def _render_historico(self, movs, dados):
        for w in self._hist_inner.winfo_children(): w.destroy()

        # Linha de horas extras (banco de horas)
        if dados["saldo_min"] > 0:
            self._hist_linha(
                data="Acumulado",
                tipo="Horas extras",
                valor=dados["valor_bruto"],
                obs=f"{self._hmin(dados['saldo_min'])} · banco de horas",
                cor=self._TEAL,
                sinal="+",
                mov_id=None,
            )
            Sep(self._hist_inner).pack(fill="x")

        # Pagamentos da folha
        rows_pag = self.db.conn.execute(
            "SELECT * FROM pagamentos_horas WHERE funcionario_id=? ORDER BY criado_em DESC",
            (self._sel_id,)).fetchall()
        for p in rows_pag:
            self._hist_linha(
                data=(p["criado_em"] or "")[:10],
                tipo="Pagamento registrado",
                valor=p["valor_pago"],
                obs=p["descricao"] or "",
                cor=C["accent"],
                sinal="-",
                mov_id=None,
            )
            Sep(self._hist_inner).pack(fill="x")

        # Movimentos do financeiro
        for m in movs:
            cor   = self._CORAL if m["tipo"] == "adiantamento" else self._TEAL
            sinal = "-" if m["tipo"] == "adiantamento" else "+"
            self._hist_linha(
                data=m["data"],
                tipo=m["tipo"].capitalize(),
                valor=m["valor"],
                obs=m["observacao"] or "",
                cor=cor,
                sinal=sinal,
                mov_id=m["id"],
            )
            Sep(self._hist_inner).pack(fill="x")

        if not movs and dados["saldo_min"] == 0 and not rows_pag:
            tk.Label(self._hist_inner, text="Nenhum movimento registrado.",
                     bg=C["bg_card"], fg=C["text_lo"], font=FONT_BODY,
                     pady=20).pack(expand=True)

    def _hist_linha(self, data, tipo, valor, obs, cor, sinal, mov_id):
        rw = tk.Frame(self._hist_inner, bg=C["bg_card"], pady=6); rw.pack(fill="x")
        tk.Frame(rw, bg=cor, width=4).pack(side="left", fill="y", padx=(12,8))
        lb = tk.Frame(rw, bg=C["bg_card"]); lb.pack(side="left", fill="x", expand=True)

        # Data formatada
        try:
            dt = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m")
        except Exception:
            dt = str(data)[:5] if data else "--"

        tk.Label(lb, text=f"{dt}  ·  {tipo}", bg=C["bg_card"],
                 fg=C["text_hi"], font=FONT_BODY, anchor="w").pack(anchor="w")
        if obs:
            tk.Label(lb, text=obs, bg=C["bg_card"],
                     fg=C["text_lo"], font=FONT_SMALL, anchor="w").pack(anchor="w")

        val_txt = f"{sinal}{self._brl(valor)}"
        tk.Label(rw, text=val_txt, bg=C["bg_card"],
                 fg=cor, font=("Segoe UI",12,"bold")).pack(side="right", padx=8)

        # Botão excluir (só para movimentos do financeiro)
        if mov_id:
            def _del(mid=mov_id):
                if messagebox.askyesno("Excluir", "Remover este movimento?"):
                    self._del_movimento(mid)
                    self._refresh()
            tk.Label(rw, text="✕", bg=C["bg_card"], fg=C["text_lo"],
                     font=FONT_SMALL, cursor="hand2").pack(side="right")
            rw.winfo_children()[-1].bind("<Button-1>", lambda e: _del())

    # ── Ações ─────────────────────────────────────────────────────────────────
    def _salvar_mult(self, fid):
        try:
            m = float(str(self._mult_var.get()).replace(",", "."))
            if m <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Erro", "Multiplicador inválido (ex: 1.5)"); return
        self._set_mult(fid, m)
        self._refresh()

    def _registrar_mov(self, f):
        tipo  = self._tipo_mov.get()
        raw_v = self._ent_valor_mov.get().strip().replace(",", ".")
        obs   = self._ent_obs.get().strip()
        raw_d = self._ent_data.get().strip()
        try:
            valor = float(raw_v)
            if valor <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Erro", "Digite um valor válido (ex: 50,00)"); return
        try:
            dt = datetime.strptime(raw_d, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            messagebox.showerror("Erro", "Data inválida (use DD/MM/AAAA)"); return
        tipo_label = "Adiantamento" if tipo == "adiantamento" else "Bônus"
        ok = messagebox.askyesno("Confirmar",
            f"Registrar {tipo_label} de {self._brl(valor)} para {f['nome']}?")
        if not ok: return
        self._add_movimento(f["id"], dt, tipo, valor, obs)
        self._ent_valor_mov.delete(0, "end")
        self._ent_obs.delete(0, "end")
        self._refresh()

    def _refresh(self):
        if self._sel_id:
            funcs = self.db.get_banco_horas()
            self._funcs = funcs
            f = next((x for x in funcs if x["id"] == self._sel_id), None)
            if f: self._render_financeiro(f)
        self._load_funcionarios()

    # ── Comprovantes ──────────────────────────────────────────────────────────
    def _gerar_whatsapp(self, f, dados):
        tipo_str = "Semanal" if dados["tipo_pag"] == "semana" else "Mensal"
        linhas = [
            f"📋 *RESUMO FINANCEIRO — {f['nome'].upper()}*",
            f"{'─'*34}",
            f"💰 Salário {tipo_str}: {self._brl(dados['salario'])}",
            f"⏱ Valor da hora: {self._brl(dados['valor_hora'])}",
            f"⚡ Hora extra (×{dados['mult']}): {self._brl(dados['valor_hora_ext'])}",
            f"{'─'*34}",
            f"🕐 Horas extras: {self._hmin(dados['saldo_min'])}",
            f"💵 Valor bruto: {self._brl(dados['valor_bruto'])}",
            f"💸 Adiantamentos: {self._brl(dados['total_adiant'])}",
            f"✅ Saldo restante: {self._brl(dados['saldo_pendente'])}",
            f"{'─'*34}",
            f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ]
        texto = "\n".join(linhas)
        try:
            self.clipboard_clear(); self.clipboard_append(texto)
            messagebox.showinfo("WhatsApp", "Texto copiado para a área de transferência!\nCole no WhatsApp.")
        except Exception:
            self._mostrar_texto(texto)

    def _gerar_resumo(self, f, dados):
        tipo_str = "Semanal" if dados["tipo_pag"] == "semana" else "Mensal"
        movs = dados["movimentos"]
        linhas = [
            "=" * 50,
            f"  RESUMO FINANCEIRO — {f['nome'].upper()}",
            f"  Emitido em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            "=" * 50,
            "",
            "DADOS DO FUNCIONÁRIO",
            f"  Salário {tipo_str}:      {self._brl(dados['salario'])}",
            f"  Valor da hora:         {self._brl(dados['valor_hora'])}",
            f"  Hora extra (×{dados['mult']}):    {self._brl(dados['valor_hora_ext'])}",
            "",
            "HORAS EXTRAS",
            f"  Total acumulado:       {self._hmin(dados['saldo_min'])}",
            f"  Valor bruto:           {self._brl(dados['valor_bruto'])}",
            "",
            "PAGAMENTOS",
            f"  Adiantamentos:         {self._brl(dados['total_adiant'])}",
            f"  Bônus registrados:     {self._brl(dados['total_bonus'])}",
            f"  Pago (folha):          {self._brl(dados['pago_oficial'])}",
            "",
            f"  SALDO RESTANTE:        {self._brl(dados['saldo_pendente'])}",
            "",
            "HISTÓRICO DE MOVIMENTOS",
        ]
        for m in movs:
            try:
                dt = datetime.strptime(m["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                dt = m["data"]
            sinal = "-" if m["tipo"] == "adiantamento" else "+"
            linhas.append(f"  {dt}  {m['tipo'].capitalize():<16} {sinal}{self._brl(m['valor'])}")
            if m["observacao"]: linhas.append(f"          Obs: {m['observacao']}")
        linhas += ["", "=" * 50]
        self._mostrar_texto("\n".join(linhas))

    def _gerar_pdf(self, f, dados):
        """Gera PDF do resumo financeiro usando reportlab se disponível."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            messagebox.showwarning("PDF indisponível",
                "reportlab não instalado.\nExecute: pip install reportlab\n\nUsando resumo em texto.")
            self._gerar_resumo(f, dados)
            return

        from tkinter import filedialog as fd
        path = fd.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF","*.pdf")],
            initialfile=f"financeiro_{f['nome']}_{datetime.now().strftime('%Y%m%d')}.pdf",
            title="Salvar comprovante PDF")
        if not path: return

        try:
            doc = SimpleDocTemplate(path, pagesize=A4,
                                    topMargin=2*cm, bottomMargin=2*cm,
                                    leftMargin=2.5*cm, rightMargin=2.5*cm)
            styles = getSampleStyleSheet()
            BG     = colors.HexColor("#13161E")
            ACCENT = colors.HexColor("#00C2FF")
            GOLD   = colors.HexColor("#FFD166")
            TEAL   = colors.HexColor("#06D6A0")
            CORAL  = colors.HexColor("#EF476F")
            WHITE  = colors.white

            title_style = ParagraphStyle("titulo", parent=styles["Title"],
                                          fontSize=18, textColor=ACCENT, spaceAfter=4)
            sub_style   = ParagraphStyle("sub",    parent=styles["Normal"],
                                          fontSize=10, textColor=colors.grey, spaceAfter=12)
            sect_style  = ParagraphStyle("sect",   parent=styles["Normal"],
                                          fontSize=9,  textColor=colors.grey,
                                          fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=4)

            story = [
                Paragraph(f"COMPROVANTE FINANCEIRO", title_style),
                Paragraph(f"{f['nome'].upper()} · {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style),
                HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=12),
                Paragraph("DADOS SALARIAIS", sect_style),
            ]

            tipo_str = "Semanal (÷ 44h)" if dados["tipo_pag"] == "semana" else "Mensal (÷ 220h)"
            t1 = Table([
                ["Salário",          self._brl(dados["salario"]),         tipo_str],
                ["Valor da Hora",    self._brl(dados["valor_hora"]),      "Base de cálculo"],
                [f"Hora Extra (×{dados['mult']})", self._brl(dados["valor_hora_ext"]), f"Multiplicador: {dados['mult']}×"],
            ], colWidths=[5*cm, 4*cm, 6.5*cm])
            t1.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1A1E2A")),
                ("TEXTCOLOR",  (0,0),(-1,-1), WHITE),
                ("FONTSIZE",   (0,0),(-1,-1), 10),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.HexColor("#1A1E2A"), colors.HexColor("#13161E")]),
                ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#1E2235")),
                ("LEFTPADDING",(0,0),(-1,-1), 10),
                ("RIGHTPADDING",(0,0),(-1,-1), 10),
                ("TOPPADDING", (0,0),(-1,-1), 8),
                ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ]))

            story += [t1, Spacer(1, 0.4*cm), Paragraph("HORAS EXTRAS & SALDO", sect_style)]

            t2 = Table([
                ["Horas extras acumuladas", self._hmin(dados["saldo_min"])],
                ["Valor bruto",             self._brl(dados["valor_bruto"])],
                ["Adiantamentos",           f"- {self._brl(dados['total_adiant'])}"],
                ["Bônus",                   f"+ {self._brl(dados['total_bonus'])}"],
                ["Pago (folha)",            f"- {self._brl(dados['pago_oficial'])}"],
                ["SALDO RESTANTE",          self._brl(dados["saldo_pendente"])],
            ], colWidths=[8*cm, 7.5*cm])
            t2.setStyle(TableStyle([
                ("TEXTCOLOR",    (0,0),(-1,-1), WHITE),
                ("FONTSIZE",     (0,0),(-1,-1), 10),
                ("ROWBACKGROUNDS",(0,0),(-1,-2),[colors.HexColor("#1A1E2A"),colors.HexColor("#13161E")]),
                ("BACKGROUND",   (0,-1),(-1,-1), colors.HexColor("#005599")),
                ("FONTNAME",     (0,-1),(-1,-1), "Helvetica-Bold"),
                ("FONTSIZE",     (0,-1),(-1,-1), 12),
                ("GRID",         (0,0),(-1,-1), 0.3, colors.HexColor("#1E2235")),
                ("LEFTPADDING",  (0,0),(-1,-1), 10),
                ("RIGHTPADDING", (0,0),(-1,-1), 10),
                ("TOPPADDING",   (0,0),(-1,-1), 8),
                ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ]))
            story += [t2]

            # Histórico
            movs = dados["movimentos"]
            if movs:
                story.append(Paragraph("HISTÓRICO DE MOVIMENTOS", sect_style))
                rows_h = [["Data", "Tipo", "Valor", "Observação"]]
                for m in movs:
                    try: dt = datetime.strptime(m["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except: dt = m["data"]
                    sinal = "-" if m["tipo"] == "adiantamento" else "+"
                    rows_h.append([dt, m["tipo"].capitalize(),
                                   f"{sinal}{self._brl(m['valor'])}", m["observacao"] or ""])
                th = Table(rows_h, colWidths=[2.5*cm, 3.5*cm, 3.5*cm, 6*cm])
                th.setStyle(TableStyle([
                    ("BACKGROUND",  (0,0),(-1,0), colors.HexColor("#005599")),
                    ("TEXTCOLOR",   (0,0),(-1,-1), WHITE),
                    ("FONTNAME",    (0,0),(-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",    (0,0),(-1,-1), 9),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#1A1E2A"),colors.HexColor("#13161E")]),
                    ("GRID",        (0,0),(-1,-1), 0.3, colors.HexColor("#1E2235")),
                    ("LEFTPADDING", (0,0),(-1,-1), 8),
                    ("RIGHTPADDING",(0,0),(-1,-1), 8),
                    ("TOPPADDING",  (0,0),(-1,-1), 6),
                    ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ]))
                story.append(th)

            story.append(Spacer(1, 0.6*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
            story.append(Paragraph(f"Documento gerado pelo Sistema de Ponto {APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                                   ParagraphStyle("footer", parent=styles["Normal"],
                                                  fontSize=7, textColor=colors.grey, spaceAfter=0)))

            doc.build(story)
            messagebox.showinfo("PDF Gerado", f"Comprovante salvo em:\n{path}")
        except Exception as ex:
            messagebox.showerror("Erro ao gerar PDF", str(ex))

    def _mostrar_texto(self, texto):
        """Exibe texto em janela popup copiável."""
        win = tk.Toplevel(self); win.title("Resumo Financeiro")
        win.configure(bg=C["bg_panel"]); win.geometry("560x480")
        win.resizable(True, True)
        tk.Label(win, text="Resumo Financeiro", bg=C["bg_panel"],
                 fg=C["text_hi"], font=("Segoe UI",13,"bold")).pack(anchor="w", padx=20, pady=(16,4))
        Sep(win).pack(fill="x", padx=20)
        txt = tk.Text(win, bg=C["bg_card"], fg=C["text_hi"], font=FONT_MONO,
                      padx=16, pady=12, relief="flat", wrap="word",
                      highlightbackground=C["border"], highlightthickness=1)
        txt.pack(fill="both", expand=True, padx=20, pady=12)
        txt.insert("1.0", texto)
        txt.config(state="disabled")
        bf = tk.Frame(win, bg=C["bg_panel"], padx=20, pady=8); bf.pack(fill="x")
        def _copy():
            win.clipboard_clear(); win.clipboard_append(texto)
        Btn(bf, "📋  Copiar",  _copy,       style="primary").pack(side="left", padx=4)
        Btn(bf, "✕  Fechar",  win.destroy,  style="outline").pack(side="right", padx=4)


# ──────────────────────────────────────────────────────────────────────────────
# ETAPA 4.4 — SISTEMA DE CORREÇÃO E VALIDAÇÃO DE PONTO
# ──────────────────────────────────────────────────────────────────────────────

class ValidacaoEngine:
    """
    Detecta problemas em registros de batida de forma independente do CalcEngine.
    Retorna lista de dicts: {codigo, nivel, descricao, detalhe}
    nivel: "critico" | "aviso" | "info"
    """

    JORNADA_MAX_MIN  = 12 * 60   # acima disso → crítico
    JORNADA_LONG_MIN = 10 * 60   # acima disso → aviso
    DUP_TOLERANCIA   = 2         # minutos: batidas iguais nesse intervalo = duplicada
    ALMOCO_MIN_MIN   = 30        # almoço menor que isso → sem almoço real
    AUSENCIA_MIN_MIN = 30        # "sem saída" quando só há entrada

    @staticmethod
    def _hm(s: str) -> int:
        try:
            h, m = map(int, s.strip().split(":"))
            return h * 60 + m
        except Exception:
            return -1

    def validar(self, horarios: list[str], data_str: str, jornada_h: float) -> list[dict]:
        """Retorna lista de alertas detectados."""
        alertas = []
        hrs_min  = [self._hm(h) for h in horarios if self._hm(h) >= 0]
        qtd      = len(hrs_min)

        # ── Sem batidas ───────────────────────────────────────────────────────
        if qtd == 0:
            alertas.append({"codigo": "SEM_BATIDAS", "nivel": "critico",
                            "descricao": "Sem batidas registradas",
                            "detalhe": "Nenhum horário importado para este dia."})
            return alertas

        # ── Horários impossíveis (fora de 00:00–23:59) ───────────────────────
        for i, (h_orig, h_min) in enumerate(zip(horarios, hrs_min)):
            if h_min < 0 or h_min > 23*60+59:
                alertas.append({"codigo": "HORARIO_INVALIDO", "nivel": "critico",
                                "descricao": f"Horário impossível: {h_orig}",
                                "detalhe": f"Batida {i+1} fora do intervalo 00:00–23:59."})

        # ── Batidas duplicadas ────────────────────────────────────────────────
        for i in range(len(hrs_min)):
            for j in range(i+1, len(hrs_min)):
                if abs(hrs_min[i] - hrs_min[j]) <= ValidacaoEngine.DUP_TOLERANCIA:
                    alertas.append({"codigo": "DUPLICADA", "nivel": "aviso",
                                    "descricao": f"Batida duplicada: {horarios[i]} ≈ {horarios[j]}",
                                    "detalhe": "Duas batidas muito próximas — possível erro do relógio."})

        # ── Saída antes da entrada ────────────────────────────────────────────
        for i in range(0, qtd - 1, 2):
            if i+1 < qtd and hrs_min[i+1] < hrs_min[i]:
                alertas.append({"codigo": "SAIDA_ANTES_ENTRADA", "nivel": "critico",
                                "descricao": f"Saída ({horarios[i+1]}) antes da entrada ({horarios[i]})",
                                "detalhe": f"Par de batidas {i+1}/{i+2} tem ordem invertida."})

        # ── Número ímpar de batidas ───────────────────────────────────────────
        if qtd % 2 != 0:
            alertas.append({"codigo": "IMPAR", "nivel": "critico",
                            "descricao": f"Número ímpar de batidas ({qtd})",
                            "detalhe": "Possível batida esquecida — entrada ou saída faltando."})

        # ── Sem saída registrada (apenas 1 batida) ───────────────────────────
        if qtd == 1:
            alertas.append({"codigo": "SEM_SAIDA", "nivel": "critico",
                            "descricao": "Funcionário sem saída registrada",
                            "detalhe": f"Apenas entrada ({horarios[0]}) encontrada."})

        # ── Jornada calculada ─────────────────────────────────────────────────
        if qtd >= 2 and qtd % 2 == 0:
            trab = sum(max(0, hrs_min[i+1] - hrs_min[i]) for i in range(0, qtd, 2))

            if trab > ValidacaoEngine.JORNADA_MAX_MIN:
                alertas.append({"codigo": "JORNADA_EXCESSIVA", "nivel": "critico",
                                "descricao": f"Jornada acima de 12h ({_fmt(trab)})",
                                "detalhe": "Provável erro de batida — confira os horários."})
            elif trab > ValidacaoEngine.JORNADA_LONG_MIN:
                alertas.append({"codigo": "JORNADA_LONGA", "nivel": "aviso",
                                "descricao": f"Jornada acima de 10h ({_fmt(trab)})",
                                "detalhe": "Jornada acima do normal — confirme se é real."})

        # ── Falta de almoço (4 batidas, intervalo < 30 min) ──────────────────
        if qtd == 4:
            almoco = hrs_min[2] - hrs_min[1]
            if 0 < almoco < ValidacaoEngine.ALMOCO_MIN_MIN:
                alertas.append({"codigo": "ALMOCO_CURTO", "nivel": "aviso",
                                "descricao": f"Almoço muito curto ({_fmt(almoco)})",
                                "detalhe": "Intervalo de almoço menor que 30 min. Confirmar trabalho direto?"})
            if almoco <= 0:
                alertas.append({"codigo": "SEM_ALMOCO", "nivel": "aviso",
                                "descricao": "Almoço não registrado corretamente",
                                "detalhe": "Retorno antes da saída para almoço."})

        # ── Dia incompleto (< 50% da jornada esperada) ───────────────────────
        if qtd >= 2 and qtd % 2 == 0:
            trab = sum(max(0, hrs_min[i+1] - hrs_min[i]) for i in range(0, qtd, 2))
            jorn = int(jornada_h * 60)
            if 0 < trab < jorn * 0.5:
                alertas.append({"codigo": "DIA_INCOMPLETO", "nivel": "aviso",
                                "descricao": f"Dia incompleto ({_fmt(trab)} de {_fmt(jorn)})",
                                "detalhe": "Menos de 50% da jornada esperada registrada."})

        return alertas


class PageValidacao(tk.Frame):
    """
    Etapa 4.4 — Validação e Correção de Registros de Ponto.
    Detecta erros automaticamente e permite correção manual com histórico de auditoria.
    """

    # Paleta de severidade
    _COR = {
        "critico": "#FF4D6A",
        "aviso":   "#FFB830",
        "info":    "#00C2FF",
        "ok":      "#00E5A0",
    }
    _ICONE = {"critico": "✕", "aviso": "⚠", "info": "ℹ", "ok": "✓"}

    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db      = db
        self.engine  = CalcEngine()
        self.vld     = ValidacaoEngine()
        self._todos_problemas = []   # lista de dicts com problema + referência à batida
        self._sel_problema    = None
        self._filtro_nivel    = tk.StringVar(value="todos")
        self._build()
        self._ensure_audit_table()
        self._scan()

    # ── Migração segura ───────────────────────────────────────────────────────
    def _ensure_audit_table(self):
        self.db.conn.executescript("""
        CREATE TABLE IF NOT EXISTS ponto_auditoria (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            batida_id     INTEGER NOT NULL,
            admin         TEXT    NOT NULL DEFAULT 'admin',
            criado_em     TEXT    NOT NULL,
            acao          TEXT    NOT NULL,
            horarios_ant  TEXT,
            horarios_nov  TEXT,
            justificativa TEXT,
            FOREIGN KEY (batida_id) REFERENCES batidas(id)
        );
        CREATE TABLE IF NOT EXISTS ponto_flags (
            batida_id        INTEGER PRIMARY KEY,
            trabalho_direto  INTEGER NOT NULL DEFAULT 0,
            corrigido        INTEGER NOT NULL DEFAULT 0,
            ignorado         INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (batida_id) REFERENCES batidas(id)
        );
        """)
        try:
            self.db.conn.execute("ALTER TABLE batidas ADD COLUMN horarios_corrigidos TEXT DEFAULT NULL")
        except Exception:
            pass
        self.db.conn.commit()

    def _get_flag(self, batida_id: int) -> dict:
        row = self.db.conn.execute(
            "SELECT * FROM ponto_flags WHERE batida_id=?", (batida_id,)).fetchone()
        if row:
            return dict(row)
        return {"batida_id": batida_id, "trabalho_direto": 0, "corrigido": 0, "ignorado": 0}

    def _set_flag(self, batida_id: int, **kwargs):
        existing = self.db.conn.execute(
            "SELECT 1 FROM ponto_flags WHERE batida_id=?", (batida_id,)).fetchone()
        if existing:
            sets = ", ".join(f"{k}=?" for k in kwargs)
            self.db.conn.execute(
                f"UPDATE ponto_flags SET {sets} WHERE batida_id=?",
                list(kwargs.values()) + [batida_id])
        else:
            keys = ["batida_id"] + list(kwargs.keys())
            vals = [batida_id]   + list(kwargs.values())
            self.db.conn.execute(
                f"INSERT INTO ponto_flags({','.join(keys)}) VALUES({','.join('?'*len(keys))})",
                vals)
        self.db.conn.commit()

    def _registrar_auditoria(self, batida_id, acao, ant, nov, justificativa):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.db.conn.execute(
            "INSERT INTO ponto_auditoria(batida_id,admin,criado_em,acao,horarios_ant,horarios_nov,justificativa)"
            " VALUES(?,?,?,?,?,?,?)",
            (batida_id, "admin", now, acao, ant, nov, justificativa))
        self.db.conn.commit()

    # ── Scanner de problemas ──────────────────────────────────────────────────
    def _scan(self):
        """Varre TODOS os registros e detecta problemas."""
        self._todos_problemas = []
        funcs = {f["id"]: f for f in self.db.get_funcionarios(apenas_ativos=False)}
        batidas = self.db.conn.execute("""
            SELECT b.*, f.nome as func_nome, f.jornada_h
            FROM batidas b
            LEFT JOIN funcionarios f ON f.id = b.funcionario_id
            ORDER BY b.data DESC, b.funcionario_raw
        """).fetchall()

        for bat in batidas:
            flag = self._get_flag(bat["id"])
            if flag["ignorado"]:
                continue
            # Usa horários corrigidos se existirem
            hrs_raw = bat["horarios_corrigidos"] or bat["horarios"]
            hrs = [h.strip() for h in hrs_raw.split(",") if h.strip()] if hrs_raw else []
            jorn = bat["jornada_h"] or 8.0
            alertas = self.vld.validar(hrs, bat["data"], jorn)
            if alertas:
                for al in alertas:
                    self._todos_problemas.append({
                        "batida": bat,
                        "flag":   flag,
                        "alerta": al,
                        "hrs":    hrs,
                    })

        self._render_lista()
        self._update_contadores()

    # ── Construção da interface ───────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=22); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="🔍  Validação de Registros", bg=C["bg_panel"],
                 fg=C["text_hi"], font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Detecção automática de erros · Correção manual · Histórico de auditoria",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w")
        Btn(hdr, "↻  Escanear", self._scan, style="primary").pack(side="right", pady=4)
        Btn(hdr, "📋  Auditoria", self._abrir_auditoria, style="outline").pack(side="right", pady=4, padx=6)
        Sep(self).pack(fill="x", padx=36)

        # Contadores
        self._cnt_frame = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=10)
        self._cnt_frame.pack(fill="x")
        self._build_contadores()

        # Filtros de nível
        flt = tk.Frame(self, bg=C["bg_card"],
                       highlightbackground=C["border"], highlightthickness=1)
        flt.pack(fill="x", padx=36, pady=0)
        tk.Frame(flt, bg=C["warning"], height=2).pack(fill="x")
        fi = tk.Frame(flt, bg=C["bg_card"], padx=20, pady=10); fi.pack(fill="x")
        tk.Label(fi, text="Filtrar:", bg=C["bg_card"], fg=C["text_mid"],
                 font=FONT_BODY).pack(side="left")
        for lbl, val, cor in [
            ("Todos",    "todos",   C["text_mid"]),
            ("Críticos", "critico", self._COR["critico"]),
            ("Avisos",   "aviso",   self._COR["aviso"]),
        ]:
            rb = tk.Radiobutton(fi, text=lbl, variable=self._filtro_nivel, value=val,
                                bg=C["bg_card"], fg=cor, selectcolor=C["bg_dark"],
                                activebackground=C["bg_card"], activeforeground=cor,
                                font=FONT_BODY, command=self._render_lista)
            rb.pack(side="left", padx=12)

        # Corpo: lista + painel
        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=10)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=3); body.columnconfigure(1, weight=2)

        # Lista de problemas (esquerda)
        lw = tk.Frame(body, bg=C["bg_panel"]); lw.grid(row=0, column=0, sticky="nsew", padx=(0,12))
        tk.Label(lw, text="PROBLEMAS DETECTADOS", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,6))

        list_card = tk.Frame(lw, bg=C["bg_card"],
                              highlightbackground=C["border"], highlightthickness=1)
        list_card.pack(fill="both", expand=True)
        tk.Frame(list_card, bg=self._COR["critico"], height=3).pack(fill="x")

        self._list_canvas = tk.Canvas(list_card, bg=C["bg_card"], highlightthickness=0)
        list_vsb = ttk.Scrollbar(list_card, orient="vertical", command=self._list_canvas.yview)
        self._list_inner = tk.Frame(self._list_canvas, bg=C["bg_card"])
        self._list_inner.bind("<Configure>",
            lambda e: self._list_canvas.configure(scrollregion=self._list_canvas.bbox("all")))
        self._list_canvas.create_window((0,0), window=self._list_inner, anchor="nw")
        self._list_canvas.configure(yscrollcommand=list_vsb.set)
        list_vsb.pack(side="right", fill="y")
        self._list_canvas.pack(fill="both", expand=True)

        # Painel direito
        rw = tk.Frame(body, bg=C["bg_panel"]); rw.grid(row=0, column=1, sticky="nsew")
        self._build_painel_direito(rw)

    def _build_contadores(self):
        for w in self._cnt_frame.winfo_children(): w.destroy()
        for title, attr, cor in [
            ("Críticos",    "_cnt_critico", self._COR["critico"]),
            ("Avisos",      "_cnt_aviso",   self._COR["aviso"]),
            ("Total",       "_cnt_total",   C["text_mid"]),
            ("Ignorados",   "_cnt_ignorado",C["text_lo"]),
        ]:
            cf = tk.Frame(self._cnt_frame, bg=C["bg_card"],
                          highlightbackground=C["border"], highlightthickness=1)
            cf.pack(side="left", padx=4, pady=2)
            tk.Frame(cf, bg=cor, height=3).pack(fill="x")
            ci = tk.Frame(cf, bg=C["bg_card"], padx=20, pady=10); ci.pack()
            lv = tk.Label(ci, text="0", bg=C["bg_card"], fg=cor,
                          font=("Segoe UI",22,"bold"))
            lv.pack()
            tk.Label(ci, text=title, bg=C["bg_card"], fg=C["text_lo"],
                     font=(FONT_SMALL[0],8,"bold")).pack()
            setattr(self, attr, lv)

    def _update_contadores(self):
        n_crit = sum(1 for p in self._todos_problemas if p["alerta"]["nivel"] == "critico")
        n_avis = sum(1 for p in self._todos_problemas if p["alerta"]["nivel"] == "aviso")
        n_tot  = len(self._todos_problemas)
        n_ign  = self.db.conn.execute(
            "SELECT COUNT(*) FROM ponto_flags WHERE ignorado=1").fetchone()[0]
        self._cnt_critico.config(text=str(n_crit))
        self._cnt_aviso.config(text=str(n_avis))
        self._cnt_total.config(text=str(n_tot))
        self._cnt_ignorado.config(text=str(n_ign))

    def _build_painel_direito(self, parent):
        tk.Label(parent, text="DETALHES E CORREÇÃO", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,6))
        self._painel = tk.Frame(parent, bg=C["bg_card"],
                                 highlightbackground=C["border"], highlightthickness=1)
        self._painel.pack(fill="both", expand=True)
        tk.Frame(self._painel, bg=C["accent"], height=3).pack(fill="x")
        self._painel_inner = tk.Frame(self._painel, bg=C["bg_card"], padx=20, pady=16)
        self._painel_inner.pack(fill="both", expand=True)
        self._painel_placeholder()

    def _painel_placeholder(self):
        for w in self._painel_inner.winfo_children(): w.destroy()
        tk.Label(self._painel_inner, text="🔍", bg=C["bg_card"],
                 fg=C["text_lo"], font=("Segoe UI",32)).pack(pady=10, expand=True)
        tk.Label(self._painel_inner, text="Selecione um problema\npara ver detalhes e corrigir",
                 bg=C["bg_card"], fg=C["text_lo"], font=FONT_BODY,
                 justify="center").pack(expand=True)

    # ── Lista de problemas ────────────────────────────────────────────────────
    def _render_lista(self):
        for w in self._list_inner.winfo_children(): w.destroy()
        filtro = self._filtro_nivel.get()
        probs  = [p for p in self._todos_problemas
                  if filtro == "todos" or p["alerta"]["nivel"] == filtro]

        if not probs:
            tk.Label(self._list_inner,
                     text="✓  Nenhum problema encontrado!" if filtro == "todos"
                          else f"Nenhum problema '{filtro}' encontrado.",
                     bg=C["bg_card"], fg=self._COR["ok"],
                     font=("Segoe UI",12), pady=30).pack(expand=True)
            return

        # Agrupa por funcionário/dia para evitar repetição
        for i, prob in enumerate(probs):
            bat  = prob["batida"]
            al   = prob["alerta"]
            cor  = self._COR.get(al["nivel"], C["text_mid"])
            flag = prob["flag"]
            corrigido = flag.get("corrigido", 0)

            row = tk.Frame(self._list_inner, bg=C["bg_card"], cursor="hand2")
            row.pack(fill="x")

            # Barra lateral colorida
            tk.Frame(row, bg=cor, width=5).pack(side="left", fill="y")

            inner = tk.Frame(row, bg=C["bg_card"], padx=12, pady=10)
            inner.pack(side="left", fill="x", expand=True)

            # Linha 1: ícone + descrição
            l1 = tk.Frame(inner, bg=C["bg_card"]); l1.pack(fill="x")
            icone = self._ICONE.get(al["nivel"], "?")
            tk.Label(l1, text=icone, bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",10,"bold"), width=2).pack(side="left")
            tk.Label(l1, text=al["descricao"], bg=C["bg_card"],
                     fg=C["text_hi"], font=("Segoe UI",10,"bold"), anchor="w").pack(side="left")
            if corrigido:
                tk.Label(l1, text="✓ corrigido", bg=C["bg_card"],
                         fg=self._COR["ok"], font=FONT_SMALL).pack(side="right")

            # Linha 2: func + data
            func_nome = bat["func_nome"] or bat["funcionario_raw"]
            try:
                dt_fmt = datetime.strptime(bat["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                dt_fmt = bat["data"]
            tk.Label(inner, text=f"{func_nome}  ·  {dt_fmt}",
                     bg=C["bg_card"], fg=C["text_mid"], font=FONT_SMALL,
                     anchor="w").pack(anchor="w")

            # Botão corrigir no lado direito
            btn_frame = tk.Frame(row, bg=C["bg_card"], padx=8); btn_frame.pack(side="right", fill="y")
            Btn(btn_frame, "Corrigir", cmd=lambda p=prob: self._abrir_correcao(p),
                style="primary" if not corrigido else "outline").pack(expand=True)

            # Clique na linha seleciona o problema
            def _click(e=None, p=prob): self._selecionar_problema(p)
            for w in (row, inner, l1):
                w.bind("<Button-1>", _click)

            Sep(self._list_inner).pack(fill="x")

    # ── Seleção e painel de detalhes ─────────────────────────────────────────
    def _selecionar_problema(self, prob):
        self._sel_problema = prob
        self._render_painel_detalhe(prob)

    def _render_painel_detalhe(self, prob):
        for w in self._painel_inner.winfo_children(): w.destroy()
        bat  = prob["batida"]
        al   = prob["alerta"]
        flag = prob["flag"]
        cor  = self._COR.get(al["nivel"], C["text_mid"])

        func_nome = bat["func_nome"] or bat["funcionario_raw"]
        try:
            dt_fmt = datetime.strptime(bat["data"], "%Y-%m-%d").strftime("%d/%m/%Y (%A)")
        except Exception:
            dt_fmt = bat["data"]

        # Cabeçalho
        tk.Label(self._painel_inner, text=func_nome, bg=C["bg_card"],
                 fg=C["text_hi"], font=("Segoe UI",13,"bold")).pack(anchor="w")
        tk.Label(self._painel_inner, text=dt_fmt, bg=C["bg_card"],
                 fg=C["accent"], font=FONT_BODY).pack(anchor="w", pady=2)
        Sep(self._painel_inner).pack(fill="x", pady=8)

        # Alerta
        af = tk.Frame(self._painel_inner, bg=C["bg_dark"],
                      highlightbackground=cor, highlightthickness=1)
        af.pack(fill="x", pady=4)
        tk.Frame(af, bg=cor, height=3).pack(fill="x")
        ai = tk.Frame(af, bg=C["bg_dark"], padx=12, pady=10); ai.pack(fill="x")
        tk.Label(ai, text=f"{self._ICONE.get(al['nivel'],'?')}  {al['descricao']}",
                 bg=C["bg_dark"], fg=cor, font=("Segoe UI",10,"bold")).pack(anchor="w")
        tk.Label(ai, text=al["detalhe"], bg=C["bg_dark"],
                 fg=C["text_mid"], font=FONT_SMALL, wraplength=260, justify="left").pack(anchor="w")

        Sep(self._painel_inner).pack(fill="x", pady=8)

        # Horários atuais
        tk.Label(self._painel_inner, text="HORÁRIOS ATUAIS", bg=C["bg_card"],
                 fg=C["text_lo"], font=(FONT_SMALL[0],8,"bold")).pack(anchor="w")
        hrs = prob["hrs"]
        LABELS = ["Entrada","Saída almoço","Retorno","Saída","Bat.5","Bat.6","Bat.7","Bat.8"]
        for i, h in enumerate(hrs):
            rw = tk.Frame(self._painel_inner, bg=C["bg_input"], pady=4, padx=10)
            rw.pack(fill="x", pady=2)
            lbl = LABELS[i] if i < len(LABELS) else f"Batida {i+1}"
            tk.Label(rw, text=lbl, bg=C["bg_input"], fg=C["text_mid"],
                     font=FONT_SMALL, width=12, anchor="w").pack(side="left")
            tk.Label(rw, text=h, bg=C["bg_input"], fg=C["text_hi"],
                     font=("Segoe UI",12,"bold")).pack(side="right")
        if not hrs:
            tk.Label(self._painel_inner, text="Nenhum horário registrado",
                     bg=C["bg_card"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w", pady=4)

        # Trabalho direto (se aplicável)
        if flag.get("trabalho_direto"):
            tf = tk.Frame(self._painel_inner, bg="#1A2E1A", padx=12, pady=6)
            tf.pack(fill="x", pady=4)
            tk.Label(tf, text="✓  Marcado como trabalho direto",
                     bg="#1A2E1A", fg=self._COR["ok"], font=FONT_SMALL).pack(anchor="w")

        Sep(self._painel_inner).pack(fill="x", pady=8)

        # Botões de ação
        Btn(self._painel_inner, "✏  Corrigir Registro",
            cmd=lambda: self._abrir_correcao(prob), style="primary").pack(fill="x", pady=2)
        Btn(self._painel_inner, "⚡  Trabalho Direto",
            cmd=lambda: self._marcar_trabalho_direto(prob), style="warning").pack(fill="x", pady=2)
        Btn(self._painel_inner, "○  Ignorar este alerta",
            cmd=lambda: self._ignorar(prob), style="outline").pack(fill="x", pady=2)

    # ── Modal de correção ─────────────────────────────────────────────────────
    def _abrir_correcao(self, prob):
        bat  = prob["batida"]
        hrs  = list(prob["hrs"])

        win = tk.Toplevel(self); win.title("Corrigir Registro")
        win.configure(bg=C["bg_panel"]); win.geometry("560x640")
        win.resizable(False, True); win.grab_set()

        # Header
        tk.Frame(win, bg=C["danger"], height=4).pack(fill="x")
        hf = tk.Frame(win, bg=C["bg_panel"], padx=24, pady=16); hf.pack(fill="x")
        func_nome = bat["func_nome"] or bat["funcionario_raw"]
        try:
            dt_fmt = datetime.strptime(bat["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            dt_fmt = bat["data"]
        tk.Label(hf, text="✏  Corrigir Registro", bg=C["bg_panel"],
                 fg=C["text_hi"], font=("Segoe UI",14,"bold")).pack(anchor="w")
        tk.Label(hf, text=f"{func_nome}  ·  {dt_fmt}",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w")
        Sep(win).pack(fill="x", padx=24)

        # Scroll area
        canvas = tk.Canvas(win, bg=C["bg_panel"], highlightthickness=0)
        vsb    = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scroll_f = tk.Frame(canvas, bg=C["bg_panel"])
        scroll_f.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=scroll_f, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)

        body = tk.Frame(scroll_f, bg=C["bg_panel"], padx=24, pady=12); body.pack(fill="x")

        # ── Alerta atual ──────────────────────────────────────────────────────
        al  = prob["alerta"]
        cor = self._COR.get(al["nivel"], C["text_mid"])
        af  = tk.Frame(body, bg=C["bg_dark"],
                       highlightbackground=cor, highlightthickness=1)
        af.pack(fill="x", pady=(0,12))
        tk.Frame(af, bg=cor, height=3).pack(fill="x")
        tk.Label(af, text=f"  {self._ICONE.get(al['nivel'],'?')}  {al['descricao']}",
                 bg=C["bg_dark"], fg=cor, font=("Segoe UI",10,"bold"), pady=8).pack(anchor="w")

        # ── Editor de batidas ─────────────────────────────────────────────────
        tk.Label(body, text="EDITAR BATIDAS", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(8,4))
        tk.Label(body, text="Use HH:MM  ·  Deixe em branco para remover  ·  Ordem será reordenada",
                 bg=C["bg_panel"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w", pady=(0,8))

        # Garante ao menos 6 campos (para adicionar batidas)
        max_campos = max(8, len(hrs) + 2)
        LABELS = ["Entrada","Saída almoço","Retorno","Saída","Bat.5","Bat.6","Bat.7","Bat.8"]
        entries = []
        for i in range(max_campos):
            rw = tk.Frame(body, bg=C["bg_panel"]); rw.pack(fill="x", pady=3)
            lbl = LABELS[i] if i < len(LABELS) else f"Batida {i+1}"
            tk.Label(rw, text=lbl, bg=C["bg_panel"], fg=C["text_mid"],
                     font=FONT_BODY, width=14, anchor="w").pack(side="left")
            ent = tk.Entry(rw, bg=C["bg_input"], fg=C["text_hi"], font=("Segoe UI",13,"bold"),
                           width=10, insertbackground=C["accent"], relief="flat",
                           highlightbackground=C["border"], highlightthickness=1)
            if i < len(hrs):
                ent.insert(0, hrs[i])
            ent.pack(side="left", padx=8)
            # Indicador de problema
            if i < len(hrs):
                # Verifica se esta batida específica tem problema
                pass
            entries.append(ent)

        # ── Trabalho direto ───────────────────────────────────────────────────
        Sep(body).pack(fill="x", pady=12)
        flag = prob["flag"]
        td_var = tk.BooleanVar(value=bool(flag.get("trabalho_direto", 0)))
        td_frame = tk.Frame(body, bg=C["bg_dark"], padx=14, pady=10)
        td_frame.pack(fill="x", pady=4)
        tk.Label(td_frame, text="⚡", bg=C["bg_dark"], fg=C["warning"],
                 font=("Segoe UI",14)).pack(side="left", padx=4)
        tf_txt = tk.Frame(td_frame, bg=C["bg_dark"]); tf_txt.pack(side="left", fill="x", expand=True)
        tk.Label(tf_txt, text="Confirmar como trabalho direto",
                 bg=C["bg_dark"], fg=C["text_hi"], font=("Segoe UI",10,"bold")).pack(anchor="w")
        tk.Label(tf_txt, text="Não desconta almoço. Calcula horas corridas.",
                 bg=C["bg_dark"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w")
        tk.Checkbutton(td_frame, variable=td_var, bg=C["bg_dark"],
                       activebackground=C["bg_dark"], selectcolor=C["bg_dark"]).pack(side="right")

        # ── Justificativa ─────────────────────────────────────────────────────
        Sep(body).pack(fill="x", pady=12)
        tk.Label(body, text="JUSTIFICATIVA (obrigatório)", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0],9,"bold")).pack(anchor="w", pady=(0,4))
        just_ent = tk.Entry(body, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                             insertbackground=C["accent"], relief="flat",
                             highlightbackground=C["border"], highlightthickness=1)
        just_ent.pack(fill="x", pady=2)
        tk.Label(body, text="Ex: 'Batida de saída esquecida — confirmado com funcionário'",
                 bg=C["bg_panel"], fg=C["text_lo"], font=FONT_SMALL).pack(anchor="w")

        # ── Ações ─────────────────────────────────────────────────────────────
        Sep(body).pack(fill="x", pady=12)
        # Preview do cálculo
        self._preview_label = tk.Label(body, text="", bg=C["bg_panel"],
                                        fg=C["text_mid"], font=FONT_SMALL, justify="left")
        self._preview_label.pack(anchor="w", pady=4)

        def _preview():
            novos = [e.get().strip() for e in entries if e.get().strip()]
            novos = sorted([h for h in novos if re.match(r'^\d{1,2}:\d{2}$', h)])
            if novos:
                jorn = bat["jornada_h"] or 8.0
                calc = self.engine.calc_dia(bat["data"], novos, jorn)
                td   = td_var.get()
                trab = calc["trabalhados_min"]
                if td and len(novos) == 2:
                    pass  # já correto para 2 batidas
                saldo_cor = "green" if calc["saldo_min"] >= 0 else "red"
                self._preview_label.config(
                    text=f"Preview: {','.join(novos)}\n"
                         f"Trabalhado: {calc['trabalhados_fmt']}  |  "
                         f"Extras: {calc['extras_fmt']}  |  Saldo: {calc['saldo_fmt']}",
                    fg=C["success"] if calc["saldo_min"] >= 0 else C["warning"])
            else:
                self._preview_label.config(text="Preencha os horários para ver o preview.")

        Btn(body, "👁  Preview do cálculo", _preview, style="outline").pack(anchor="w", pady=4)

        def _salvar():
            # Coleta e valida
            novos = [e.get().strip() for e in entries if e.get().strip()]
            novos = [h for h in novos if re.match(r'^\d{1,2}:\d{2}$', h)]
            novos = sorted(novos)
            just  = just_ent.get().strip()
            if not just:
                messagebox.showerror("Justificativa obrigatória",
                    "Digite uma justificativa antes de salvar.", parent=win); return
            ant_str = bat["horarios_corrigidos"] or bat["horarios"]
            nov_str = ",".join(novos)
            # Salva horários corrigidos na batida
            self.db.conn.execute(
                "UPDATE batidas SET horarios_corrigidos=? WHERE id=?",
                (nov_str, bat["id"]))
            self.db.conn.commit()
            # Flags
            self._set_flag(bat["id"], corrigido=1,
                           trabalho_direto=1 if td_var.get() else 0)
            # Auditoria
            self._registrar_auditoria(bat["id"], "correcao_manual",
                                       ant_str, nov_str, just)
            messagebox.showinfo("Salvo", "Registro corrigido e auditoria registrada!", parent=win)
            win.destroy()
            self._scan()

        def _remover_batida():
            """Remove a batida inteira (marca como ignorada)."""
            just = just_ent.get().strip()
            if not just:
                messagebox.showerror("Justificativa obrigatória",
                    "Digite uma justificativa antes de remover.", parent=win); return
            ok = messagebox.askyesno("Confirmar remoção",
                "Remover todas as batidas deste dia?\n"
                "A batida ficará com 0 horários e será ignorada na validação.", parent=win)
            if not ok: return
            ant_str = bat["horarios_corrigidos"] or bat["horarios"]
            self.db.conn.execute(
                "UPDATE batidas SET horarios_corrigidos='' WHERE id=?", (bat["id"],))
            self.db.conn.commit()
            self._set_flag(bat["id"], corrigido=1, ignorado=1)
            self._registrar_auditoria(bat["id"], "remocao_manual", ant_str, "", just)
            messagebox.showinfo("Removido", "Batidas removidas e auditoria registrada.", parent=win)
            win.destroy()
            self._scan()

        bf = tk.Frame(body, bg=C["bg_panel"]); bf.pack(fill="x", pady=8)
        Btn(bf, "✓  Salvar Correção",    _salvar,         style="success").pack(side="left")
        Btn(bf, "🗑  Remover Batida",     _remover_batida, style="danger").pack(side="left", padx=8)
        Btn(bf, "Cancelar",              win.destroy,     style="outline").pack(side="right")

    # ── Trabalho direto rápido ────────────────────────────────────────────────
    def _marcar_trabalho_direto(self, prob):
        bat  = prob["batida"]
        nome = bat["func_nome"] or bat["funcionario_raw"]
        try:
            dt = datetime.strptime(bat["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            dt = bat["data"]
        ok = messagebox.askyesno("Trabalho Direto",
            f"Confirmar trabalho direto para {nome} em {dt}?\n\n"
            "O almoço NÃO será descontado. As horas serão calculadas de forma contínua.")
        if not ok: return
        self._set_flag(bat["id"], trabalho_direto=1, corrigido=1)
        self._registrar_auditoria(bat["id"], "trabalho_direto",
                                   bat["horarios"], bat["horarios_corrigidos"] or bat["horarios"],
                                   "Confirmado como trabalho direto pelo administrador")
        self._scan()

    def _ignorar(self, prob):
        bat  = prob["batida"]
        nome = bat["func_nome"] or bat["funcionario_raw"]
        try:
            dt = datetime.strptime(bat["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            dt = bat["data"]
        ok = messagebox.askyesno("Ignorar alerta",
            f"Ignorar o alerta '{prob['alerta']['descricao']}'\npara {nome} em {dt}?\n\n"
            "O registro não aparecerá mais na lista de problemas.")
        if not ok: return
        self._set_flag(bat["id"], ignorado=1)
        self._registrar_auditoria(bat["id"], "ignorado",
                                   bat["horarios"], "", "Alerta ignorado pelo administrador")
        self._scan()

    # ── Tela de auditoria ─────────────────────────────────────────────────────
    def _abrir_auditoria(self):
        win = tk.Toplevel(self); win.title("Histórico de Auditoria")
        win.configure(bg=C["bg_panel"]); win.geometry("820x520")
        win.resizable(True, True)

        tk.Frame(win, bg=C["accent"], height=4).pack(fill="x")
        hf = tk.Frame(win, bg=C["bg_panel"], padx=24, pady=14); hf.pack(fill="x")
        tk.Label(hf, text="📋  Histórico de Auditoria", bg=C["bg_panel"],
                 fg=C["text_hi"], font=("Segoe UI",13,"bold")).pack(side="left")
        Btn(hf, "✕  Fechar", win.destroy, style="outline").pack(side="right")
        Sep(win).pack(fill="x", padx=24)

        # Treeview
        tf = tk.Frame(win, bg=C["bg_panel"], padx=24, pady=12); tf.pack(fill="both", expand=True)
        _tv_style()
        cols = ("data_hr","admin","func","data_ponto","acao","horarios_ant","horarios_nov","justificativa")
        tree = ttk.Treeview(tf, columns=cols, show="headings",
                             style="Dark.Treeview", selectmode="browse")
        hdrs = [("data_hr","DATA/HORA",130),("admin","ADMIN",70),
                ("func","FUNCIONÁRIO",110),("data_ponto","DIA",90),
                ("acao","AÇÃO",120),("horarios_ant","ANTES",130),
                ("horarios_nov","DEPOIS",130),("justificativa","JUSTIFICATIVA",200)]
        for col, lbl, w in hdrs:
            tree.heading(col, text=lbl)
            tree.column(col, width=w, anchor="w")
        tree.tag_configure("correcao",      foreground=C["accent"])
        tree.tag_configure("trabalho_dir",  foreground=C["warning"])
        tree.tag_configure("ignorado",      foreground=C["text_lo"])
        tree.tag_configure("remocao",       foreground=C["danger"])

        vsb = ttk.Scrollbar(tf, orient="vertical",  command=tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        # Carrega dados
        rows = self.db.conn.execute("""
            SELECT a.*, b.data as data_ponto,
                   COALESCE(f.nome, b.funcionario_raw) as func_nome
            FROM ponto_auditoria a
            JOIN batidas b ON b.id = a.batida_id
            LEFT JOIN funcionarios f ON f.id = b.funcionario_id
            ORDER BY a.criado_em DESC
            LIMIT 500
        """).fetchall()

        for r in rows:
            try:
                dt_fmt = datetime.strptime(r["data_ponto"], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                dt_fmt = r["data_ponto"]
            acao = r["acao"] or ""
            tag  = ("correcao"     if "correcao" in acao else
                    "trabalho_dir" if "direto"   in acao else
                    "remocao"      if "remocao"  in acao else
                    "ignorado")
            tree.insert("", "end", values=(
                (r["criado_em"] or "")[:16], r["admin"] or "admin",
                r["func_nome"] or "", dt_fmt,
                acao, r["horarios_ant"] or "", r["horarios_nov"] or "",
                r["justificativa"] or ""),
                tags=(tag,))

        if not rows:
            tk.Label(tf, text="Nenhum registro de auditoria ainda.",
                     bg=C["bg_panel"], fg=C["text_lo"], font=FONT_BODY).pack(pady=20)


# ──────────────────────────────────────────────────────────────────────────────
# APLICATIVO PRINCIPAL
# ──────────────────────────────────────────────────────────────────────────────

class PageFolhaPagamento(tk.Frame):
    """Página de Folha de Pagamento de Horas Extras."""
    def __init__(self, parent, db, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db
        self._dados = []
        self._sel_id = None
        self._build()
        self._load()

    def _build(self):
        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=26); hdr.pack(fill="x")
        lh  = tk.Frame(hdr, bg=C["bg_panel"]); lh.pack(side="left", fill="x", expand=True)
        tk.Label(lh, text="Folha de Pagamento", bg=C["bg_panel"], fg=C["text_hi"],
                 font=FONT_TITLE).pack(anchor="w")
        tk.Label(lh, text="Horas extras acumuladas · Valor a pagar por funcionário",
                 bg=C["bg_panel"], fg=C["text_mid"], font=FONT_BODY).pack(anchor="w")
        Btn(hdr, "↻  Atualizar", self._load, style="outline").pack(side="right", pady=4)
        Sep(self).pack(fill="x", padx=36)

        # ── Resumo total (topo) ───────────────────────────────────────────────
        res = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=16); res.pack(fill="x")
        self._total_card = tk.Frame(res, bg=C["bg_card"],
                                     highlightbackground=C["border"], highlightthickness=1)
        self._total_card.pack(fill="x")
        tk.Frame(self._total_card, bg=C["accent"], height=3).pack(fill="x")
        ti = tk.Frame(self._total_card, bg=C["bg_card"], padx=24, pady=16); ti.pack(fill="x")
        row_t = tk.Frame(ti, bg=C["bg_card"]); row_t.pack(fill="x")
        tk.Label(row_t, text="TOTAL A PAGAR (todos os funcionários)", bg=C["bg_card"],
                 fg=C["text_mid"], font=(FONT_SMALL[0], 9, "bold")).pack(side="left")
        self._lbl_total = tk.Label(row_t, text="R$ 0,00", bg=C["bg_card"],
                                    fg=C["accent"], font=("Segoe UI", 22, "bold"))
        self._lbl_total.pack(side="right")
        self._lbl_resumo = tk.Label(ti, text="", bg=C["bg_card"],
                                     fg=C["text_lo"], font=FONT_SMALL)
        self._lbl_resumo.pack(anchor="w", pady=2)

        # ── Corpo: tabela + painel lateral ────────────────────────────────────
        body = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=4); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=3); body.columnconfigure(1, weight=2)

        # Tabela esquerda
        lw = tk.Frame(body, bg=C["bg_panel"]); lw.grid(row=0, column=0, sticky="nsew", padx=(0,16))
        tk.Label(lw, text="FUNCIONÁRIOS", bg=C["bg_panel"], fg=C["text_mid"],
                 font=(FONT_SMALL[0], 9, "bold")).pack(anchor="w", pady=8)
        tw = tk.Frame(lw, bg=C["bg_panel"]); tw.pack(fill="both", expand=True)
        _tv_style()
        cols = ("nome","horas","valor_hora","valor_devido","pago","a_pagar")
        self.tree = ttk.Treeview(tw, columns=cols, show="headings",
                                  style="Dark.Treeview", selectmode="browse")
        hdrs = [("nome","FUNCIONÁRIO",160),("horas","SALDO HRS",90),
                ("valor_hora","R$/HORA",80),("valor_devido","TOTAL DEVIDO",110),
                ("pago","JÁ PAGO",100),("a_pagar","A PAGAR",100)]
        for col, lbl, w in hdrs:
            self.tree.heading(col, text=lbl)
            anchor = "w" if col == "nome" else "center"
            self.tree.column(col, width=w, anchor=anchor)
        self.tree.tag_configure("positivo", foreground=C["success"])
        self.tree.tag_configure("negativo", foreground=C["danger"])
        self.tree.tag_configure("zero",     foreground=C["text_mid"])
        self.tree.tag_configure("pago",     foreground=C["text_lo"])
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sel)

        # Painel direito
        rw = tk.Frame(body, bg=C["bg_panel"]); rw.grid(row=0, column=1, sticky="nsew")

        # Card saldo do funcionário selecionado
        tk.Label(rw, text="FUNCIONÁRIO SELECIONADO", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0], 9, "bold")).pack(anchor="w", pady=6)
        self._det_card = tk.Frame(rw, bg=C["bg_card"],
                                   highlightbackground=C["border"], highlightthickness=1)
        self._det_card.pack(fill="x")
        self._det_bar = tk.Frame(self._det_card, bg=C["accent"], height=3); self._det_bar.pack(fill="x")
        di = tk.Frame(self._det_card, bg=C["bg_card"], padx=20, pady=16); di.pack(fill="x")
        self._lbl_det_nome  = tk.Label(di, text="Selecione um funcionário", bg=C["bg_card"],
                                        fg=C["text_mid"], font=("Segoe UI", 12))
        self._lbl_det_nome.pack(anchor="w")
        self._lbl_det_valor = tk.Label(di, text="R$ --,--", bg=C["bg_card"],
                                        fg=C["accent"], font=("Segoe UI", 32, "bold"))
        self._lbl_det_valor.pack(anchor="w")
        self._lbl_det_sub   = tk.Label(di, text="", bg=C["bg_card"],
                                        fg=C["text_lo"], font=FONT_SMALL)
        self._lbl_det_sub.pack(anchor="w")

        # Formulário de pagamento
        Sep(rw).pack(fill="x", pady=10)
        tk.Label(rw, text="REGISTRAR PAGAMENTO", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0], 9, "bold")).pack(anchor="w", pady=4)
        pag_card = tk.Frame(rw, bg=C["bg_card"],
                             highlightbackground=C["border"], highlightthickness=1)
        pag_card.pack(fill="x")
        tk.Frame(pag_card, bg=C["success"], height=3).pack(fill="x")
        pi = tk.Frame(pag_card, bg=C["bg_card"], padx=20, pady=16); pi.pack(fill="x")

        # Tipo de pagamento
        tp = tk.Frame(pi, bg=C["bg_card"]); tp.pack(fill="x", pady=4)
        tk.Label(tp, text="Tipo:", bg=C["bg_card"], fg=C["text_mid"],
                 font=FONT_BODY, width=9, anchor="w").pack(side="left")
        self._v_tipo = tk.StringVar(value="total")
        for txt, val in [("Pagamento total", "total"), ("Pagamento parcial", "parcial")]:
            tk.Radiobutton(tp, text=txt, variable=self._v_tipo, value=val,
                           bg=C["bg_card"], fg=C["text_hi"], selectcolor=C["bg_card"],
                           activebackground=C["bg_card"], activeforeground=C["accent"],
                           font=FONT_BODY, command=self._toggle_parcial).pack(side="left", padx=8)

        # Valor parcial (só aparece se parcial)
        self._fr_parcial = tk.Frame(pi, bg=C["bg_card"]); self._fr_parcial.pack(fill="x", pady=4)
        self._fr_parcial.pack_forget()
        vp = tk.Frame(self._fr_parcial, bg=C["bg_card"]); vp.pack(fill="x")
        tk.Label(vp, text="Valor (R$):", bg=C["bg_card"], fg=C["text_mid"],
                 font=FONT_BODY, width=9, anchor="w").pack(side="left")
        self.ent_valor = tk.Entry(vp, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                   insertbackground=C["accent"], width=12,
                                   highlightbackground=C["border"], highlightthickness=1)
        self.ent_valor.pack(side="left", padx=8)

        # Motivo
        mr = tk.Frame(pi, bg=C["bg_card"]); mr.pack(fill="x", pady=8)
        tk.Label(mr, text="Motivo:", bg=C["bg_card"], fg=C["text_mid"],
                 font=FONT_BODY, width=9, anchor="w").pack(side="left")
        self.ent_motivo = tk.Entry(mr, bg=C["bg_input"], fg=C["text_hi"], font=FONT_BODY,
                                    insertbackground=C["accent"],
                                    highlightbackground=C["border"], highlightthickness=1)
        self.ent_motivo.pack(side="left", fill="x", expand=True)

        Btn(pi, "✓  Registrar Pagamento", self._registrar_pagamento, style="success").pack(anchor="w", pady=4)

        # Histórico de pagamentos
        Sep(rw).pack(fill="x", pady=10)
        tk.Label(rw, text="HISTÓRICO DE PAGAMENTOS", bg=C["bg_panel"],
                 fg=C["text_mid"], font=(FONT_SMALL[0], 9, "bold")).pack(anchor="w", pady=4)
        self._hist_frame = tk.Frame(rw, bg=C["bg_card"],
                                     highlightbackground=C["border"], highlightthickness=1)
        self._hist_frame.pack(fill="both", expand=True)
        tk.Frame(self._hist_frame, bg=C["success"], height=3).pack(fill="x")
        self._hist_inner = tk.Frame(self._hist_frame, bg=C["bg_card"], padx=16, pady=12)
        self._hist_inner.pack(fill="both", expand=True)
        tk.Label(self._hist_inner, text="Selecione um funcionário",
                 bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI", 10)).pack(expand=True)

    def _toggle_parcial(self):
        if self._v_tipo.get() == "parcial":
            self._fr_parcial.pack(fill="x", pady=4)
        else:
            self._fr_parcial.pack_forget()

    def _fmt_brl(self, valor):
        s = f"R$ {valor:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        return s

    def _load(self):
        self._dados = self.db.get_folha_pagamento()
        for r in self.tree.get_children(): self.tree.delete(r)
        total_a_pagar = 0.0
        funcs_com_saldo = 0
        for d in self._dados:
            saldo_h  = d["saldo_min"] // 60
            saldo_m  = d["saldo_min"] % 60
            horas_txt = f"{saldo_h}h {saldo_m:02d}m"
            vh_txt    = self._fmt_brl(d["valor_hora"])
            vd_txt    = self._fmt_brl(d["valor_devido"])
            pg_txt    = self._fmt_brl(d["total_pago"])
            ap_txt    = self._fmt_brl(d["a_pagar"])
            if d["saldo_min"] <= 0:
                tag = "zero"
            elif d["a_pagar"] <= 0:
                tag = "pago"
            else:
                tag = "positivo"
                total_a_pagar += d["a_pagar"]
                funcs_com_saldo += 1
            self.tree.insert("", "end", iid=str(d["id"]),
                             values=(d["nome"], horas_txt, vh_txt, vd_txt, pg_txt, ap_txt),
                             tags=(tag,))
        # Atualiza card de total
        self._lbl_total.config(text=self._fmt_brl(total_a_pagar))
        self._lbl_resumo.config(text=f"{funcs_com_saldo} funcionário(s) com horas a receber"
                                      f" · {len(self._dados)} total cadastrados")

    def _on_sel(self, e=None):
        sel = self.tree.selection()
        if not sel: return
        fid  = int(sel[0])
        self._sel_id = fid
        d    = next((x for x in self._dados if x["id"] == fid), None)
        if not d: return
        cor  = C["success"] if d["a_pagar"] > 0 else C["text_mid"]
        self._det_bar.config(bg=cor)
        self._lbl_det_nome.config(text=d["nome"], fg=C["text_hi"])
        self._lbl_det_valor.config(text=self._fmt_brl(d["a_pagar"]), fg=cor)
        h = d["saldo_min"] // 60; m = d["saldo_min"] % 60
        tipo_str = "Semanal" if d["tipo_pag"] == "semana" else "Mensal"
        self._lbl_det_sub.config(
            text=f"{h}h {m:02d}min extras · {self._fmt_brl(d['valor_hora'])}/h"
                 f" · Salário {tipo_str} {self._fmt_brl(d['salario'])}")
        # Preenche valor parcial com o valor a pagar
        self.ent_valor.delete(0,"end")
        self.ent_valor.insert(0, f"{d['a_pagar']:.2f}".replace(".", ","))
        # Histórico
        hists = self.db.get_historico_pagamentos(fid)
        self._render_historico(hists)

    def _render_historico(self, hists):
        for w in self._hist_inner.winfo_children(): w.destroy()
        if not hists:
            tk.Label(self._hist_inner, text="Nenhum pagamento registrado ainda.",
                     bg=C["bg_card"], fg=C["text_lo"], font=("Segoe UI",10)).pack(expand=True)
            return
        canvas = tk.Canvas(self._hist_inner, bg=C["bg_card"], highlightthickness=0, height=120)
        sb = ttk.Scrollbar(self._hist_inner, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=C["bg_card"])
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        for h in hists:
            tipo_h = h["tipo"]
            cor    = C["success"]
            rw = tk.Frame(inner, bg=C["bg_card"], pady=5); rw.pack(fill="x")
            tk.Frame(rw, bg=cor, width=3).pack(side="left", fill="y", padx=8)
            lb = tk.Frame(rw, bg=C["bg_card"]); lb.pack(side="left", fill="x", expand=True)
            tipo_label = "Pagamento total" if tipo_h == "total" else "Pagamento parcial"
            desc = h["descricao"] or tipo_label
            tk.Label(lb, text=desc, bg=C["bg_card"], fg=C["text_hi"],
                     font=FONT_SMALL, anchor="w").pack(anchor="w")
            dt = (h["criado_em"] or "")[:16]
            tk.Label(lb, text=f"{dt} · {tipo_label}", bg=C["bg_card"],
                     fg=C["text_lo"], font=("Segoe UI",8), anchor="w").pack(anchor="w")
            val_txt = self._fmt_brl(h["valor_pago"])
            tk.Label(rw, text=val_txt, bg=C["bg_card"], fg=cor,
                     font=("Segoe UI",12,"bold")).pack(side="right", padx=6)
            Sep(inner).pack(fill="x")

    def _registrar_pagamento(self):
        if not self._sel_id:
            messagebox.showwarning("Atenção", "Selecione um funcionário primeiro."); return
        d = next((x for x in self._dados if x["id"] == self._sel_id), None)
        if not d: return
        tipo = self._v_tipo.get()
        if tipo == "total":
            valor = d["a_pagar"]
            if valor <= 0:
                messagebox.showinfo("Sem pendência",
                    f"{d['nome']} não tem valor pendente a receber."); return
        else:
            raw = self.ent_valor.get().strip().replace(",",".")
            try:
                valor = float(raw)
            except ValueError:
                messagebox.showerror("Erro", "Digite um valor válido (ex: 150,00)"); return
            if valor <= 0:
                messagebox.showerror("Erro", "O valor deve ser maior que zero."); return
        motivo = self.ent_motivo.get().strip() or ("Pagamento total" if tipo == "total" else "Pagamento parcial")
        tipo_label = "total" if tipo == "total" else "parcial"
        msg_pag = (
            "Registrar pagamento de " + self._fmt_brl(valor) + " para " + d["nome"] + "?\n"
            + "Tipo: " + tipo_label.upper() + "\nMotivo: " + motivo
        )
        ok = messagebox.askyesno("Confirmar pagamento", msg_pag)
        if not ok: return
        # Calcula minutos correspondentes
        min_pagos = int((valor / d["valor_hora"]) * 60) if d["valor_hora"] > 0 else 0
        self.db.registrar_pagamento_horas(self._sel_id, min_pagos, valor, tipo_label, motivo)
        messagebox.showinfo("Sucesso",
            f"Pagamento de {self._fmt_brl(valor)} registrado para {d['nome']}!")
        self.ent_motivo.delete(0, "end")
        self._load()
        try: self.tree.selection_set(str(self._sel_id)); self._on_sel()
        except Exception: pass


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA — NOTIFICAÇÕES (Etapa 4.6)
# ──────────────────────────────────────────────────────────────────────────────
class PageNotificacoes(tk.Frame):
    """
    Central de Notificações como página completa do app.
    Inclui lista filtrada, estatísticas e ações rápidas.
    """

    def __init__(self, parent, db, app_ref=None, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self.db = db
        self.app_ref = app_ref
        self._filtro = tk.StringVar(value="todas")
        self._build()

    def _build(self):
        if not _NOTIF_OK or not self.db.notif:
            self._sem_suporte()
            return

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=26)
        hdr.pack(fill="x")

        left = tk.Frame(hdr, bg=C["bg_panel"])
        left.pack(side="left", fill="y")
        tk.Label(left, text="🔔  Notificações", bg=C["bg_panel"],
                 fg=C["text_hi"], font=("Segoe UI", 22, "bold")).pack(anchor="w")
        tk.Label(left, text="Central de avisos automáticos do sistema",
                 bg=C["bg_panel"], fg=C["text_mid"],
                 font=("Segoe UI", 11)).pack(anchor="w", pady=2)

        # Botões de ação
        right = tk.Frame(hdr, bg=C["bg_panel"])
        right.pack(side="right", fill="y")
        Btn(right, "Marcar todas como lidas", self._marcar_todas,
            style="outline", icon="✓").pack(side="right", padx=6)
        Btn(right, "Limpar antigas", self._limpar_antigas,
            style="outline", icon="🗑").pack(side="right", padx=0)

        Sep(self).pack(fill="x", padx=36)

        # ── Cards de estatísticas ─────────────────────────────────────────────
        stats_frame = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=20)
        stats_frame.pack(fill="x")
        self._stat_cards = {}

        notifs_todas = self.db.notif.db.listar(limite=500)
        nao_lidas    = sum(1 for n in notifs_todas if not n["lida"])
        extras = sum(1 for n in notifs_todas if n["tipo"] == TipoNotif.HORAS_EXTRAS)
        pagamentos = sum(1 for n in notifs_todas if n["tipo"] == TipoNotif.PAGAMENTO)
        negativas = sum(1 for n in notifs_todas if n["tipo"] == TipoNotif.HORAS_NEGATIVAS)

        for i, (t, v, c) in enumerate([
            ("Não Lidas",       nao_lidas,   C["danger"]),
            ("Horas Extras",    extras,      C["success"]),
            ("Pagamentos",      pagamentos,  C["purple"]),
            ("Horas Negativas", negativas,   C["warning"]),
        ]):
            Card(stats_frame, t, v, color=c).grid(
                row=0, column=i, padx=(0 if i else 0, 12), sticky="ew")
            stats_frame.columnconfigure(i, weight=1)

        Sep(self).pack(fill="x", padx=36)

        # ── Filtros ───────────────────────────────────────────────────────────
        filtro_frame = tk.Frame(self, bg=C["bg_panel"], padx=36, pady=14)
        filtro_frame.pack(fill="x")
        tk.Label(filtro_frame, text="Filtrar:", bg=C["bg_panel"],
                 fg=C["text_mid"], font=FONT_SMALL).pack(side="left", padx=(0, 10))
        for val, txt in [
            ("todas",            "📋 Todas"),
            ("nao_lidas",        "🔴 Não lidas"),
            (TipoNotif.HORAS_EXTRAS,     "🟢 Horas extras"),
            (TipoNotif.HORAS_NEGATIVAS,  "⚠ Horas negativas"),
            (TipoNotif.PAGAMENTO,        "💵 Pagamentos"),
            (TipoNotif.FECHAMENTO,       "📊 Fechamentos"),
        ]:
            rb = tk.Radiobutton(
                filtro_frame, text=txt, variable=self._filtro, value=val,
                bg=C["bg_panel"], fg=C["text_mid"],
                activebackground=C["bg_panel"], activeforeground=C["accent"],
                selectcolor=C["bg_card"], font=("Segoe UI", 9),
                command=self._carregar)
            rb.pack(side="left", padx=6)

        # ── Lista scrollável ──────────────────────────────────────────────────
        container = tk.Frame(self, bg=C["bg_panel"])
        container.pack(fill="both", expand=True, padx=36, pady=(0, 20))

        self._canvas = tk.Canvas(container, bg=C["bg_panel"],
                                 highlightthickness=0, bd=0)
        sb = ttk.Scrollbar(container, orient="vertical",
                           command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._lista_frame = tk.Frame(self._canvas, bg=C["bg_panel"])
        self._win_id = self._canvas.create_window(
            (0, 0), window=self._lista_frame, anchor="nw")
        self._lista_frame.bind("<Configure>", lambda e: self._canvas.configure(
            scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>", lambda e: self._canvas.itemconfig(
            self._win_id, width=e.width))
        self._canvas.bind_all("<MouseWheel>", lambda e: self._canvas.yview_scroll(
            -1 * (e.delta // 120), "units"))

        # Status bar
        self._status = tk.Label(self, text="", bg=C["bg_dark"],
                                fg=C["text_lo"], font=FONT_SMALL, pady=6)
        self._status.pack(fill="x", side="bottom")

        self._carregar()

    def _carregar(self):
        for w in self._lista_frame.winfo_children():
            w.destroy()

        filtro = self._filtro.get()
        apenas_nao_lidas = (filtro == "nao_lidas")
        notifs = self.db.notif.db.listar(apenas_nao_lidas=apenas_nao_lidas, limite=200)

        if filtro not in ("todas", "nao_lidas"):
            notifs = [n for n in notifs if n["tipo"] == filtro]

        self._status.config(
            text=f"  {len(notifs)} notificações  ·  "
                 f"{sum(1 for n in notifs if not n['lida'])} não lidas  ·  "
                 f"Última: {notifs[0]['criado_em'][:16] if notifs else '—'}")

        if not notifs:
            tk.Label(self._lista_frame, text="🔕\n\nNenhuma notificação encontrada",
                     bg=C["bg_panel"], fg=C["text_lo"],
                     font=("Segoe UI", 14), justify="center",
                     pady=60).pack(fill="x")
            return

        hoje = datetime.now().strftime("%d/%m/%Y")
        ontem = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")

        grupos: dict = {}
        for n in notifs:
            try:
                dia = datetime.strptime(n["criado_em"], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
            except Exception:
                dia = "—"
            grupos.setdefault(dia, []).append(n)

        label_dia = {hoje: "Hoje", ontem: "Ontem"}

        for dia, items in grupos.items():
            # Separador de dia
            dh = tk.Frame(self._lista_frame, bg=C["bg_panel"],
                          padx=0, pady=8)
            dh.pack(fill="x")
            tk.Label(dh, text=label_dia.get(dia, dia),
                     bg=C["bg_panel"], fg=C["text_lo"],
                     font=("Segoe UI", 9, "bold")).pack(side="left")
            tk.Frame(dh, bg=C["border"], height=1).pack(
                side="left", fill="x", expand=True, padx=8)

            for n in items:
                self._card(n)

    def _card(self, n):
        tipo  = n["tipo"]
        lida  = bool(n["lida"])
        cor   = TipoNotif.CORES.get(tipo, C["accent"])
        icone = TipoNotif.ICONES.get(tipo, "🔔")
        nav   = n["navegacao"] or ""
        bg    = C["bg_card"] if not lida else C["bg_panel"]

        outer = tk.Frame(self._lista_frame, bg=C["border"], pady=1)
        outer.pack(fill="x", pady=1)

        card = tk.Frame(outer, bg=bg, padx=0, pady=10, cursor="hand2")
        card.pack(fill="x")

        barra = tk.Frame(card, bg=cor if not lida else C["border"], width=4)
        barra.pack(side="left", fill="y", padx=(0, 16))

        corpo = tk.Frame(card, bg=bg)
        corpo.pack(side="left", fill="both", expand=True)

        topo = tk.Frame(corpo, bg=bg)
        topo.pack(fill="x")

        tk.Label(topo, text=icone, bg=bg, fg=cor,
                 font=("Segoe UI", 13)).pack(side="left", padx=(0, 8))
        tk.Label(topo, text=n["titulo"], bg=bg,
                 fg=C["text_hi"] if not lida else C["text_mid"],
                 font=("Segoe UI", 11, "bold" if not lida else "normal"),
                 anchor="w").pack(side="left", fill="x", expand=True)

        try:
            ts = datetime.strptime(n["criado_em"], "%Y-%m-%d %H:%M:%S")
            hora_txt = ts.strftime("%H:%M")
        except Exception:
            hora_txt = ""
        tk.Label(topo, text=hora_txt, bg=bg, fg=C["text_lo"],
                 font=FONT_SMALL).pack(side="right", padx=(0, 16))

        tk.Label(corpo, text=n["mensagem"], bg=bg,
                 fg=C["text_mid"], font=FONT_SMALL,
                 anchor="w", justify="left",
                 wraplength=700).pack(fill="x", pady=(4, 0))

        if nav:
            nav_lbl = tk.Label(corpo, text="Ver detalhes →",
                               bg=bg, fg=cor, font=("Segoe UI", 9),
                               cursor="hand2")
            nav_lbl.pack(anchor="w", pady=(4, 0))
            nav_lbl.bind("<Button-1>",
                         lambda e, nid=n["id"], d=nav: self._ir(nid, d))

        if not lida:
            tk.Label(card, text="●", bg=bg, fg=cor,
                     font=("Segoe UI", 8)).pack(
                         side="right", padx=(0, 16))

        def _enter(e, f=card, b=bg):
            f.config(bg=C["hover"])
            for c in f.winfo_children(): _rc(c, C["hover"])
        def _leave(e, f=card, b=bg):
            f.config(bg=b)
            for c in f.winfo_children(): _rc(c, b)
        def _rc(w, bg):
            try:
                w.config(bg=bg)
                for c in w.winfo_children(): _rc(c, bg)
            except Exception: pass

        card.bind("<Enter>", _enter)
        card.bind("<Leave>", _leave)
        card.bind("<Button-1>",
                  lambda e, nid=n["id"], d=nav: self._ir(nid, d))

    def _ir(self, notif_id: int, destino: str):
        self.db.notif.db.marcar_lida(notif_id)
        if self.app_ref and destino:
            self.app_ref._navegar_notif(destino)
        else:
            self._carregar()

    def _marcar_todas(self):
        self.db.notif.db.marcar_todas_lidas()
        self.db.notif._notificar_callbacks()
        self._carregar()

    def _limpar_antigas(self):
        self.db.notif.db.deletar_antigas(7)
        self._carregar()

    def _sem_suporte(self):
        tk.Label(self, text="🔔\n\nMódulo de notificações não disponível.\n"
                 "Verifique se o arquivo notificacoes.py está na mesma pasta.",
                 bg=C["bg_panel"], fg=C["text_lo"],
                 font=("Segoe UI", 14), justify="center",
                 pady=80).pack(expand=True)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} {APP_VERSION}")
        self.configure(bg=C["bg_dark"])
        self.minsize(1100, 680)
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h   = min(1400, sw-60), min(840, sh-60)
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.db = Database(); self.current_page = None
        self._build_layout(); self._show("dashboard")
        # ── Registra toast interno (Etapa 4.6) ───────────────────────────────
        if _NOTIF_OK and self.db.notif:
            self.db.notif.registrar_callback(self._on_nova_notif)
            # Limpa notificações lidas com mais de 30 dias na inicialização
            self.db.notif.db.deletar_antigas(30)

    def _on_nova_notif(self):
        """Chamado quando nova notificação é emitida — mostra toast interno."""
        try:
            notifs = self.db.notif.db.listar(apenas_nao_lidas=True, limite=1)
            if notifs:
                n = notifs[0]
                ToastInterno.mostrar(
                    root=self,
                    titulo=n["titulo"],
                    mensagem=n["mensagem"],
                    tipo=n["tipo"],
                    duracao_ms=5500,
                    on_click=lambda dest=n["navegacao"]: self._navegar_notif(dest or "")
                )
        except Exception:
            pass

    def _build_layout(self):
        self.sidebar = tk.Frame(self, bg=C["sidebar"], width=232)
        self.sidebar.pack(side="left", fill="y"); self.sidebar.pack_propagate(False)

        brand = tk.Frame(self.sidebar, bg=C["sidebar"], pady=26); brand.pack(fill="x")
        tk.Label(brand, text="◉",     bg=C["sidebar"], fg=C["accent"],  font=("Segoe UI",20,"bold")).pack()
        tk.Label(brand, text="PONTO", bg=C["sidebar"], fg=C["text_hi"], font=("Segoe UI",13,"bold")).pack()
        tk.Label(brand, text="Sistema de Controle", bg=C["sidebar"], fg=C["text_lo"], font=FONT_SMALL).pack(pady=0)
        Sep(self.sidebar).pack(fill="x", padx=16, pady=8)

        self.nav_btns = {}
        nf = tk.Frame(self.sidebar, bg=C["sidebar"]); nf.pack(fill="x", padx=8, pady=8)
        for key, icon, label in [
            ("dashboard",    "◫",  "Dashboard"),
            ("funcionarios", "👤", "Funcionários"),
            ("importacao",   "↑",  "Importar Ponto"),
            ("relatorios",   "≡",  "Relatórios Importados"),
            ("calculo",      "⏱",  "Cálculo de Horas"),
            ("banco_horas",    "⌛", "Banco de Horas"),
            ("folha_pagamento","💰", "Folha de Pagamento"),
            ("financeiro",     "💳", "Financeiro"),
            ("validacao",      "🔍", "Validação de Ponto"),
            ("notificacoes",   "🔔", "Notificações"),
            ("configuracoes",  "⚙",  "Configurações"),
        ]:
            self.nav_btns[key] = self._nav_btn(nf, key, icon, label)

        tk.Label(self.sidebar, text=APP_VERSION, bg=C["sidebar"],
                 fg=C["text_lo"], font=FONT_SMALL).pack(side="bottom", pady=16)

        # ── Botão de Notificações (sino) — Etapa 4.6 ─────────────────────────
        if _NOTIF_OK and self.db.notif:
            Sep(self.sidebar).pack(fill="x", padx=16, side="bottom", pady=0)
            self._sino_btn = BotaoSino(
                self.sidebar,
                gerenciador=self.db.notif,
                on_navegar=self._navegar_notif,
                bg_color=C["sidebar"])
            self._sino_btn.pack(side="bottom", pady=4)
            tk.Label(self.sidebar, text="Notificações", bg=C["sidebar"],
                     fg=C["text_lo"], font=("Segoe UI", 8)).pack(side="bottom")

        self.main = tk.Frame(self, bg=C["bg_panel"]); self.main.pack(side="right", fill="both", expand=True)

    def _nav_btn(self, parent, key, icon, label):
        frame    = tk.Frame(parent, bg=C["sidebar"], cursor="hand2"); frame.pack(fill="x", pady=2)
        icon_lbl = tk.Label(frame, text=icon,  bg=C["sidebar"], fg=C["text_mid"], font=("Segoe UI",13,"bold"), width=4)
        icon_lbl.pack(side="left", padx=4, pady=11)
        text_lbl = tk.Label(frame, text=label, bg=C["sidebar"], fg=C["text_mid"], font=("Segoe UI",11), anchor="w")
        text_lbl.pack(side="left", fill="x", expand=True)
        bar = tk.Frame(frame, bg=C["sidebar"], width=4); bar.pack(side="right", fill="y")

        def click(e=None): self._show(key)
        def enter(e=None):
            if self.current_page != key:
                for w in (frame, icon_lbl, text_lbl, bar): w.config(bg=C["hover"])
        def leave(e=None):
            if self.current_page != key:
                for w in (frame, icon_lbl, text_lbl, bar): w.config(bg=C["sidebar"])
        for w in (frame, icon_lbl, text_lbl, bar):
            w.bind("<Button-1>", click); w.bind("<Enter>", enter); w.bind("<Leave>", leave)
        return {"frame":frame,"icon":icon_lbl,"text":text_lbl,"bar":bar}

    def _navegar_notif(self, destino: str):
        """
        Navega para a página correta ao clicar em uma notificação.
        destino pode ser: 'calculo', 'financeiro:3', 'banco_horas', etc.
        """
        if not destino:
            return
        partes = destino.split(":")
        pagina = partes[0]
        paginas_validas = {
            "dashboard", "funcionarios", "importacao", "relatorios",
            "calculo", "banco_horas", "folha_pagamento", "financeiro",
            "validacao", "configuracoes"
        }
        if pagina in paginas_validas:
            self._show(pagina)

    def _show(self, key):
        if self.current_page == key: return
        if self.current_page and self.current_page in self.nav_btns:
            old = self.nav_btns[self.current_page]
            for w in ("frame","icon","text","bar"): old[w].config(bg=C["sidebar"])
            old["icon"].config(fg=C["text_mid"]); old["text"].config(fg=C["text_mid"])
        self.current_page = key
        btn = self.nav_btns[key]
        for w in ("frame","icon","text","bar"): btn[w].config(bg=C["bg_card"])
        btn["bar"].config(bg=C["accent"]); btn["icon"].config(fg=C["accent"]); btn["text"].config(fg=C["text_hi"])
        for child in self.main.winfo_children(): child.destroy()
        {
            "dashboard":    lambda: PageDashboard(self.main, self.db),
            "funcionarios": lambda: PageFuncionarios(self.main, self.db),
            "importacao":   lambda: PageImportacao(self.main, self.db, app_ref=self),
            "relatorios":   lambda: PageRelatorios(self.main, self.db),
            "calculo":      lambda: PageCalculo(self.main, self.db),
            "banco_horas":    lambda: PageBancoHoras(self.main, self.db),
            "folha_pagamento":lambda: PageFolhaPagamento(self.main, self.db),
            "financeiro":     lambda: PageFinanceiro(self.main, self.db),
            "validacao":      lambda: PageValidacao(self.main, self.db),
            "notificacoes":   lambda: PageNotificacoes(self.main, self.db, app_ref=self),
            "configuracoes":  lambda: PageConfiguracoes(self.main, self.db),
        }[key]().pack(fill="both", expand=True)


# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()
